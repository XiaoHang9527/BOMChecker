import tkinter as tk
from tkinter import filedialog, ttk, messagebox, scrolledtext, simpledialog
import threading
import pandas as pd
import fitz  # PyMuPDF
import re
import os
import sys  # 用于系统相关功能
import time  # 用于性能测量和延迟控制
import traceback  # 用于详细错误信息
import queue  # 用于线程间通信
import gc  # 用于垃圾回收控制
import json  # 用于配置文件存储和加载
import logging  # 用于日志记录
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor, as_completed  # 用于并行处理
from datetime import datetime  # 用于生成时间戳
# 引入多进程相关库
import multiprocessing as mp
from multiprocessing import Process, Queue, freeze_support
# 引入自动更新相关库
import requests
import tempfile
import shutil
import zipfile
import subprocess
import platform  # 用于获取系统信息
from packaging import version as pkg_version  # 用于版本比较

# 定义版本信息和更新相关常量
APP_VERSION = "1.8"
GITHUB_REPO = "XiaoHang9527/BOMChecker"
GITHUB_API_URL = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
UPDATE_CHECK_INTERVAL = 7  # 天

# 统计功能已移除

# 尝试导入chardet库，用于自动检测文件编码
try:
    import chardet
except ImportError:
    chardet = None
    print("提示: chardet库未安装，将使用默认编码读取文件")

# 定义下载重试次数和超时时间
DOWNLOAD_MAX_RETRIES = 3  # 最大重试次数
DOWNLOAD_TIMEOUT = 30     # 下载超时时间(秒)
DOWNLOAD_CHUNK_SIZE = 8192  # 下载块大小

try:
    from PIL import Image, ImageTk  # 用于加载图标
except ImportError:
    pass  # 如果PIL库不存在，忽略

# 尝试导入psutil库，但如果不可用则忽略
try:
    import psutil
except ImportError:
    psutil = None
    print("提示: psutil库未安装，部分性能优化功能将不可用")

# 如果是Windows系统，尝试启用内存优化
if sys.platform == 'win32':
    try:
        import ctypes
        from ctypes import wintypes

        # 正确定义Windows API常量
        PROCESS_SET_INFORMATION = 0x0200

        # Windows 10/11支持的内存优先级参数
        ProcessMemoryPriority = 0  # 正确的ProcessInformationClass值

        # 定义内存优先级结构
        class MEMORY_PRIORITY_INFORMATION(ctypes.Structure):
            _fields_ = [("MemoryPriority", wintypes.ULONG)]

        # 定义优先级级别常量
        MEMORY_PRIORITY_VERY_LOW = 1
        MEMORY_PRIORITY_LOW = 2
        MEMORY_PRIORITY_MEDIUM = 3  # 默认值
        MEMORY_PRIORITY_BELOW_NORMAL = 4
        MEMORY_PRIORITY_NORMAL = 5
        MEMORY_PRIORITY_ABOVE_NORMAL = 6  # 我们将使用这个
        MEMORY_PRIORITY_HIGH = 7
        MEMORY_PRIORITY_VERY_HIGH = 8

        # 获取当前进程句柄
        process_handle = ctypes.windll.kernel32.OpenProcess(PROCESS_SET_INFORMATION, False, os.getpid())

        # 设置为高内存优先级
        if process_handle:
            try:
                # 创建并设置内存优先级信息结构
                priority_info = MEMORY_PRIORITY_INFORMATION()
                priority_info.MemoryPriority = MEMORY_PRIORITY_ABOVE_NORMAL

                # 调用API设置内存优先级
                result = ctypes.windll.kernel32.SetProcessInformation(
                    process_handle,
                    ProcessMemoryPriority,  # 正确的ProcessInformationClass
                    ctypes.byref(priority_info),  # 正确的结构指针
                    ctypes.sizeof(priority_info)  # 正确的结构大小
                )

                if result:
                    print("已成功设置为高内存优先级")
                else:
                    error_code = ctypes.windll.kernel32.GetLastError()
                    # 错误87表示参数错误，可能是当前Windows版本不支持
                    if error_code == 87:
                        print("当前Windows版本可能不支持内存优先级设置")
                    else:
                        print(f"设置内存优先级失败，错误码: {error_code}")

                    # 尝试使用备选方案：工作集大小设置
                    try:
                        kernel32 = ctypes.windll.kernel32
                        c_size_t = ctypes.c_size_t

                        # 获取当前工作集大小
                        min_ws_size = c_size_t()
                        max_ws_size = c_size_t()

                        if kernel32.GetProcessWorkingSetSize(process_handle, ctypes.byref(min_ws_size), ctypes.byref(max_ws_size)):
                            # 增加工作集大小限制，请求更多内存
                            new_min = min_ws_size.value
                            new_max = max(max_ws_size.value * 2, 1024 * 1024 * 256)  # 256MB或当前的2倍

                            if kernel32.SetProcessWorkingSetSize(process_handle, new_min, new_max):
                                print(f"已成功增加内存工作集上限到 {new_max/(1024*1024):.1f} MB")
                            else:
                                print(f"设置工作集大小失败: {kernel32.GetLastError()}")
                    except Exception as ws_error:
                        # 这个错误不应影响程序运行
                        print(f"设置工作集大小错误: {str(ws_error)}")

                # 尝试设置进程CPU优先级（这与内存优先级不同）
                try:
                    # 定义CPU优先级常量
                    NORMAL_PRIORITY_CLASS = 0x0020
                    ABOVE_NORMAL_PRIORITY_CLASS = 0x0008000
                    HIGH_PRIORITY_CLASS = 0x0080

                    # 设置为高于正常优先级，不用最高优先级以避免影响系统稳定性
                    if kernel32.SetPriorityClass(process_handle, ABOVE_NORMAL_PRIORITY_CLASS):
                        print("已设置高优先级进程")
                    else:
                        print(f"设置进程优先级失败: {kernel32.GetLastError()}")
                except Exception as priority_error:
                    print(f"设置进程优先级错误: {str(priority_error)}")
            finally:
                ctypes.windll.kernel32.CloseHandle(process_handle)
    except Exception as e:
        print(f"内存优化初始化失败: {str(e)}")
        # 这个错误不影响核心功能，所以我们只记录而不抛出异常

# 显示系统信息，帮助排查问题
try:
    import platform

    print(f"\n====== 系统信息 ======")
    print(f"操作系统: {sys.platform} ({platform.system()} {platform.release()})")

    # 更复杂的系统位数检测方法
    def get_os_bits():
        # 方法1: 使用platform.machine()
        machine = platform.machine().lower()
        if machine in ('amd64', 'x86_64', 'ia64', 'em64t', 'x64'):
            return 64
        elif machine in ('i386', 'i686', 'x86'):
            return 32

        # 方法2: 使用环境变量 (仅在Windows上)
        if 'PROCESSOR_ARCHITECTURE' in os.environ:
            if os.environ['PROCESSOR_ARCHITECTURE'].lower() == 'amd64' or \
               os.environ.get('PROCESSOR_ARCHITEW6432', '').lower() == 'amd64':
                return 64

        # 方法3: 使用ctypes (仅在Windows上)
        if sys.platform == 'win32':
            try:
                import ctypes
                if hasattr(ctypes, 'windll') and hasattr(ctypes.windll, 'kernel32'):
                    is_wow64 = ctypes.c_bool(False)
                    if ctypes.windll.kernel32.IsWow64Process(
                        ctypes.windll.kernel32.GetCurrentProcess(),
                        ctypes.byref(is_wow64)
                    ):
                        if is_wow64.value:
                            return 64
            except:
                pass

        # 默认尝试platform.architecture()
        try:
            return 64 if platform.architecture()[0] == '64bit' else 32
        except:
            # 最后的猜测: 系统最大整数
            return 64 if sys.maxsize > 2**32 else 32

    # 检测操作系统位数
    os_bits = get_os_bits()
    # 检测Python位数
    is_64bit_python = sys.maxsize > 2**32

    print(f"系统架构: {os_bits}位 Windows")
    print(f"Python架构: {'64位' if is_64bit_python else '32位'}")

    # 如果是64位系统但使用32位Python，发出警告
    if os_bits == 64 and not is_64bit_python:
        print("注意: 您正在64位Windows上使用32位Python，这可能限制可用内存")

    print(f"Python版本: {sys.version}")
    print(f"Pandas版本: {pd.__version__}")
    try:
        import fitz
        print(f"PyMuPDF版本: {fitz.version}")
    except:
        print("PyMuPDF版本: 未知")

    if psutil:
        try:
            mem = psutil.virtual_memory()
            print(f"系统总内存: {mem.total / (1024**3):.2f} GB")
            print(f"可用内存: {mem.available / (1024**3):.2f} GB ({mem.percent}% 已使用)")
            print(f"CPU核心数: {psutil.cpu_count(logical=False) or 'N/A'} 物理核心 / {psutil.cpu_count()} 逻辑核心")
        except:
            pass
    print(f"=========================\n")
except Exception as e:
    print(f"获取系统信息时出错: {str(e)}")

def prepare_regex_patterns(ref_designators):
    """
    为位号列表准备正则表达式模式，将位号分成小块以提高性能
    支持特殊逻辑：当BOM中位号是U401，但PDF中的位号是U401A/U401B等带字母后缀时，也视为匹配
    其他情况需要精确匹配，如B3202不应匹配到FB3202
    """
    # 转换为集合以提高查找效率
    ref_set = set(ref_designators)

    # 确定合适的块大小，平衡内存和性能
    chunk_size = min(50, max(5, len(ref_set) // 100))

    # 将位号分成多个块
    ref_chunks = [list(ref_set)[i:i+chunk_size] for i in range(0, len(ref_set), chunk_size)]

    # 为每个块创建单独的正则表达式
    # 1. 用于字母后缀匹配的模式（仅用于特殊情况，如U401匹配U401A）
    suffix_patterns = []
    # 2. 用于精确匹配的模式（确保完全精确匹配，如B3201必须精确匹配，不匹配FB3201）
    exact_patterns = []

    for chunk in ref_chunks:
        # 排序使得更长的位号先匹配，避免子串匹配问题
        sorted_chunk = sorted(chunk, key=len, reverse=True)

        # 分离字母开头的位号（如U401, R2, C10等）
        suffix_candidates = []
        for ref in sorted_chunk:
            # 只有字母开头的位号才会有字母后缀匹配
            if ref and ref[0].isalpha():
                suffix_candidates.append(ref)

        # 1. 创建支持字母后缀匹配的模式（仅用于字母开头的位号）
        if suffix_candidates:
            suffix_pattern_parts = []
            for ref in suffix_candidates:
                # 使用前后边界条件确保匹配精确:
                # (?<![A-Za-z0-9]) - 负向后瞻，确保位号前面不是字母或数字
                # (?![A-Za-z0-9]) - 正向前瞻，确保后缀后面不是字母或数字
                pattern = f"(?<![A-Za-z0-9]){re.escape(ref)}[A-Za-z]+(?![A-Za-z0-9])"
                suffix_pattern_parts.append(pattern)

            if suffix_pattern_parts:
                suffix_pattern_str = '|'.join(suffix_pattern_parts)
                suffix_pattern = re.compile(suffix_pattern_str)
                suffix_patterns.append(suffix_pattern)

        # 2. 精确匹配模式 - 使用严格的边界条件
        exact_pattern_parts = []
        for ref in sorted_chunk:
            # 使用前后边界条件确保匹配精确
            pattern = f"(?<![A-Za-z0-9]){re.escape(ref)}(?![A-Za-z0-9])"
            exact_pattern_parts.append(pattern)

        if exact_pattern_parts:
            exact_pattern_str = '|'.join(exact_pattern_parts)
            exact_pattern = re.compile(exact_pattern_str)
            exact_patterns.append(exact_pattern)

    # 将两种模式都返回
    return {'suffix_patterns': suffix_patterns, 'exact_patterns': exact_patterns}

def page_task_generator(doc, ref_set, patterns):
    """生成页面处理任务"""
    for page_num in range(len(doc)):
        try:
            page = doc[page_num]
            yield (page_num + 1, page, ref_set, patterns)
        except Exception as e:
            print(f"警告: 加载页面 {page_num + 1} 失败: {str(e)}")
            continue

def _highlight_safe_ref(page, ref, highlight_color, text_instances=None):
    """
    安全地高亮页面上的位号，确保精确匹配

    Args:
        page: PDF页面对象
        ref: 要高亮的位号
        highlight_color: 高亮颜色
        text_instances: 可选的预定义文本实例列表

    Returns:
        找到并高亮的位号区域列表
    """
    highlighted_areas = []

    try:
        # 如果没有提供文本实例，就在页面上搜索
        if not text_instances:
            # 首先获取页面上所有包含ref的区域
            text_instances = page.search_for(ref)

        if not text_instances:
            return highlighted_areas

        # 对每个找到的区域进行验证
        for inst in text_instances:
            # 获取区域周围的文本
            rect = inst  # 区域坐标
            # 略微扩大搜索区域，确保能获取到完整的上下文
            expanded_rect = fitz.Rect(rect.x0 - 10, rect.y0 - 2,
                                     rect.x1 + 10, rect.y1 + 2)

            # 获取这个区域的文本
            area_text = page.get_text("text", clip=expanded_rect)

            # 使用正则表达式验证这个区域包含的是完整位号（前后有非字母数字字符或文本边界）
            # 创建一个严格的边界检查模式
            pattern = r'(?:^|[^A-Za-z0-9])' + re.escape(ref) + r'(?:[^A-Za-z0-9]|$)'

            matches = re.search(pattern, area_text)
            if matches:
                # 这是一个有效匹配，添加高亮
                annot = page.add_highlight_annot(inst)
                annot.set_colors(stroke=highlight_color)
                annot.update()
                highlighted_areas.append(inst)
    except Exception as e:
        print(f"安全高亮位号 {ref} 时出错: {str(e)}")

    return highlighted_areas

def _process_page(page_data):
    """
    处理单个PDF页面并高亮匹配的位号

    实现特殊匹配逻辑：
    1. 当BOM中位号是U401，PDF中的位号是U401A/U401B等带字母后缀时，也允许匹配
    2. 其他情况保持精确匹配，如B3202不允许匹配到FB3202
    """
    page_num, page, ref_set, patterns, highlight_color, add_annotation, bom_info = page_data

    # 获取页面文本
    try:
        text = page.get_text()

        # 使用所有分块的正则表达式进行匹配
        all_matches = set()  # 保存所有匹配到的原始位号
        suffix_matches = {}  # 存储带后缀的匹配及其原始位号

        # 使用精确匹配模式（所有位号都进行精确匹配）
        for pattern in patterns['exact_patterns']:
            try:
                # 查找所有精确匹配
                matches = pattern.findall(text)
                if matches:
                    for match in matches:
                        if match in ref_set:  # 确保匹配的是我们需要的位号
                            all_matches.add(match)
            except Exception as pattern_error:
                # 单个模式匹配失败不应影响整个处理
                print(f"页面 {page_num} 的精确模式匹配失败: {str(pattern_error)}")

        # 使用后缀匹配模式（仅用于字母开头的位号，如U401匹配U401A/U401B）
        for pattern in patterns['suffix_patterns']:
            try:
                # 找到所有带后缀的匹配（如U401A, R302B等）
                suffix_matches_found = pattern.findall(text)

                if suffix_matches_found:
                    for suffix_match in suffix_matches_found:
                        # 找到匹配的原始位号（不带后缀）
                        for ref in ref_set:
                            # 确保ref是以字母开头，且suffix_match以ref开头
                            if ref and ref[0].isalpha() and suffix_match.startswith(ref):
                                # 确保匹配与位号的差异仅为字母后缀
                                suffix = suffix_match[len(ref):]
                                if suffix and suffix.isalpha():
                                    all_matches.add(ref)  # 添加原始位号到匹配集合
                                    suffix_matches[suffix_match] = ref  # 记录后缀匹配与原始位号的关系
                                    break
            except Exception as pattern_error:
                # 单个模式匹配失败不应影响整个处理
                print(f"页面 {page_num} 的后缀模式匹配失败: {str(pattern_error)}")

        # 记录已找到的位号
        found_refs = set()

        # 高亮所有匹配实例 - 使用安全高亮函数
        for ref in all_matches:
            if ref in ref_set:
                try:
                    # 使用安全高亮函数处理原始位号
                    highlighted_areas = _highlight_safe_ref(page, ref, highlight_color)

                    # 如果成功高亮了位号，将其添加到找到的位号集合中
                    if highlighted_areas:
                        found_refs.add(ref)

                        # 对成功高亮的区域添加注解（如果需要）
                        if add_annotation and ref in bom_info:
                            for inst in highlighted_areas:
                                part_info = bom_info[ref]
                                # 创建注解内容
                                annot_content = f"位号: {ref}\n"

                                if 'PN' in part_info and part_info['PN']:
                                    annot_content += f"料号: {part_info['PN']}\n"

                                if 'Description' in part_info and part_info['Description']:
                                    annot_content += f"描述: {part_info['Description']}\n"

                                # 确保厂商料号与制造商分开显示
                                if 'MfrPN' in part_info and part_info['MfrPN'] and part_info['MfrPN'].strip():
                                    annot_content += f"厂商料号: {part_info['MfrPN']}\n"

                                # 确保制造商信息正确显示
                                if 'Manufacturer' in part_info and part_info['Manufacturer'] and part_info['Manufacturer'].strip():
                                    # 如果是注解的最后一项，不添加换行符
                                    annot_content += f"制造商: {part_info['Manufacturer']}"

                                # 创建文本注解
                                info_annot = page.add_text_annot(
                                    inst.br,  # 在高亮区域右下角添加注解
                                    annot_content,
                                    icon="Note"
                                )
                                # 设置注解默认关闭
                                info_annot.set_open(False)
                                info_annot.update()

                    # 处理后缀匹配情况
                    if ref and ref[0].isalpha():
                        for suffix_match, orig_ref in suffix_matches.items():
                            if orig_ref == ref and suffix_match != ref:
                                # 对于每个后缀匹配，单独处理匹配和高亮
                                suffix_areas = page.search_for(suffix_match)
                                highlighted_suffix_areas = []

                                for area in suffix_areas:
                                    # 获取区域周围的文本
                                    expanded_rect = fitz.Rect(area.x0 - 10, area.y0 - 2,
                                                             area.x1 + 10, area.y1 + 2)
                                    area_text = page.get_text("text", clip=expanded_rect)

                                    # 验证这个区域包含的是完整的后缀匹配位号
                                    pattern = r'(?:^|[^A-Za-z0-9])' + re.escape(suffix_match) + r'(?:[^A-Za-z0-9]|$)'
                                    if re.search(pattern, area_text):
                                        # 添加高亮
                                        annot = page.add_highlight_annot(area)
                                        annot.set_colors(stroke=highlight_color)
                                        annot.update()
                                        highlighted_suffix_areas.append(area)

                                # 如果找到了后缀匹配，将原始位号添加到找到的位号集合
                                if highlighted_suffix_areas and ref not in found_refs:
                                    found_refs.add(ref)

                                    # 对后缀匹配也添加注解
                                    if add_annotation and ref in bom_info:
                                        for inst in highlighted_suffix_areas:
                                            part_info = bom_info[ref]
                                            # 创建注解内容
                                            annot_content = f"位号: {ref} (匹配: {suffix_match})\n"

                                            if 'PN' in part_info and part_info['PN']:
                                                annot_content += f"料号: {part_info['PN']}\n"

                                            if 'Description' in part_info and part_info['Description']:
                                                annot_content += f"描述: {part_info['Description']}\n"

                                            # 确保厂商料号与制造商分开显示
                                            if 'MfrPN' in part_info and part_info['MfrPN'] and part_info['MfrPN'].strip():
                                                annot_content += f"厂商料号: {part_info['MfrPN']}\n"

                                            # 确保制造商信息正确显示
                                            if 'Manufacturer' in part_info and part_info['Manufacturer'] and part_info['Manufacturer'].strip():
                                                annot_content += f"制造商: {part_info['Manufacturer']}"

                                            # 创建文本注解
                                            info_annot = page.add_text_annot(
                                                inst.br,  # 在高亮区域右下角添加注解
                                                annot_content,
                                                icon="Note"
                                            )
                                            info_annot.set_open(False)
                                            info_annot.update()
                except Exception as highlight_error:
                    # 高亮单个位号失败不应该导致整个页面处理失败
                    print(f"高亮位号 {ref} 时出错: {str(highlight_error)}")

        return page_num, found_refs

    except Exception as e:
        # 捕获页面处理中的任何异常，避免整个进程崩溃
        print(f"处理页面 {page_num} 时发生错误: {str(e)}")
        return page_num, set()  # 返回空结果集

def set_windows_process_priority():
    """在Windows系统上设置进程优先级为低于正常"""
    if sys.platform == 'win32':
        try:
            import ctypes
            process_handle = ctypes.windll.kernel32.GetCurrentProcess()
            ctypes.windll.kernel32.SetPriorityClass(process_handle, 0x00004000)  # BELOW_NORMAL_PRIORITY_CLASS
        except Exception as e:
            print(f"设置Windows进程优先级失败: {str(e)}")

def pdf_processing_worker(pdf_path, ref_designators, num_threads, highlight_color, result_queue, progress_queue, add_annotation=False, bom_path=None, header_mapping=None):
    """
    在单独进程中处理PDF文件，支持BOM数据验证、替代料处理等增强功能
    Args:
        pdf_path: PDF文件路径
        ref_designators: 位号列表
        num_threads: 线程数
        highlight_color: 高亮颜色
        result_queue: 结果队列
        progress_queue: 进度队列
        add_annotation: 是否添加注解
        bom_path: BOM文件路径
        header_mapping: BOM表头映射字典
    """
    # 打印进程信息，帮助调试
    print(f"PDF处理子进程已启动 - PID: {os.getpid()}")

    try:
        # 设置当前进程优先级较低，避免影响UI响应
        try:
            if sys.platform == 'win32':
                set_windows_process_priority()
            elif sys.platform in ['linux', 'linux2']:
                os.nice(10)  # 增加nice值，降低优先级
        except Exception as e:
            print(f"设置进程优先级失败: {str(e)}")

        start_time = time.time()

        # 初始化结果字典
        total_refs = len(ref_designators)
        found_refs = set()  # 使用集合存储已找到的位号，避免重复

        # 加载BOM信息，用于添加注解
        bom_info = {}
        if add_annotation and bom_path:
            try:
                # 尝试加载BOM文件并提取每个位号对应的物料信息
                # 检测真正的表头行，处理包含项目信息的BOM
                detected_header_row = 0
                try:
                    # 读取Excel文件的前15行进行分析
                    max_rows_to_check = 15
                    # 获取文件扩展名
                    file_ext = os.path.splitext(bom_path)[1].lower()

                    # 增强Excel文件读取，尝试不同引擎
                    try:
                        # 首先尝试openpyxl引擎（适用于.xlsx文件）
                        df_preview = pd.read_excel(bom_path, engine='openpyxl', nrows=max_rows_to_check, header=None)
                    except Exception as e1:
                        print(f"使用openpyxl引擎读取预览失败: {str(e1)}")
                        try:
                            # 尝试xlrd引擎（适用于.xls文件）
                            df_preview = pd.read_excel(bom_path, engine='xlrd', nrows=max_rows_to_check, header=None)
                        except Exception as e2:
                            print(f"使用xlrd引擎读取预览失败: {str(e2)}")
                            # 使用默认引擎
                            df_preview = pd.read_excel(bom_path, nrows=max_rows_to_check, header=None)

                    # 统计每行非空单元格数量和关键词匹配
                    row_scores = []
                    for row_idx in range(min(max_rows_to_check, len(df_preview))):
                        row = df_preview.iloc[row_idx]
                        non_empty_cells = row.notna().sum()
                        if non_empty_cells > 0:
                            row_scores.append((row_idx, non_empty_cells))

                    # 选择非空单元格最多的行作为表头
                    if row_scores:
                        detected_header_row = max(row_scores, key=lambda x: x[1])[0]
                except Exception as e:
                    print(f"检测表头行出错: {str(e)}")

                # 读取BOM文件，使用适当的引擎
                try:
                    # 首先尝试openpyxl引擎（适用于.xlsx文件）
                    df = pd.read_excel(bom_path, engine='openpyxl', header=detected_header_row, dtype=str)
                except Exception as e1:
                    print(f"使用openpyxl引擎读取BOM失败: {str(e1)}")
                    try:
                        # 尝试xlrd引擎（适用于.xls文件）
                        df = pd.read_excel(bom_path, engine='xlrd', header=detected_header_row, dtype=str)
                    except Exception as e2:
                        print(f"使用xlrd引擎读取BOM失败: {str(e2)}")
                        # 使用默认引擎
                        df = pd.read_excel(bom_path, header=detected_header_row, dtype=str)

                # 处理可能的NaN值
                for col in df.columns:
                    df[col] = df[col].astype(str).replace('nan', '')

                # 确定关键列
                ref_col = None
                pn_col = None
                desc_col = None
                mfr_pn_col = None
                mfr_col = None

                # 使用传入的表头映射进行列匹配
                if header_mapping:
                    # 遍历所有列，与header_mapping中的关键词进行匹配
                    for col in df.columns:
                        col_str = str(col).lower()

                        # 位号列
                        if ref_col is None:
                            for key in header_mapping.get('reference', []):
                                if key.lower() in col_str:
                                    ref_col = col
                                    break

                        # 料号列
                        if pn_col is None:
                            for key in header_mapping.get('pn', []):
                                if key.lower() in col_str:
                                    pn_col = col
                                    break

                        # 描述列
                        if desc_col is None:
                            for key in header_mapping.get('description', []):
                                if key.lower() in col_str:
                                    desc_col = col
                                    break

                        # 先尝试匹配制造商列（精确匹配以避免误识别）
                        if mfr_col is None:
                            for key in header_mapping.get('manufacturer', []):
                                # 使用精确匹配或包含"名称"、"name"的关键词来减少误判
                                if (key.lower() == col_str or
                                    ("名称" in key.lower() and "名称" in col_str) or
                                    ("name" in key.lower() and "name" in col_str)):
                                    mfr_col = col
                                    break

                        # 再尝试匹配厂商料号列（避免与制造商列混淆）
                        if mfr_pn_col is None:
                            for key in header_mapping.get('mfr_pn', []):
                                # 使用精确匹配或包含"料号"、"型号"等关键词
                                if (key.lower() == col_str or
                                    ("料号" in key.lower() and "料号" in col_str) or
                                    ("pn" in key.lower() and "pn" in col_str) or
                                    ("型号" in key.lower() and "型号" in col_str)):
                                    mfr_pn_col = col
                                    break
                else:
                    # 如果没有传入表头映射，则使用简化的默认匹配逻辑
                    for col in df.columns:
                        col_str = str(col).lower()
                        # 位号列
                        if any(key in col_str for key in ['reference', 'ref', 'designator', '位号']):
                            ref_col = col
                        # 料号列
                        elif any(key in col_str for key in ['part number', 'pn', 'part no', '料号']):
                            pn_col = col
                        # 描述列
                        elif any(key in col_str for key in ['description', 'desc', 'comment', '描述', '说明']):
                            desc_col = col
                        # 厂商料号列 - 使用更精确的匹配条件
                        elif any(key in col_str for key in ['manufacturer part', 'mfr pn', 'part number']):
                            mfr_pn_col = col
                        # 制造商列 - 确保不被识别为厂商料号
                        elif any(key in col_str for key in ['manufacturer name', 'mfr name', 'vendor name', '制造商名称', '厂商名称']) or (
                            ('manufacturer' in col_str or 'mfr' in col_str or '制造商' in col_str or '厂商' in col_str) and
                            not any(term in col_str for term in ['part', 'pn', 'p/n', '料号', '型号'])):
                            mfr_col = col

                # 如果找到位号列，提取物料信息
                if ref_col is not None:
                    for _, row in df.iterrows():
                        # 获取位号
                        ref_value = row[ref_col]
                        if pd.isna(ref_value):
                            continue

                        # 处理位号（可能包含多个位号）
                        ref_str = str(ref_value).strip()
                        refs = [r.strip() for r in re.split(r'[,;，；\s]+', ref_str) if r.strip()]

                        # 获取其他信息
                        pn = row[pn_col] if pn_col is not None and not pd.isna(row[pn_col]) else ""
                        desc = row[desc_col] if desc_col is not None and not pd.isna(row[desc_col]) else ""
                        mfr_pn = row[mfr_pn_col] if mfr_pn_col is not None and not pd.isna(row[mfr_pn_col]) else ""
                        mfr = row[mfr_col] if mfr_col is not None and not pd.isna(row[mfr_col]) else ""

                        # 为每个位号添加信息
                        for ref in refs:
                            bom_info[ref] = {
                                'PN': str(pn).strip(),
                                'Description': str(desc).strip(),
                                'MfrPN': str(mfr_pn).strip(),
                                'Manufacturer': str(mfr).strip()
                            }
            except Exception as e:
                print(f"加载BOM信息时出错: {str(e)}")
                import traceback
                traceback.print_exc()

        # 打开PDF文件
        doc = fitz.open(pdf_path)

        # 统计总页数
        page_count = len(doc)
        progress_queue.put(("text", f"PDF共{page_count}页，开始处理...\n", False))

        # 转换为集合以提高查找效率
        ref_set = set(ref_designators)

        # 预编译正则表达式，支持字母后缀匹配
        patterns = prepare_regex_patterns(ref_designators)

        # 要处理的页面和参数组合
        page_tasks = []
        for page_num in range(page_count):
            try:
                page = doc[page_num]
                page_tasks.append((page_num + 1, page, ref_set, patterns, highlight_color, add_annotation, bom_info))
            except Exception as e:
                print(f"警告: 加载页面 {page_num + 1} 失败: {str(e)}")
                continue

        # 创建线程池
        with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
            futures = []
            for task in page_tasks:
                future = executor.submit(_process_page, task)
                futures.append(future)

            total_tasks = len(futures)
            completed = 0

            # 收集结果
            for future in concurrent.futures.as_completed(futures):
                try:
                    page_num, found = future.result()
                    found_refs.update(found)  # 合并找到的位号集合

                    # 更新进度
                    completed += 1
                    progress_percentage = (completed / total_tasks) * 100
                    progress_queue.put(("progress", progress_percentage))

                    # 每10页或完成时报告
                    if completed % 10 == 0 or completed == total_tasks:
                        elapsed = time.time() - start_time
                        remaining = (elapsed / completed) * (total_tasks - completed) if completed > 0 else 0
                        progress_queue.put(("text", f"已处理 {completed}/{total_tasks} 页 "
                                          f"({progress_percentage:.1f}%)，"
                                          f"已用时 {int(elapsed/60)}分{int(elapsed%60)}秒，"
                                          f"预计剩余 {int(remaining/60)}分{int(remaining%60)}秒\n", False))
                except Exception as e:
                    print(f"处理结果时出错: {str(e)}")

        # 保存修改后的PDF
        output_pdf = pdf_path.replace(".pdf", "_标注.pdf")
        if os.path.exists(output_pdf):
            # 如果文件已存在，添加时间戳
            timestamp = time.strftime("%Y%m%d%H%M%S")
            output_pdf = pdf_path.replace(".pdf", f"_标注_{timestamp}.pdf")

        doc.save(output_pdf)
        doc.close()

        # 清理内存
        import gc
        gc.collect()

        # 计算未找到的位号
        not_found = list(set(ref_designators) - found_refs)

        # 将结果放入队列
        result = {
            'total': total_refs,
            'found': len(found_refs),
            'not_found': not_found,
            'output_pdf': output_pdf
        }

        result_queue.put(result)

    except Exception as e:
        # 捕获所有异常，避免进程崩溃
        error_msg = f"PDF处理错误: {str(e)}"

        # 将错误信息转换为中文
        # 由于在子进程中无法访问主程序的_translate_error_to_chinese方法
        # 所以我们在这里直接进行简单的错误翻译
        chinese_error_msg = error_msg
        for eng_error, cn_error in {
            "name 'logging' is not defined": "程序内部错误：日志模块未定义，请联系开发者",
            "name 'chardet' is not defined": "缺少chardet模块，请安装：pip install chardet",
            "No module named": "缺少必要的模块，请安装相应的Python库",
            "Permission denied": "文件权限不足，请以管理员身份运行或检查文件权限",
            "File not found": "找不到指定的文件，请检查文件路径是否正确",
            "Invalid file": "无效的文件，请检查文件格式是否正确",
            "Failed to open": "打开文件失败，请确保文件未被其他程序占用",
            "Memory error": "内存不足，请关闭其他程序或增加系统内存",
            "Disk full": "磁盘空间不足，请清理磁盘空间",
            "Timeout": "操作超时，请检查网络连接或重试",
            "Connection error": "连接错误，请检查网络连接",
            "Invalid format": "无效的格式，请检查文件格式是否正确",
            "Cannot read": "无法读取文件，请检查文件是否损坏或格式不正确",
            "Cannot write": "无法写入文件，请检查文件权限或磁盘空间",
            "PDF file format": "PDF文件格式错误，请确保使用标准的PDF格式",
            "UnicodeDecodeError": "文件编码错误，请使用UTF-8或GBK编码保存文件",
            "KeyError": "程序内部错误：键值错误，请联系开发者",
            "IndexError": "程序内部错误：索引超出范围，请联系开发者",
            "TypeError": "程序内部错误：类型错误，请联系开发者",
            "ValueError": "程序内部错误：值错误，请联系开发者",
            "AttributeError": "程序内部错误：属性错误，请联系开发者",
            "ImportError": "程序内部错误：导入错误，请联系开发者",
            "ModuleNotFoundError": "缺少必要的模块，请安装相应的Python库",
            "OSError": "操作系统错误，请检查文件权限或磁盘空间",
            "FileNotFoundError": "找不到指定的文件，请检查文件路径是否正确",
            "PermissionError": "文件权限不足，请以管理员身份运行或检查文件权限",
            "FileExistsError": "文件已存在，请尝试使用其他文件名",
            "NotImplementedError": "功能尚未实现，请等待后续版本",
            "RuntimeError": "运行时错误，请重新启动程序或联系开发者",
            "Exception": "程序异常，请重新启动程序或联系开发者"
        }.items():
            if eng_error in str(e):
                chinese_error_msg = f"PDF处理错误: {cn_error}"
                break

        progress_queue.put(("warning", chinese_error_msg))

        # 记录详细错误
        import traceback
        traceback.print_exc()

        # 将错误结果放入队列
        result = {
            'error': chinese_error_msg,
            'total': len(ref_designators) if 'ref_designators' in locals() else 0,
            'found': 0,
            'not_found': [],
            'output_pdf': None
        }
        result_queue.put(result)

def verify_regex_patterns():
    """
    验证正则表达式模式是否能正确工作
    测试两个关键场景:
    1. B3201不应该匹配到FB3201
    2. U401应该匹配到U401A
    """
    print("正在验证位号匹配规则...")
    # 测试数据
    test_refs = ["B3201", "U401", "R302"]
    test_text = "FB3201 B3201 U401 U401A R302 R3021"

    # 生成测试模式
    patterns = prepare_regex_patterns(test_refs)

    # 测试精确匹配
    print("测试精确匹配...")
    exact_matches = set()
    for pattern in patterns['exact_patterns']:
        matches = pattern.findall(test_text)
        for match in matches:
            if match in test_refs:
                exact_matches.add(match)

    expected_exact = {"B3201", "U401", "R302"}
    if exact_matches == expected_exact:
        print("✓ 精确匹配正确")
    else:
        print(f"✗ 精确匹配错误! 期望: {expected_exact}, 实际: {exact_matches}")

    # 测试后缀匹配
    print("测试后缀匹配...")
    suffix_matches = {}
    for pattern in patterns['suffix_patterns']:
        suffix_matches_found = pattern.findall(test_text)
        for suffix_match in suffix_matches_found:
            for ref in test_refs:
                if ref and ref[0].isalpha() and suffix_match.startswith(ref):
                    suffix = suffix_match[len(ref):]
                    if suffix and suffix.isalpha():
                        suffix_matches[suffix_match] = ref

    # 检查U401A是否被映射到U401
    if "U401A" in suffix_matches and suffix_matches["U401A"] == "U401":
        print("✓ 后缀匹配正确 (U401 -> U401A)")
    else:
        print(f"✗ 后缀匹配错误! U401A 未正确匹配到 U401")

    # 检查FB3201是否被错误地匹配到B3201
    if any(suffix == "FB3201" for suffix in suffix_matches.keys()):
        mapped_ref = suffix_matches.get("FB3201", "")
        print(f"✗ 错误匹配! FB3201 被错误地匹配到 {mapped_ref}")
    else:
        print("✓ FB3201 未被错误匹配")

    # 检查R3021是否被错误地匹配到R302
    if any(suffix == "R3021" for suffix in suffix_matches.keys()):
        mapped_ref = suffix_matches.get("R3021", "")
        print(f"✗ 错误匹配! R3021 被错误地匹配到 {mapped_ref}")
    else:
        print("✓ R3021 未被错误匹配")

    print("位号匹配规则验证完成!\n")

# 自动更新相关函数
def check_for_updates(current_version):
    """检查是否有新版本可用

    Args:
        current_version: 当前版本号

    Returns:
        tuple: (has_update, latest_version, download_url, changelog, is_exe_update)
    """
    try:
        # 使用GitHub API获取最新版本信息
        print(f"正在检查更新，当前版本: {current_version}")
        print(f"API URL: {GITHUB_API_URL}")

        response = requests.get(GITHUB_API_URL, timeout=5)
        print(f"响应状态码: {response.status_code}")

        if response.status_code == 200:
            data = response.json()
            latest_version = data["tag_name"].replace("v", "")
            print(f"最新版本: {latest_version}")

            # 比较版本号
            if pkg_version.parse(latest_version) > pkg_version.parse(current_version):
                download_url = None
                is_exe_update = False

                # 检查当前是否是单文件exe
                is_frozen = getattr(sys, 'frozen', False)

                # 首先检查是否有任何资源文件
                if "assets" in data and len(data["assets"]) > 0:
                    # 如果是单文件exe，优先查找.exe资源
                    if is_frozen:
                        for asset in data["assets"]:
                            if asset["name"].endswith(".exe"):
                                download_url = asset["browser_download_url"]
                                is_exe_update = True
                                print(f"找到exe更新: {asset['name']}")
                                break

                    # 如果没有找到exe或不是单文件exe，查找.zip资源
                    if not download_url:
                        for asset in data["assets"]:
                            if asset["name"].endswith(".zip"):
                                download_url = asset["browser_download_url"]
                                print(f"找到zip更新: {asset['name']}")
                                break

                    # 如果还是没有找到，使用任何可用的资源
                    if not download_url and len(data["assets"]) > 0:
                        download_url = data["assets"][0]["browser_download_url"]
                        if data["assets"][0]["name"].endswith(".exe"):
                            is_exe_update = True
                        print(f"使用可用资源: {data['assets'][0]['name']}")

                # 如果没有资源文件，使用源代码下载链接
                if not download_url:
                    download_url = data["zipball_url"]
                    print("使用源代码链接作为备用")

                # 获取更新日志
                changelog = data["body"] if "body" in data else "无可用的更新日志"

                return True, latest_version, download_url, changelog, is_exe_update

        # 如果没有新版本或请求失败
        return False, current_version, "", "", False
    except Exception as e:
        print(f"检查更新失败: {str(e)}")
        return False, current_version, "", "", False

def download_update(download_url, save_path):
    """下载更新包

    Args:
        download_url: 下载链接
        save_path: 保存路径

    Returns:
        bool: 下载是否成功
    """
    try:
        response = requests.get(download_url, stream=True, timeout=30)
        total_size = int(response.headers.get('content-length', 0))

        with open(save_path, 'wb') as f:
            downloaded = 0
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    downloaded += len(chunk)
                    # 可以在这里添加进度回调
        return True
    except Exception as e:
        print(f"下载更新失败: {str(e)}")
        return False

def extract_update(zip_path, extract_path):
    """解压更新包

    Args:
        zip_path: 压缩包路径
        extract_path: 解压目标路径

    Returns:
        bool: 解压是否成功
    """
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_path)
        return True
    except Exception as e:
        print(f"解压更新包失败: {str(e)}")
        return False

def install_update(extract_path, current_path):
    """安装更新

    Args:
        extract_path: 解压目录
        current_path: 当前程序路径

    Returns:
        bool: 安装是否成功
    """
    try:
        # 查找解压目录中的第一个目录（GitHub压缩包通常有一个根目录）
        root_dirs = [d for d in os.listdir(extract_path) if os.path.isdir(os.path.join(extract_path, d))]
        if root_dirs:
            source_dir = os.path.join(extract_path, root_dirs[0])
        else:
            source_dir = extract_path

        # 复制文件
        for item in os.listdir(source_dir):
            s = os.path.join(source_dir, item)
            d = os.path.join(current_path, item)

            if os.path.isdir(s):
                if os.path.exists(d):
                    shutil.rmtree(d)
                shutil.copytree(s, d)
            else:
                if item.endswith('.py') or item.endswith('.ico') or item.endswith('.txt'):
                    shutil.copy2(s, d)

        return True
    except Exception as e:
        print(f"安装更新失败: {str(e)}")
        return False

# 统计功能已移除

def download_with_resume(url, dest_path, progress_callback=None, status_callback=None):
    """
    支持断点续传的下载函数

    Args:
        url: 下载URL
        dest_path: 保存路径
        progress_callback: 进度回调函数，接收参数(downloaded, total_size, progress_percent)
        status_callback: 状态回调函数，接收参数(status_message)

    Returns:
        bool: 下载是否成功
    """
    # 检查是否存在部分下载的文件
    file_size = 0
    if os.path.exists(dest_path):
        file_size = os.path.getsize(dest_path)
        if status_callback:
            status_callback(f"发现已下载的文件({file_size/1024:.1f}KB)，继续下载...")

    # 设置HTTP头，支持断点续传
    headers = {}
    if file_size > 0:
        headers['Range'] = f'bytes={file_size}-'

    # 打开文件，使用追加模式
    with open(dest_path, 'ab' if file_size > 0 else 'wb') as f:
        # 设置重试次数
        retry_count = 0

        while retry_count < DOWNLOAD_MAX_RETRIES:
            try:
                # 发起请求
                response = requests.get(url, headers=headers, stream=True, timeout=DOWNLOAD_TIMEOUT)

                # 检查响应状态码
                if file_size > 0 and response.status_code == 416:
                    # 范围请求错误，文件可能已经完整下载
                    if status_callback:
                        status_callback("文件已完整下载")
                    return True
                elif file_size > 0 and response.status_code != 206:
                    # 不支持断点续传，重新下载
                    if status_callback:
                        status_callback("服务器不支持断点续传，重新下载...")
                    f.close()
                    os.remove(dest_path)
                    return download_with_resume(url, dest_path, progress_callback, status_callback)
                elif response.status_code not in [200, 206]:
                    # 其他错误
                    raise Exception(f"下载失败，HTTP状态码: {response.status_code}")

                # 获取总大小
                if file_size > 0 and response.status_code == 206:
                    # 断点续传情况
                    content_range = response.headers.get('Content-Range', '')
                    match = re.search(r'bytes\s+\d+-\d+/(\d+)', content_range)
                    total_size = int(match.group(1)) if match else file_size + int(response.headers.get('content-length', 0))
                else:
                    # 普通下载情况
                    total_size = int(response.headers.get('content-length', 0)) + file_size

                # 下载文件
                downloaded = file_size
                for chunk in response.iter_content(chunk_size=DOWNLOAD_CHUNK_SIZE):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        if progress_callback and total_size > 0:
                            progress_percent = (downloaded / total_size) * 100
                            progress_callback(downloaded, total_size, progress_percent)

                # 下载成功，返回
                return True

            except (requests.exceptions.RequestException, IOError) as e:
                # 网络错误或IO错误
                retry_count += 1
                if retry_count >= DOWNLOAD_MAX_RETRIES:
                    if status_callback:
                        status_callback(f"下载失败，已重试{DOWNLOAD_MAX_RETRIES}次: {str(e)}")
                    raise Exception(f"下载失败，已重试{DOWNLOAD_MAX_RETRIES}次: {str(e)}")

                # 等待一段时间后重试，使用指数退避策略
                wait_time = 2 ** retry_count
                if status_callback:
                    status_callback(f"网络错误，{wait_time}秒后重试({retry_count}/{DOWNLOAD_MAX_RETRIES})...")
                time.sleep(wait_time)

    return False

def create_updater_script(current_exe, temp_dir):
    """创建更新脚本，在当前程序退出后运行

    Args:
        current_exe: 当前程序的可执行文件路径
        temp_dir: 临时目录

    Returns:
        str: 脚本路径
    """
    script_path = os.path.join(temp_dir, "updater.bat")
    with open(script_path, 'w') as f:
        f.write(f"@echo off\n")
        f.write(f"echo 正在更新 BOM Checker...\n")
        f.write(f"timeout /t 2 /nobreak > nul\n")
        f.write(f"start "" \"{current_exe}\"\n")

    return script_path

# 在程序开始时运行自检
try:
    verify_regex_patterns()
except Exception as e:
    print(f"位号匹配规则验证失败: {str(e)}")

class BOMCheckerApp:
    def __init__(self, master):
        """初始化应用程序"""
        self.master = master
        self.version = APP_VERSION
        self.master.title(f"BOM Checker v{self.version} | 小航  2025.5.27")

        # 设置主窗口大小和位置
        window_width = 900
        window_height = 600
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.master.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.master.minsize(800, 550)  # 设置最小尺寸

        # 配置按钮样式、颜色和图标
        self.setup_styles()

        # 设置变量
        self.bom_path = ""
        self.pdf_path = ""
        self.output_pdf_path = None

        # 处理标志
        self.processing = False

        # 更新相关变量
        self.update_available = False
        self.latest_version = ""
        self.download_url = ""
        self.update_changelog = ""
        self.is_exe_update = False
        self.new_exe_name = ""  # 智能文件名处理后的新文件名
        self.update_window_open = False  # 更新窗口是否打开

        # 添加注解选项变量
        self.add_annotation_var = tk.BooleanVar(value=False)

        # 初始化保存格式选项变量
        self.save_as_excel_var = tk.BooleanVar(value=False)  # 默认为txt格式

        # 创建消息队列，用于多线程通信
        self.message_queue = queue.Queue()

        # 加载用户设置
        self.load_settings()

        # 初始化默认表头映射
        self.load_header_mapping()

        # 创建界面组件
        self.create_widgets()

        # 启动时检查更新（非阻塞方式）
        threading.Thread(target=self.check_updates_on_startup, daemon=True).start()

    def create_menu(self):
        """创建菜单栏"""
        menubar = tk.Menu(self.master)
        self.master.config(menu=menubar)

        # 文件菜单
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="文件", menu=file_menu)
        file_menu.add_command(label="选择BOM文件", command=self.select_bom)
        file_menu.add_command(label="选择PDF文件", command=self.select_pdf)
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.on_closing)

        # 工具菜单
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="工具", menu=tools_menu)
        tools_menu.add_command(label="设置", command=self.show_settings)
        tools_menu.add_command(label="表头映射编辑器", command=self.show_mapping_settings)

        # 帮助菜单
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="帮助", menu=help_menu)
        help_menu.add_command(label="使用帮助", command=self.show_help)
        help_menu.add_command(label="检查更新", command=self.check_updates_manually)
        help_menu.add_separator()
        help_menu.add_command(label="关于", command=self.show_about)

    def check_updates_on_startup(self):
        """程序启动时检查更新"""
        # 延迟几秒，让主界面先加载完成
        time.sleep(2)

        # 检查上次更新时间，如果距离上次检查时间不足UPDATE_CHECK_INTERVAL天，则不检查
        config = self._load_config()
        last_check = config.get('last_update_check', 0)
        now = time.time()

        if now - last_check < UPDATE_CHECK_INTERVAL * 24 * 60 * 60:
            return

        # 检查更新
        has_update, latest_version, download_url, changelog, is_exe_update = check_for_updates(self.version)

        # 更新最后检查时间
        config['last_update_check'] = now
        self._save_config(config)

        if has_update:
            # 保存更新信息
            self.update_available = True
            self.latest_version = latest_version
            self.download_url = download_url
            self.update_changelog = changelog
            self.is_exe_update = is_exe_update

            # 显示更新提示
            self.master.after(0, self.show_update_notification)

    def check_updates_manually(self):
        """手动检查更新"""
        # 如果更新窗口已经打开，则返回
        if self.update_window_open:
            return

        # 显示正在检查的提示
        self._update_status("正在检查更新...", self.info_color)

        # 在单独的线程中检查更新
        threading.Thread(target=self._check_updates_thread, daemon=True).start()

    def _check_updates_thread(self):
        """在单独线程中检查更新"""
        try:
            has_update, latest_version, download_url, changelog, is_exe_update = check_for_updates(self.version)

            # 更新最后检查时间
            config = self._load_config()
            config['last_update_check'] = time.time()
            self._save_config(config)

            if has_update:
                # 保存更新信息
                self.update_available = True
                self.latest_version = latest_version
                self.download_url = download_url
                self.update_changelog = changelog
                self.is_exe_update = is_exe_update

                # 显示更新对话框
                self.master.after(0, self.show_update_dialog)
            else:
                # 显示无更新的提示
                self.master.after(0, lambda: messagebox.showinfo("检查更新", f"当前版本 {self.version} 已是最新版本"))
                self.master.after(0, lambda: self._update_status("当前已是最新版本", self.success_color))
                # 2秒后恢复状态栏
                self.master.after(2000, lambda: self._update_status("就绪", self.text_color))
        except Exception as e:
            # 显示错误提示
            self.master.after(0, lambda: messagebox.showerror("检查更新失败", f"检查更新时出错：{str(e)}"))
            self.master.after(0, lambda: self._update_status("检查更新失败", self.error_color))
            # 2秒后恢复状态栏
            self.master.after(2000, lambda: self._update_status("就绪", self.text_color))

    def show_update_notification(self):
        """显示更新通知"""
        # 如果更新窗口已经打开，则返回
        if self.update_window_open:
            return

        if messagebox.askyesno("发现新版本",
                            f"发现新版本 v{self.latest_version}，当前版本 v{self.version}。\n\n是否查看更新内容并更新？"):
            self.show_update_dialog()

    def show_update_dialog(self):
        """显示更新对话框"""
        # 如果更新窗口已经打开，则返回
        if self.update_window_open:
            return

        # 设置窗口打开标志
        self.update_window_open = True

        update_window = tk.Toplevel(self.master)
        update_window.title("发现新版本")
        update_window.geometry("500x500")
        update_window.transient(self.master)  # 设置为主窗口的子窗口
        update_window.grab_set()  # 模态对话框
        update_window.withdraw()  # 先隐藏窗口，等配置完成后再显示

        # 添加窗口关闭事件处理
        def on_window_close():
            self.update_window_open = False
            # 更新状态栏，清除"正在检查更新..."的提示
            self._update_status("就绪", self.text_color)
            update_window.destroy()

        update_window.protocol("WM_DELETE_WINDOW", on_window_close)

        # 添加内边距
        main_frame = ttk.Frame(update_window, padding=(20, 15))
        main_frame.pack(fill='both', expand=True)

        # 标题
        ttk.Label(main_frame, text=f"发现新版本: v{self.latest_version}",
                font=('\u5fae\u8f6f\u96c5\u9ed1', 12, 'bold')).pack(pady=(0, 10))

        # 版本信息
        version_frame = ttk.Frame(main_frame)
        version_frame.pack(fill='x', pady=(0, 10))

        ttk.Label(version_frame, text=f"当前版本: v{self.version}",
                font=('\u5fae\u8f6f\u96c5\u9ed1', 10)).pack(side='left')
        ttk.Label(version_frame, text=f"最新版本: v{self.latest_version}",
                font=('\u5fae\u8f6f\u96c5\u9ed1', 10, 'bold')).pack(side='right')

        # 更新日志
        ttk.Label(main_frame, text="更新内容:",
                font=('\u5fae\u8f6f\u96c5\u9ed1', 10, 'bold')).pack(anchor='w', pady=(10, 5))

        # 更新日志文本框
        changelog_frame = ttk.Frame(main_frame)
        changelog_frame.pack(fill='both', expand=True, pady=(0, 10))

        changelog_text = scrolledtext.ScrolledText(changelog_frame, wrap=tk.WORD, height=10)
        changelog_text.pack(fill='both', expand=True)
        changelog_text.insert(tk.END, self.update_changelog)
        changelog_text.config(state=tk.DISABLED)  # 设置为只读

        # 按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x', pady=(10, 0))

        ttk.Button(button_frame, text="稍后提醒", command=update_window.destroy).pack(side='left')

        # 下载并安装按钮
        update_button = ttk.Button(button_frame, text="下载并安装",
                                command=lambda: self.download_and_install_update(update_window))
        update_button.pack(side='right')

        # 居中窗口
        update_window.update_idletasks()
        width = update_window.winfo_width()
        height = update_window.winfo_height()
        x = (update_window.winfo_screenwidth() // 2) - (width // 2)
        y = (update_window.winfo_screenheight() // 2) - (height // 2)
        update_window.geometry(f"+{x}+{y}")

        # 显示窗口
        update_window.deiconify()

    def download_and_install_update(self, update_window=None):
        """下载并安装更新"""
        if update_window:
            self.update_window_open = False  # 重置窗口状态
            # 更新状态栏，清除"正在检查更新..."的提示
            self._update_status("就绪", self.text_color)
            update_window.destroy()

        # 创建进度对话框
        progress_window = tk.Toplevel(self.master)
        progress_window.title("正在更新")
        progress_window.geometry("400x150")
        progress_window.transient(self.master)  # 设置为主窗口的子窗口
        progress_window.grab_set()  # 模态对话框
        progress_window.withdraw()  # 先隐藏窗口，等配置完成后再显示

        # 添加内边距
        main_frame = ttk.Frame(progress_window, padding=(20, 15))
        main_frame.pack(fill='both', expand=True)

        # 标题
        ttk.Label(main_frame, text="正在下载更新...",
                font=('\u5fae\u8f6f\u96c5\u9ed1', 10, 'bold')).pack(pady=(0, 10))

        # 进度条
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(main_frame, variable=progress_var, maximum=100)
        progress_bar.pack(fill='x', pady=(0, 10))

        # 状态标签
        status_var = tk.StringVar(value="正在准备下载...")
        status_label = ttk.Label(main_frame, textvariable=status_var)
        status_label.pack(pady=(0, 10))

        # 居中窗口
        progress_window.update_idletasks()
        width = progress_window.winfo_width()
        height = progress_window.winfo_height()
        x = (progress_window.winfo_screenwidth() // 2) - (width // 2)
        y = (progress_window.winfo_screenheight() // 2) - (height // 2)
        progress_window.geometry(f"+{x}+{y}")

        # 显示窗口
        progress_window.deiconify()

        # 在单独的线程中下载和安装更新
        threading.Thread(target=self._download_and_install_thread,
                        args=(progress_var, status_var, progress_window),
                        daemon=True).start()

    def _download_and_install_thread(self, progress_var, status_var, progress_window):
        """在单独线程中下载和安装更新"""
        try:
            # 创建临时目录
            temp_dir = tempfile.mkdtemp()

            # 检查是否是单文件exe更新
            if self.is_exe_update and getattr(sys, 'frozen', False):
                # 单文件exe的更新方式
                exe_name = os.path.basename(sys.executable)

                # 智能文件名处理逻辑
                new_exe_name = self._get_updated_filename(exe_name, self.latest_version)
                print(f"原始文件名: {exe_name}, 新文件名: {new_exe_name}")

                # 保存新文件名供后续使用
                self.new_exe_name = new_exe_name

                download_path = os.path.join(temp_dir, f"BOM_Checker_new_{self.latest_version}.exe")

                # 更新状态
                self.master.after(0, lambda: status_var.set("正在下载更新..."))

                # 使用支持断点续传的下载函数
                def update_progress(downloaded, total_size, progress_percent):
                    # 下载占总进度的90%
                    progress = progress_percent * 0.9
                    self.master.after(0, lambda p=progress: progress_var.set(p))

                def update_status(status_message):
                    self.master.after(0, lambda msg=status_message: status_var.set(msg))

                # 执行下载
                try:
                    download_success = download_with_resume(
                        self.download_url,
                        download_path,
                        progress_callback=update_progress,
                        status_callback=update_status
                    )

                    if not download_success:
                        raise Exception("下载失败")
                except Exception as e:
                    raise Exception(f"下载更新失败: {str(e)}")

                # 更新状态
                self.master.after(0, lambda: status_var.set("准备安装更新..."))
                self.master.after(0, lambda: progress_var.set(95))

                # 创建更新批处理脚本
                updater_script = os.path.join(temp_dir, "updater.bat")
                current_exe = sys.executable
                install_dir = os.path.dirname(current_exe)

                # 打印调试信息
                print(f"当前程序路径: {current_exe}")
                print(f"下载的更新文件: {download_path}")

                # 创建更简单的更新脚本，使用管理员权限运行
                with open(updater_script, 'w', encoding='gbk') as f:
                    f.write('@echo off\n')
                    f.write('echo 正在更新 BOM Checker...\n')
                    f.write('echo 请关闭原程序后更新将自动继续...\n')

                    # 等待原程序退出的循环
                    f.write(':wait_loop\n')
                    f.write(f'tasklist /FI "IMAGENAME eq {os.path.basename(sys.executable)}" /NH | find "{os.path.basename(sys.executable)}" >nul\n')
                    f.write('if %ERRORLEVEL% EQU 0 (\n')
                    f.write('    timeout /t 1 /nobreak >nul\n')
                    f.write('    goto wait_loop\n')
                    f.write(')\n')
                    f.write('echo 原程序已关闭，继续更新...\n')
                    f.write('timeout /t 2 /nobreak >nul\n')

                    # 创建管理员权限脚本
                    admin_script = os.path.join(temp_dir, "admin_update.bat")
                    f.write(f'echo @echo off > "{admin_script}"\n')
                    f.write(f'echo cd /d "%~dp0" >> "{admin_script}"\n')

                    # 复制新文件
                    target_path = os.path.join(os.path.dirname(current_exe), self.new_exe_name)
                    f.write(f'echo copy /Y "{download_path}" "{target_path}" >> "{admin_script}"\n')

                    # 清理临时文件
                    f.write(f'echo timeout /t 2 /nobreak >nul >> "{admin_script}"\n')
                    f.write(f'echo rmdir /S /Q "{temp_dir}" >> "{admin_script}"\n')

                    # 启动新版本
                    f.write(f'echo start "" "{target_path}" >> "{admin_script}"\n')

                    # 使用管理员权限运行更新脚本
                    f.write(f'powershell -Command "Start-Process -FilePath \'{admin_script}\' -Verb RunAs"\n')

                # 关闭进度窗口
                self.master.after(0, progress_window.destroy)

                # 显示提示并退出
                def show_info_and_quit():
                    result = messagebox.showinfo("更新就绪",
                        "更新已下载完成。\n点击确定后程序将自动关闭，更新将自动继续。")
                    # 在消息框关闭后自动退出程序
                    self.master.quit()

                self.master.after(0, show_info_and_quit)

                # 启动更新脚本
                subprocess.Popen(updater_script, shell=True)
                # 不再强制退出程序，等待用户手动关闭
            else:
                # 非单文件exe的更新方式（使用zip包）
                zip_path = os.path.join(temp_dir, "update.zip")
                extract_path = os.path.join(temp_dir, "extract")
                os.makedirs(extract_path, exist_ok=True)

                # 更新状态
                self.master.after(0, lambda: status_var.set("正在下载更新..."))

                # 使用支持断点续传的下载函数
                def update_progress(downloaded, total_size, progress_percent):
                    # 下载占总进度的50%
                    progress = progress_percent * 0.5
                    self.master.after(0, lambda p=progress: progress_var.set(p))

                def update_status(status_message):
                    self.master.after(0, lambda msg=status_message: status_var.set(msg))

                # 执行下载
                try:
                    download_success = download_with_resume(
                        self.download_url,
                        zip_path,
                        progress_callback=update_progress,
                        status_callback=update_status
                    )

                    if not download_success:
                        raise Exception("下载失败")
                except Exception as e:
                    raise Exception(f"下载更新失败: {str(e)}")

                # 更新状态
                self.master.after(0, lambda: status_var.set("正在解压更新包..."))
                self.master.after(0, lambda: progress_var.set(60))  # 解压占总进度的10%

                # 解压更新包
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_path)

                # 更新状态
                self.master.after(0, lambda: status_var.set("正在安装更新..."))
                self.master.after(0, lambda: progress_var.set(80))  # 安装占总进度的20%

                # 获取当前程序路径
                current_path = os.path.dirname(os.path.abspath(__file__))

                # 查找解压目录中的第一个目录（GitHub压缩包通常有一个根目录）
                root_dirs = [d for d in os.listdir(extract_path) if os.path.isdir(os.path.join(extract_path, d))]
                if root_dirs:
                    source_dir = os.path.join(extract_path, root_dirs[0])
                else:
                    source_dir = extract_path

                # 创建更新脚本
                if getattr(sys, 'frozen', False) and not self.is_exe_update:
                    # 如果是打包后的可执行文件（非单文件exe）
                    current_exe = sys.executable
                    updater_script = create_updater_script(current_exe, temp_dir)

                    # 关闭进度窗口
                    self.master.after(0, progress_window.destroy)

                    # 显示成功提示
                    def show_info_and_quit():
                        result = messagebox.showinfo("更新就绪",
                            "更新已下载完成，点击确定后程序将自动关闭并重启以完成安装。")
                        # 在消息框关闭后自动退出程序
                        self.master.quit()

                    self.master.after(0, show_info_and_quit)

                    # 运行更新脚本
                    subprocess.Popen([updater_script], shell=True)
                    # 不需要再次调用quit，因为在show_info_and_quit中已经处理了
                else:
                    # 如果是源代码运行，直接复制文件
                    for item in os.listdir(source_dir):
                        s = os.path.join(source_dir, item)
                        d = os.path.join(current_path, item)

                        if os.path.isdir(s):
                            if os.path.exists(d):
                                shutil.rmtree(d)
                            shutil.copytree(s, d)
                        else:
                            if item.endswith('.py') or item.endswith('.ico') or item.endswith('.txt'):
                                shutil.copy2(s, d)

                    # 更新状态
                    self.master.after(0, lambda: progress_var.set(100))
                    self.master.after(0, lambda: status_var.set("更新完成！"))

                    # 关闭进度窗口
                    self.master.after(2000, progress_window.destroy)

                    # 显示成功提示
                    self.master.after(2100, lambda: messagebox.showinfo("更新成功",
                                                            "更新已安装完成，请重启程序以应用更新。"))
        except Exception as e:
            # 关闭进度窗口
            self.master.after(0, progress_window.destroy)

            # 显示错误提示
            self.master.after(0, lambda: messagebox.showerror("更新失败",
                                                    f"下载或安装更新时出错：{str(e)}"))

    def __init_default_mapping(self):
        """初始化默认的BOM表头映射 - 增强版本"""
        self.bom_header_mapping = {
            'reference': ['Reference', 'Ref', '位号', '元件位号', '器件位号', '元件编号', '参考标识符'],
            'pn': ['Part Number', 'PartNumber', 'PN', 'P/N', '料号', '物料编码', '物料编号', '型号'],
            'part': ['Type', 'Value', '元件类型', '器件类型', '组件类型', '类型'],
            'quantity': ['Quantity', 'QTY', 'Qty', 'Number', '数量', '用量', '元件数量', '元件用量'],
            'description': ['Description', 'Desc', 'Note', '描述', '说明', '备注', '注释', '规格'],
            'mfr_pn': [
                # 英文别名 - 精确匹配
                'ManuFacturer P/N', 'Manufacturer Part Number', 'MFR PN', 'MfrPN', 'MFR_PN', 'MFR-PN',
                'Manufacturer PN', 'Mfr Part Number', 'Mfg Part Number', 'Supplier PN', 'Vendor PN',
                'Factory PN', 'Original PN', 'OEM PN', 'Component PN', 'Device PN', 'Chip PN',
                'Supplier Part Number', 'Vendor Part Number', 'Factory Part Number', 'Original Part Number',
                'OEM Part Number', 'Component Part Number', 'Device Part Number', 'Chip Part Number',
                'MPN', 'M/P/N', 'M-P-N', 'MPNumber', 'Manufacturer P/N', 'Manuf P/N', 'Mfr P/N', 'Mfg P/N',
                # 中文别名
                '制造商零件编号', '厂商料号', '厂商型号', '厂家料号', '厂家型号', '制造商料号', '制造商型号',
                '供应商料号', '供应商型号', '原厂料号', '原厂型号', '厂商编号', '厂家编号', '制造商编号',
                '供应商编号', '原厂编号', '厂商物料号', '厂家物料号', '制造商物料号', '供应商物料号',
                '原厂物料号', '器件料号', '器件型号', '芯片料号', '芯片型号', '元件料号', '元件型号',
                '零件料号', '零件型号', '厂商料', '厂商编码', '制造商料', '制造商编码'
            ],
            'manufacturer': [
                'Manufacturer', 'Manufacturer Name', 'Mfr', 'Mfr Name', 'Vendor', 'Vendor Name',
                'Supplier', 'Supplier Name', 'Factory', 'Factory Name', 'Producer', 'Producer Name',
                'Maker', 'Maker Name', 'Brand', 'Brand Name',
                '制造商名称', '厂商名称', '厂家名称', '供应商名称', '生产商名称', '制造厂商', '生产厂商',
                '制造商', '厂商', '厂家', '供应商', '生产商', '品牌', '品牌名称'
            ]
        }

    def load_header_mapping(self):
        """从配置文件加载BOM表头映射"""
        # 先初始化默认映射
        self.__init_default_mapping()

        # 使用统一的配置文件和路径处理
        config = self._load_config()

        if 'header_mapping' in config:
            try:
                saved_mapping = config['header_mapping']

                # 合并已保存的映射和默认映射
                for key, values in saved_mapping.items():
                    if key in self.bom_header_mapping:
                        # 确保没有重复
                        for val in values:
                            if val not in self.bom_header_mapping[key]:
                                self.bom_header_mapping[key].append(val)
                    else:
                        self.bom_header_mapping[key] = values

                print(f"已加载自定义表头映射：{len(saved_mapping)}个字段")
                # 输出详细的映射信息，特别关注厂商料号
                if 'mfr_pn' in saved_mapping:
                    print(f"  厂商料号列映射: {saved_mapping['mfr_pn']}")
            except Exception as e:
                print(f"加载表头映射出错: {str(e)}")
        else:
            print("未找到自定义表头映射，使用默认映射")

        # 输出最终的表头映射状态
        print(f"最终表头映射状态:")
        for key, values in self.bom_header_mapping.items():
            print(f"  {key}: {values[:3]}{'...' if len(values) > 3 else ''} (共{len(values)}个)")

    def save_header_mapping(self):
        """保存BOM表头映射到配置文件"""
        try:
            # 先读取现有配置
            config = self._load_config()

            # 更新表头映射部分
            config['header_mapping'] = self.bom_header_mapping

            # 保存配置
            self._save_config(config)
            print("已保存表头映射配置")
        except Exception as e:
            print(f"保存表头映射出错: {str(e)}")

    def _get_config_path(self):
        """获取配置文件路径，适配exe运行环境"""
        try:
            # 针对PyInstaller打包的应用程序
            if getattr(sys, 'frozen', False):
                # 如果是打包后的exe程序
                app_dir = os.path.dirname(sys.executable)
            else:
                # 正常的Python脚本
                app_dir = os.path.dirname(os.path.abspath(__file__))

            # 确保路径存在
            os.makedirs(app_dir, exist_ok=True)

            config_file = os.path.join(app_dir, "bom_checker_config.json")
            print(f"配置文件路径: {config_file}")
            return config_file
        except Exception as e:
            print(f"获取配置路径错误: {str(e)}")
            # 回退到当前工作目录
            fallback_path = os.path.join(os.getcwd(), "bom_checker_config.json")
            print(f"使用备用配置路径: {fallback_path}")
            return fallback_path

    def _get_updated_filename(self, original_filename, new_version):
        """智能处理文件名中的版本号

        Args:
            original_filename: 原始文件名
            new_version: 新版本号

        Returns:
            更新后的文件名
        """
        import re

        # 常见的版本号模式：
        # 1. BOM_Checker_v1.6.exe
        # 2. BOM_Checker-v1.6.exe
        # 3. BOM_Checker_1.6.exe
        # 4. BOM_Checker-1.6.exe
        # 5. BOM_Checker v1.6.exe
        # 6. BOM_Checker 1.6.exe

        # 定义版本号模式
        version_patterns = [
            r'(_v)([0-9]+\.[0-9]+(?:\.[0-9]+)?)',  # BOM_Checker_v1.6.exe
            r'(-v)([0-9]+\.[0-9]+(?:\.[0-9]+)?)',  # BOM_Checker-v1.6.exe
            r'(_)([0-9]+\.[0-9]+(?:\.[0-9]+)?)',    # BOM_Checker_1.6.exe
            r'(-)([0-9]+\.[0-9]+(?:\.[0-9]+)?)',    # BOM_Checker-1.6.exe
            r'( v)([0-9]+\.[0-9]+(?:\.[0-9]+)?)',   # BOM_Checker v1.6.exe
            r'( )([0-9]+\.[0-9]+(?:\.[0-9]+)?)'     # BOM_Checker 1.6.exe
        ]

        # 尝试匹配每一种模式
        for pattern in version_patterns:
            match = re.search(pattern, original_filename)
            if match:
                # 找到版本号，替换为新版本
                prefix = match.group(1)  # 分隔符（_v, -v, _, -, 等）
                return re.sub(pattern, f"{prefix}{new_version}", original_filename)

        # 如果没有找到版本号模式，返回原始文件名
        return original_filename

    def _load_config(self):
        """加载统一配置文件"""
        config_file = self._get_config_path()

        # 默认配置
        config = {
            'header_mapping': {},
            'highlight_color': "橙色",
            'add_annotation': False,
            'last_save_format': 'xlsx'  # 添加默认保存格式设置
        }

        # 尝试读取现有配置
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    saved_config = json.load(f)
                    # 合并配置，保留默认值
                    for key, value in saved_config.items():
                        config[key] = value
                print("已加载配置文件")
            except Exception as e:
                print(f"加载配置出错: {str(e)}")

        return config

    def _save_config(self, config):
        """保存统一配置文件，增强可靠性"""
        config_file = self._get_config_path()

        try:
            # 确保目录存在
            config_dir = os.path.dirname(config_file)
            os.makedirs(config_dir, exist_ok=True)

            # 先尝试将配置写入临时文件
            temp_file = config_file + ".tmp"
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)

            # 如果临时文件写入成功，再重命名为正式文件
            if os.path.exists(temp_file):
                # 如果正式文件已存在，先尝试删除
                if os.path.exists(config_file):
                    try:
                        os.remove(config_file)
                    except:
                        pass
                # 重命名临时文件为正式文件
                os.rename(temp_file, config_file)
                print(f"已保存配置到文件: {config_file}")
            else:
                print("临时配置文件创建失败")

        except Exception as e:
            print(f"保存配置出错: {str(e)}")
            # 尝试在当前目录保存
            try:
                fallback_file = os.path.join(os.getcwd(), "bom_checker_config.json")
                with open(fallback_file, 'w', encoding='utf-8') as f:
                    json.dump(config, f, ensure_ascii=False, indent=2)
                print(f"已保存配置到备用位置: {fallback_file}")
            except Exception as nested_e:
                print(f"保存配置到备用位置也失败: {str(nested_e)}")
                # 最后尝试保存到用户目录
                try:
                    user_dir = os.path.expanduser("~")
                    user_file = os.path.join(user_dir, "bom_checker_config.json")
                    with open(user_file, 'w', encoding='utf-8') as f:
                        json.dump(config, f, ensure_ascii=False, indent=2)
                    print(f"已保存配置到用户目录: {user_file}")
                except Exception as final_e:
                    print(f"所有保存配置尝试都失败: {str(final_e)}")

    def setup_styles(self):
        """设置GUI样式"""
        # 设置主题颜色
        self.primary_color = "#4a86e8"
        self.secondary_color = "#f1f3f4"
        self.text_color = "#202124"
        self.success_color = "#34a853"
        self.warning_color = "#fbbc05"
        self.error_color = "#ea4335"
        self.info_color = "#4a86e8"  # 信息颜色，使用主色调

        # 高亮颜色设置
        self.color_var = tk.StringVar()
        self.color_choices = {
            "黄色": (1, 1, 0),
            "红色": (1, 0, 0),
            "绿色": (0, 1, 0),
            "蓝色": (0, 0, 1),
            "粉色": (1, 0.5, 0.5),
            "橙色": (1, 0.65, 0),
            "紫色": (0.5, 0, 0.5)
        }
        self.color_var.set("橙色")  # 默认选择橙色
        self.highlight_color = (1, 0.65, 0)  # 默认橙色 (RGB格式，值范围0-1)

        # 初始化颜色预览属性
        self.color_preview = None

        # 添加注解选项
        self.add_annotation_var = tk.BooleanVar(value=False)

        # 设置ttk样式
        self.style = ttk.Style()
        self.style.configure('TButton', font=('微软雅黑', 9))
        self.style.configure('TLabel', font=('微软雅黑', 9))
        self.style.configure('TCheckbutton', font=('微软雅黑', 9))
        self.style.configure('TRadiobutton', font=('微软雅黑', 9))

        # 设置窗口关闭事件处理
        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)

        # 尝试设置图标
        try:
            self.master.iconbitmap("icon.ico")
        except:
            pass  # 如果图标不存在，忽略错误

    def create_widgets(self):
        # 创建菜单栏
        self.create_menu()

        # 主容器，使用更好的内边距
        main_container = ttk.Frame(self.master, padding=(20, 15, 20, 15))
        main_container.pack(fill='both', expand=True)

        # 顶部标题
        header_frame = ttk.Frame(main_container)
        header_frame.pack(fill='x', pady=(0, 15))

        ttk.Label(header_frame, text="BOM与原理图位号检查工具",
                style='Header.TLabel').pack(side='left')

        # 设置按钮
        self.settings_button = ttk.Button(header_frame, text="设置",
                                     command=self.show_settings)
        self.settings_button.pack(side='right', padx=5)

        # 检查更新按钮
        self.update_button = ttk.Button(header_frame, text="检查更新",
                                    command=self.check_updates_manually)
        self.update_button.pack(side='right', padx=5)

        # 文件选择区域 - 使用更好的布局
        file_frame = ttk.LabelFrame(main_container, text="文件选择", padding=(10, 5))
        file_frame.pack(fill='x', pady=(0, 15))

        # BOM文件选择
        bom_frame = ttk.Frame(file_frame)
        bom_frame.pack(fill='x', pady=5)

        ttk.Button(bom_frame, text="选择BOM文件",
                  command=self.select_bom, width=15).pack(side='left', padx=(0, 10))

        self.bom_label = ttk.Label(bom_frame, text="未选择文件", style='FileInfo.TLabel')
        self.bom_label.pack(side='left', fill='x', expand=True)

        # PDF文件选择
        pdf_frame = ttk.Frame(file_frame)
        pdf_frame.pack(fill='x', pady=5)

        ttk.Button(pdf_frame, text="选择PDF原理图",
                  command=self.select_pdf, width=15).pack(side='left', padx=(0, 10))

        self.pdf_label = ttk.Label(pdf_frame, text="未选择文件", style='FileInfo.TLabel')
        self.pdf_label.pack(side='left', fill='x', expand=True)

        # 进度显示和执行按钮
        progress_frame = ttk.LabelFrame(main_container, text="执行控制", padding=(10, 5))
        progress_frame.pack(fill='x', pady=(0, 15))

        # 保存进度框架的引用
        self.progress_frame = progress_frame

        # 进度条和百分比在一行
        progress_row = ttk.Frame(progress_frame)
        progress_row.pack(fill='x', pady=5)

        # 创建进度变量并设置初始值
        self.progress_var = tk.DoubleVar()
        self.progress_var.set(0)

        self.progress = ttk.Progressbar(progress_row, orient='horizontal', mode='determinate', variable=self.progress_var)
        self.progress.pack(fill='x', side='left', expand=True, padx=(0, 10))

        self.progress_label = ttk.Label(progress_row, text="0%", width=5)
        self.progress_label.pack(side='left')

        # 运行按钮
        button_row = ttk.Frame(progress_frame)
        button_row.pack(fill='x', pady=(5, 0))

        self.status_label = ttk.Label(button_row, text="就绪", foreground=self.success_color)
        self.status_label.pack(side='left')

        self.run_button = ttk.Button(button_row, text="开始检查",
                                  command=self.run_check,
                                  style='Primary.TButton')
        self.run_button.pack(side='right')

        # 结果展示区域 - 自定义标题栏
        result_frame = ttk.Frame(main_container)
        result_frame.pack(fill='both', expand=True)

        # 创建结果标题栏
        title_frame = ttk.Frame(result_frame)
        title_frame.pack(fill='x', side='top')

        ttk.Label(title_frame, text="检查结果", font=('微软雅黑', 9, 'bold')).pack(side='left', padx=5, pady=2)

        # 保存结果按钮
        self.save_results_button = ttk.Button(title_frame, text="保存结果", command=self.save_results, width=8)
        self.save_results_button.pack(side='right', padx=5)
        self.save_results_button.config(state='disabled')  # 初始状态为禁用

        # 添加保存格式选择
        # 使用已初始化的self.save_as_excel_var变量，不再重新创建
        save_format_frame = ttk.LabelFrame(title_frame, text="保存格式")
        save_format_frame.pack(side='right', padx=5)

        # 创建单选按钮组
        ttk.Radiobutton(save_format_frame, text="TXT", variable=self.save_as_excel_var, value=False).pack(side='left', padx=5)
        ttk.Radiobutton(save_format_frame, text="Excel", variable=self.save_as_excel_var, value=True).pack(side='left', padx=5)

        # 保存PDF按钮区域的容器引用，初始为空
        self.pdf_buttons_frame = ttk.Frame(title_frame)
        self.pdf_buttons_frame.pack(side='right', padx=5)

        # 结果内容区域
        content_frame = ttk.Frame(result_frame, relief='groove', borderwidth=1)
        content_frame.pack(fill='both', expand=True, padx=5, pady=(0, 5))

        # 结果文本框和滚动条
        result_container = ttk.Frame(content_frame)
        result_container.pack(fill='both', expand=True, padx=5, pady=5)

        # 使用更大的字体
        self.result_text = tk.Text(result_container, height=10, wrap='word',
                                font=('微软雅黑', 10))
        self.result_text.pack(side='left', fill='both', expand=True)

        # 配置文本标签样式
        self.result_text.tag_configure('header', foreground='#0066CC', font=('微软雅黑', 12, 'bold'))
        self.result_text.tag_configure('subheader', foreground='#006633', font=('微软雅黑', 11, 'bold'))
        self.result_text.tag_configure('warning', foreground='#CC3300', font=('微软雅黑', 10, 'bold'))
        self.result_text.tag_configure('success', foreground='#009900', font=('微软雅黑', 10, 'bold'))
        self.result_text.tag_configure('info', foreground='#333333', font=('微软雅黑', 10))
        self.result_text.tag_configure('emphasis', foreground='#0000FF', font=('微软雅黑', 10, 'bold'))

        scroll = ttk.Scrollbar(result_container, command=self.result_text.yview)
        scroll.pack(side='right', fill='y')
        self.result_text.config(yscrollcommand=scroll.set)

        # 保存结果框架的引用
        self.result_frame = result_frame
        self.title_frame = title_frame
        self.result_container = result_container

        # 设置初始说明文本
        self.result_text.insert('end', """欢迎使用BOM检查工具！

使用步骤：
1. 选择BOM文件（Excel格式）
2. 选择PDF文件
3. 选择高亮颜色和处理线程数
4. 点击"开始检查"按钮

检查完成后，将生成标记了所有BOM位号的PDF文件。""", 'info')
        self.result_text.config(state='disabled')  # 初始设为只读

    def select_bom(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx;*.xls")])
        if filepath:
            self.bom_path = filepath
            self.bom_label.config(text=filepath)  # 显示完整路径
            self._update_status("已选择BOM文件", self.success_color)

    def show_help(self):
        help_window = tk.Toplevel(self.master)
        help_window.title("使用帮助")
        help_window.geometry("800x620")
        help_window.minsize(700, 550)

        # 先隐藏窗口，直到所有内容都加载完成
        help_window.withdraw()

        # 设置图标
        try:
            help_window.iconbitmap("BOM_Checker/icon.ico")
        except:
            pass  # 如果图标文件不存在，则忽略

        # 设置主窗口为模态，禁止操作主窗口
        help_window.transient(self.master)
        help_window.grab_set()

        # 添加滚动条和文本框
        help_container = ttk.Frame(help_window, padding=20)
        help_container.pack(fill='both', expand=True)

        # 标题区域
        header_frame = ttk.Frame(help_container)
        header_frame.pack(fill='x', pady=(0, 15))

        # 尝试加载并显示图标
        try:
            from PIL import Image, ImageTk
            icon_img = Image.open("BOM_Checker/icon.ico")
            icon_img = icon_img.resize((48, 48))  # 增大图标尺寸
            icon_photo = ImageTk.PhotoImage(icon_img)
            icon_label = ttk.Label(header_frame, image=icon_photo)
            icon_label.image = icon_photo  # 防止被垃圾回收
            icon_label.pack(side='left', padx=(0, 15))
        except:
            pass  # 如果加载失败，则不显示图标

        # 标题信息
        title_info = ttk.Frame(header_frame)
        title_info.pack(side='left', fill='y', expand=True)

        # 标题
        ttk.Label(title_info, text="BOM检查工具使用说明",
               font=('微软雅黑', 16, 'bold'),
               foreground=self.primary_color).pack(anchor='w')

        # 版本信息
        ttk.Label(header_frame, text="V1.8",
               font=('微软雅黑', 11, 'bold')).pack(side='right')

        # 分隔线
        separator = ttk.Separator(help_container, orient='horizontal')
        separator.pack(fill='x', pady=5)

        # 使用选项卡组织内容
        notebook = ttk.Notebook(help_container)
        notebook.pack(fill='both', expand=True, padx=5)

        # 设置选项卡样式
        style = ttk.Style()
        # 直接配置样式，不检查layout_names
        style.configure('TNotebook.Tab', padding=(10, 5), font=('微软雅黑', 10))

        # 快速入门选项卡
        quick_start_frame = ttk.Frame(notebook, padding=15)
        notebook.add(quick_start_frame, text=" 快速入门 ")

        # 详细说明选项卡
        detailed_frame = ttk.Frame(notebook, padding=15)
        notebook.add(detailed_frame, text=" 详细说明 ")

        # 常见问题选项卡
        faq_frame = ttk.Frame(notebook, padding=15)
        notebook.add(faq_frame, text=" 常见问题 ")

        # 更新说明选项卡
        update_frame = ttk.Frame(notebook, padding=15)
        notebook.add(update_frame, text=" 更新说明 ")

        # 快速入门内容
        quick_start_container = ttk.Frame(quick_start_frame)
        quick_start_container.pack(fill='both', expand=True)

        quick_start_text = tk.Text(quick_start_container, wrap=tk.WORD, font=('微软雅黑', 10),
                                highlightthickness=0, borderwidth=0)
        quick_start_text.pack(side='left', fill='both', expand=True)

        quick_scroll = ttk.Scrollbar(quick_start_container, command=quick_start_text.yview)
        quick_scroll.pack(side='right', fill='y')
        quick_start_text.config(yscrollcommand=quick_scroll.set)

        # 设置标签样式
        quick_start_text.tag_configure('title', font=('微软雅黑', 12, 'bold'), foreground=self.primary_color)
        quick_start_text.tag_configure('subtitle', font=('微软雅黑', 11, 'bold'))
        quick_start_text.tag_configure('normal', font=('微软雅黑', 10))
        quick_start_text.tag_configure('highlight', font=('微软雅黑', 10, 'bold'), foreground='#0066CC')

        quick_content = """✨ BOM检查工具 V1.8 ✨

🛠️ 核心功能
══════════════════════════════════════════
BOM检查工具是一款高效的PCB原理图校验解决方案，能够：
✅ 自动对比BOM与原理图中的元器件位号
✅ 高亮标注原理图中的所有位号
✅ 识别并处理替代料组关系
✅ 检测BOM数据一致性问题
✅ 生成详细的检查报告

🚀 快速上手
══════════════════════════════════════════
1️⃣ 准备文件
   • BOM表格 (Excel格式，包含位号列)
   • 原理图 (文本可选择的PDF格式，可选)

2️⃣ 基本操作
   • 【选择BOM文件】上传Excel文件
   • 【选择PDF原理图】上传PDF文件 (可选)
   • 【设置】调整高亮颜色和注解选项
   • 【开始检查】执行位号匹配/BOM检查

3️⃣ 检查结果
   • 查看BOM验证结果和详细信息
   • 生成标注好的PDF输出文件
   • 一键打开结果文件或所在文件夹

🔍 高级提示
══════════════════════════════════════════
💡 能够自动识别替代料组关系
💡 支持BOM表头自定义映射
💡 BOM验证功能可检测多种数据问题
💡 可仅验证BOM，无需提供PDF文件
"""

        quick_start_text.insert(tk.END, quick_content)
        quick_start_text.config(state=tk.DISABLED)  # 设置为只读模式

        # 详细说明内容
        detailed_container = ttk.Frame(detailed_frame)
        detailed_container.pack(fill='both', expand=True)

        detailed_text = tk.Text(detailed_container, wrap=tk.WORD, font=('微软雅黑', 10),
                              highlightthickness=0, borderwidth=0)
        detailed_text.pack(side='left', fill='both', expand=True)

        detailed_scroll = ttk.Scrollbar(detailed_container, command=detailed_text.yview)
        detailed_scroll.pack(side='right', fill='y')
        detailed_text.config(yscrollcommand=detailed_scroll.set)

        # 设置标签样式
        detailed_text.tag_configure('title', font=('微软雅黑', 12, 'bold'), foreground=self.primary_color)
        detailed_text.tag_configure('subtitle', font=('微软雅黑', 11, 'bold'))
        detailed_text.tag_configure('normal', font=('微软雅黑', 10))
        detailed_text.tag_configure('highlight', font=('微软雅黑', 10, 'bold'), foreground='#0066CC')

        detailed_content = """✨ BOM检查工具 ✨
开发者：小航
════════════════════════════════════════

📁 文件规范与数据要求
────────────────────────────────────────

【BOM文件要求】
• 格式：支持Excel文件（.xlsx, .xls）
• 必需列：位号列（Reference Designator）
• 推荐列：料号、厂商料号、数量、描述
• 数据处理：
  ▶ 自动识别表头位置（可跳过项目信息行）
  ▶ 自动处理多位号格式（支持逗号、分号、空格分隔）
  ▶ 智能识别替代料组关系（通过料号项编号）

【原理图要求】
• 格式：仅支持PDF文件
• 文本层：需包含可选择文本的PDF（非图像扫描版）
• 质量：从EDA软件直接导出的PDF效果最佳

────────────────────────────────────────
🛠️ 功能详解
────────────────────────────────────────

1️⃣ BOM数据验证（新增功能）
   • 自动检测替代料组关系
   • 识别重复位号和重复料号
   • 检查位号数量与标称数量是否一致
   • 处理同一替代料组内的位号分配
   • 支持仅BOM检查模式，无需PDF文件

2️⃣ PDF处理增强
   • 智能匹配算法支持多种位号格式
   • 可选添加物料注解（料号、描述等）
   • 生成带标记的PDF便于审阅

3️⃣ 表头映射管理
   • 自定义BOM表头识别规则
   • 记住并应用表头映射设置
   • 支持批量添加表头关键字
   • 灵活应用于各种BOM格式

────────────────────────────────────────
⚙️ 性能优化
────────────────────────────────────────

⚡ 内存使用：优化大型PDF文件处理
⚡ 用户界面：处理过程中保持界面响应
⚡ 结果展示：优化报告格式，提高可读性
"""

        detailed_text.insert(tk.END, detailed_content)
        detailed_text.config(state=tk.DISABLED)  # 设置为只读模式

        # 常见问题内容
        faq_container = ttk.Frame(faq_frame)
        faq_container.pack(fill='both', expand=True)

        faq_text = tk.Text(faq_container, wrap=tk.WORD, font=('微软雅黑', 10),
                         highlightthickness=0, borderwidth=0)
        faq_text.pack(side='left', fill='both', expand=True)

        faq_scroll = ttk.Scrollbar(faq_container, command=faq_text.yview)
        faq_scroll.pack(side='right', fill='y')
        faq_text.config(yscrollcommand=faq_scroll.set)

        # 设置标签样式
        faq_text.tag_configure('title', font=('微软雅黑', 12, 'bold'), foreground=self.primary_color)
        faq_text.tag_configure('subtitle', font=('微软雅黑', 11, 'bold'))
        faq_text.tag_configure('normal', font=('微软雅黑', 10))
        faq_text.tag_configure('highlight', font=('微软雅黑', 10, 'bold'), foreground='#0066CC')

        faq_content = """⚠️ 常见问题与解答
════════════════════════════════════════

❓ BOM验证提示大量重复位号，但实际无重复
────────────────────────────────────────
✅ 可能是替代料组识别问题。替代料使用相同位号是正常的，系统会自动识别替代料
   关系。如果料号编号格式为"基础编号.子编号"（如"1.1"、"1.2"），
   系统会将其视为替代料组。

❓ PDF中大量位号未找到或高亮显示偏移
────────────────────────────────────────
✅ 常见原因：
   • PDF不是文本可选择版本，而是扫描图像
   • 导出PDF时启用了字体替代选项，导致文本层偏移
   • 原理图使用了特殊字体，影响文本识别

   解决方案：从EDA软件重新导出PDF，禁用字体替代选项。

❓ 如何处理分块位号（如U1A/U1B/U1C）？
────────────────────────────────────────
✅ 系统可以识别多种格式的分块位号：
   • 正常连写形式：U1A, U1B, U1C
   • 如BOM中只有U1，但原理图有U1A/U1B，建议在BOM中也使用分块形式。

❓ 位号注解功能使用技巧？
────────────────────────────────────────
✅ 位号注解默认隐藏，在PDF阅读器中点击高亮区域可显示。注解包含：
   • 位号、料号、描述
   • 厂商料号、制造商（如有）
   • 在"设置"中可开启/关闭此功能

❓ BOM表头无法自动识别怎么办？
────────────────────────────────────────
✅ 使用表头映射功能：
   1. 点击"设置" → "编辑BOM表头映射"
   2. 添加您使用的表头关键字
   3. 保存设置后再次运行检查
   系统会记住您的设置，下次自动应用。

❓ 如何最大化提升处理速度？
────────────────────────────────────────
✅ 对于大型PDF文件处理：
   • 确保计算机有足够内存（建议8GB以上）
   • 关闭其他内存密集型应用
   • 性能优先的环境下，可以禁用注解功能

❓ 如何使用仅BOM检查模式？
────────────────────────────────────────
✅ 如果您只需要验证BOM数据而不需检查与原理图的一致性：
   1. 只选择BOM文件，不选择PDF文件
   2. 点击"开始检查"按钮
   3. 系统将自动进入仅BOM检查模式
   4. 检查结果将包含物料清单验证信息，如：
      • 重复料号/厂商料号警告
      • 替代料组识别结果
      • 位号数量与标称不符的警告

════════════════════════════════════════
📞 技术支持：微信 XiaoHang_Sky
"""


        faq_text.insert(tk.END, faq_content)
        faq_text.config(state=tk.DISABLED)  # 设置为只读模式

        # 更新说明内容
        update_container = ttk.Frame(update_frame)
        update_container.pack(fill='both', expand=True)

        update_text = tk.Text(update_container, wrap=tk.WORD, font=('微软雅黑', 10),
                            highlightthickness=0, borderwidth=0)
        update_text.pack(side='left', fill='both', expand=True)

        update_scroll = ttk.Scrollbar(update_container, command=update_text.yview)
        update_scroll.pack(side='right', fill='y')
        update_text.config(yscrollcommand=update_scroll.set)

        # 设置标签样式
        update_text.tag_configure('title', font=('微软雅黑', 12, 'bold'), foreground=self.primary_color)
        update_text.tag_configure('subtitle', font=('微软雅黑', 11, 'bold'))
        update_text.tag_configure('normal', font=('微软雅黑', 10))
        update_text.tag_configure('highlight', font=('微软雅黑', 10, 'bold'), foreground='#0066CC')

        update_content = """✨ BOM检查工具功能更新历史 ✨
════════════════════════════════════════
▶ 版本 V1.8 ◀
发布日期：2025年5月27日

🐞 问题修复
• 优化当位号间隔用中文逗号时，无法正确计算位号数量的问题。
• 优化检查位号数量失效的问题
• 优化厂商物料重复检测失效的问题
────────────────────────────────────────

"""

        update_text.insert(tk.END, update_content)
        update_text.config(state=tk.DISABLED)  # 设置为只读模式

        # 底部按钮区域 - 使用更现代的设计
        button_container = ttk.Frame(help_container)
        button_container.pack(pady=(15, 0))

        # 添加阴影效果的按钮样式
        style = ttk.Style()
        # 直接配置样式，不检查theme_names
        style.configure('AccentButton.TButton',
                      font=('微软雅黑', 10, 'bold'),
                      background=self.primary_color,
                      foreground='white')

        close_btn = ttk.Button(button_container, text="关闭",
                             command=help_window.destroy,
                             style='Primary.TButton',
                             width=10)
        close_btn.pack(pady=5)

        # 居中窗口
        help_window.update_idletasks()
        width = help_window.winfo_width()
        height = help_window.winfo_height()
        x = (help_window.winfo_screenwidth() // 2) - (width // 2)
        y = (help_window.winfo_screenheight() // 2) - (height // 2)
        help_window.geometry(f"+{x}+{y}")

        # 打开时默认选中第一个选项卡
        notebook.select(0)

        # 所有内容加载完成后，显示窗口
        help_window.deiconify()

    # 统计功能已移除

    def select_pdf(self):
        filepath = filedialog.askopenfilename(filetypes=[("PDF文件", "*.pdf")])
        if filepath:
            self.pdf_path = filepath
            self.pdf_label.config(text=filepath)  # 显示完整路径
            self._update_status("已选择PDF文件", self.success_color)

    def parse_bom(self):
        """解析BOM文件，提取位号信息"""

        # 检查路径是否有效
        if not self.bom_path or not os.path.exists(self.bom_path):
            self._show_error("请先选择有效的BOM文件")
            return False

        # 先显示正在解析的消息
        self._update_result_text("正在解析BOM文件...\n", True)

        # 尝试解析BOM文件
        try:
            # 记录开始时间
            t0 = time.time()

            # 状态更新
            self._update_status("解析BOM文件...")

            # 检测文件类型
            file_ext = os.path.splitext(self.bom_path)[1].lower()

            # 使用pandas读取文件
            if file_ext == '.csv':
                # 尝试自动检测编码
                encoding = 'utf-8'  # 默认编码
                try:
                    if chardet is not None:
                        with open(self.bom_path, 'rb') as f:
                            result = chardet.detect(f.read())
                            if result['confidence'] > 0.7:
                                encoding = result['encoding']
                    else:
                        # 如果chardet模块不可用，尝试使用常见编码
                        for enc in ['utf-8', 'gbk', 'gb18030', 'latin1']:
                            try:
                                with open(self.bom_path, 'r', encoding=enc) as f:
                                    f.read(100)  # 尝试读取前100个字符
                                encoding = enc
                                break
                            except UnicodeDecodeError:
                                continue
                except:
                    # 如果检测失败，使用默认编码
                    pass

                # 尝试读取CSV文件
                try:
                    df = pd.read_csv(self.bom_path, encoding=encoding, dtype=str)
                except UnicodeDecodeError:
                    # 如果UTF-8解码失败，尝试其他常见编码
                    success = False
                    for enc in ['gbk', 'gb18030', 'latin1', 'iso-8859-1']:
                        try:
                            df = pd.read_csv(self.bom_path, encoding=enc, dtype=str)
                            success = True
                            break
                        except UnicodeDecodeError:
                            continue

                    if not success:
                        # 如果所有编码都失败
                        raise UnicodeDecodeError("UTF-8", b"", 0, 1, "无法使用常见编码解析CSV文件")

            elif file_ext in ['.xls', '.xlsx']:
                # 尝试读取Excel文件
                try:
                    # 增强Excel文件读取，尝试不同引擎
                    try:
                        # 首先尝试openpyxl引擎（适用于.xlsx文件），强制将所有列转换为字符串类型
                        df = pd.read_excel(self.bom_path, engine='openpyxl', dtype=str)
                    except Exception as e1:
                        print(f"使用openpyxl引擎读取失败: {str(e1)}")
                        try:
                            # 尝试xlrd引擎（适用于.xls文件），强制将所有列转换为字符串类型
                            df = pd.read_excel(self.bom_path, engine='xlrd', dtype=str)
                        except Exception as e2:
                            print(f"使用xlrd引擎读取失败: {str(e2)}")
                            # 使用默认引擎，强制将所有列转换为字符串类型
                            df = pd.read_excel(self.bom_path, dtype=str)

                    print(f"成功读取Excel文件，列数: {len(df.columns)}，前5列: {list(df.columns)[:5]}")

                    # 如果Excel有多个工作表，尝试找到包含位号信息的工作表
                    if 'Reference' not in df.columns and 'PN' not in df.columns:
                        # 添加中文列名检查
                        cn_ref_names = ['位号', '元件位号', '参考号']
                        cn_pn_names = ['料号', '型号', '物料编码']

                        # 检查是否有中文列名
                        has_cn_columns = any(col in df.columns for col in cn_ref_names + cn_pn_names)

                        if not has_cn_columns:
                            print("未在第一个工作表中找到标准列名，尝试检查其他工作表...")
                            # 尝试逐个读取工作表，直到找到包含Reference或PN的那个
                            try:
                                xl = pd.ExcelFile(self.bom_path)
                                sheet_names = xl.sheet_names
                                print(f"Excel文件包含 {len(sheet_names)} 个工作表: {sheet_names}")

                                sheet_found = False
                                for sheet_name in sheet_names:
                                    try:
                                        sheet_df = pd.read_excel(self.bom_path, sheet_name=sheet_name, dtype=str)
                                        print(f"  检查工作表 '{sheet_name}' 的列名")

                                        # 检查标准英文列名
                                        if 'Reference' in sheet_df.columns or 'PN' in sheet_df.columns:
                                            df = sheet_df
                                            sheet_found = True
                                            print(f"  在工作表 '{sheet_name}' 中找到标准列名")
                                            break

                                        # 检查中文列名
                                        if any(col in sheet_df.columns for col in cn_ref_names + cn_pn_names):
                                            df = sheet_df
                                            sheet_found = True
                                            print(f"  在工作表 '{sheet_name}' 中找到中文列名")
                                            break
                                    except Exception as e:
                                        print(f"  读取工作表 '{sheet_name}' 时出错: {str(e)}")
                                        continue
                                                           # 未找到工作表时处理
                                if not sheet_found:
                                    print("未找到包含标准列名的工作表，将尝试智能识别列名")
                                    # 不再显示此警告信息，程序会自动尝试智能识别
                                    # self._update_result_text("⚠️ 注意: 未在Excel工作表中找到标准列名，将尝试智能识别列名\n", False, 'warning')

                            except Exception as e:
                                print(f"读取工作表时出错: {str(e)}")
                                return False
                        else:
                            print("在第一个工作表中找到标准列名")
                except Exception as e:
                    error_msg = str(e)
                    # 使用非阻塞错误显示，允许继续尝试其他方法
                    self._show_error(f"读取Excel文件出错: {error_msg}", show_dialog=False)
                    self._update_result_text("将尝试备用方法继续处理...\n", False, 'info')
            else:
                self._show_error(f"不支持的文件类型: {file_ext}")
                return False

            # 推测标题行
            # 有时Excel文件的实际标题行不是第一行，需要推测
            header_row = self._detect_header_row(self.bom_path)
            if header_row > 0:
                if file_ext == '.csv':
                    df = pd.read_csv(self.bom_path, header=header_row, dtype=str)
                else:
                    df = pd.read_excel(self.bom_path, header=header_row, dtype=str)

            # 检查是否有有效数据
            if df.empty:
                self._update_result_text("BOM文件中没有发现有效数据。\n", False, 'warning')
                return False

            # 识别关键列
            required_columns = ['PN', 'Reference']
            optional_columns = ['Item', 'Quantity', 'MfrPN']

            # 保存原始列名映射
            self.original_columns = {}

            # 检查是否有所需的列
            missing_columns = []

            # 加载已保存的列映射
            # 使用self.bom_header_mapping而不是self.header_mapping
            column_mapping = {}
            if hasattr(self, 'bom_header_mapping') and isinstance(self.bom_header_mapping, dict):
                column_mapping = self.bom_header_mapping

            # 检查并识别所需的列
            for col in required_columns + optional_columns:
                print(f"正在处理列: {col}")

                # 直接精确匹配
                if col in df.columns:
                    self.original_columns[col] = col
                    print(f"  精确匹配成功: {col}")
                    continue

                # 尝试使用用户定义的映射
                found = False

                # 获取当前列类型的所有可能的列名
                col_type = col.lower()
                # 处理特殊的列名映射
                if col_type == 'mfrpn':
                    col_type = 'mfr_pn'  # 映射到配置文件中的键名
                possible_headers = column_mapping.get(col_type, [])
                print(f"  查找列类型 '{col_type}' 的映射，找到 {len(possible_headers)} 个候选: {possible_headers[:5]}...")

                # 逐个检查是否有匹配的列名
                for header in possible_headers:
                    print(f"    检查候选: '{header}'")
                    # 检查列名是否存在于df.columns中
                    if header in df.columns:
                        # 重命名列
                        df = df.rename(columns={header: col})
                        self.original_columns[col] = header
                        found = True
                        print(f"    精确匹配成功: '{header}' -> {col}")
                        break

                    # 如果不是精确匹配，尝试模糊匹配
                    for df_col in df.columns:
                        df_col_str = str(df_col).lower()
                        header_str = str(header).lower()

                        # 检查列名是否包含在df_col中
                        if header_str in df_col_str:
                            df = df.rename(columns={df_col: col})
                            self.original_columns[col] = df_col
                            found = True
                            print(f"    模糊匹配成功: '{df_col}' (包含 '{header}') -> {col}")
                            break

                    if found:
                        break

                if found:
                    continue

                # 尝试模糊匹配
                print(f"  开始模糊匹配...")
                possible_matches = []
                for c in df.columns:
                    # 忽略空列
                    if not c or str(c).strip() == '':
                        continue

                    # 如果已经映射了，跳过
                    if c in [self.original_columns.get(k, None) for k in self.original_columns]:
                        continue

                    c_str = str(c).lower()

                    # 检查是否为常见的列名别名
                    if col == 'PN':
                        pn_aliases = ['pn', 'partnumber', 'part number', 'part-number', 'part_number', 'material',
                                    'materialno', 'material no', 'material-no', 'material_no', 'material number',
                                    'part no', 'partno', 'part-no', 'part_no', 'p/n', 'p-n', 'p_n']
                        if any(alias in c_str for alias in pn_aliases) or 'p' == c_str or 'pn' == c_str:
                            possible_matches.append((c, 1))

                    elif col == 'Reference':
                        ref_aliases = ['reference', 'refdes', 'ref des', 'ref-des', 'ref_des', 'location',
                                    'position', 'designator', 'ref', 'refdes', 'reference designator',
                                    'designators', 'references', 'positions', 'locations', 'componet', 'component']
                        if any(alias in c_str for alias in ref_aliases):
                            possible_matches.append((c, 1))

                    elif col == 'Item':
                        item_aliases = ['item', 'itemno', 'item no', 'item-no', 'item_no', 'item number',
                                      'part', 'component']
                        if any(alias in c_str for alias in item_aliases):
                            possible_matches.append((c, 1))

                    elif col == 'Quantity':
                        qty_aliases = ['quantity', 'qty', 'amount', 'count', 'num', 'number', 'q', 'quant']
                        if any(alias in c_str for alias in qty_aliases):
                            possible_matches.append((c, 1))

                    elif col == 'MfrPN':
                        print(f"    检查MfrPN列，当前列名: '{c}' (小写: '{c_str}')")
                        mfrpn_aliases = [
                            # 英文别名
                            'mfrpn', 'manufacturer part number', 'mfr part number', 'mfr-pn', 'mfr_pn',
                            'manufacturer partnumber', 'mpn', 'm/p/n', 'm-p-n', 'vendor pn', 'vendorpn',
                            'factory pn', 'factory-pn', 'factory_pn', 'manufacturer pn', 'mfrpartnumber',
                            'manuf. p/n', 'mfg p/n', 'factory part number', 'mfg part number',
                            'supplier pn', 'supplier part number', 'vendor part number', 'vendor partnumber',
                            'mfr part no', 'manufacturer part no', 'mfg part no', 'supplier part no',
                            'vendor part no', 'factory part no', 'original part number', 'oem part number',
                            'component part number', 'device part number', 'chip part number',
                            # 特殊格式
                            'manufacturer p/n', 'manuf p/n', 'mfr p/n', 'mfg p/n',
                            # 中文别名
                            '厂商料号', '厂家料号', '制造商料号', '供应商料号', '原厂料号', '厂商型号',
                            '厂家型号', '制造商型号', '供应商型号', '原厂型号', '厂商编号', '厂家编号',
                            '制造商编号', '供应商编号', '原厂编号', '厂商物料号', '厂家物料号',
                            '制造商物料号', '供应商物料号', '原厂物料号', '器件料号', '器件型号',
                            '芯片料号', '芯片型号', '元件料号', '元件型号', '零件料号', '零件型号'
                        ]
                        # 检查是否匹配任何别名
                        for alias in mfrpn_aliases:
                            if alias in c_str:
                                print(f"      匹配到别名: '{alias}' 在 '{c_str}' 中")
                                possible_matches.append((c, 1))
                                break

                # 如果有匹配项，选择第一个
                if possible_matches:
                    best_match = sorted(possible_matches, key=lambda x: x[1], reverse=True)[0][0]
                    df = df.rename(columns={best_match: col})
                    self.original_columns[col] = best_match
                    print(f"  模糊匹配成功: '{best_match}' -> {col}")
                else:
                    print(f"  未找到匹配项")
                    # 如果没有匹配项，对于必需列标记为缺失
                    if col in required_columns:
                        missing_columns.append(col)

            # 如果缺少必要的列，请用户手动指定
            if missing_columns:
                for col in missing_columns:
                    selected_col_info = self.ask_user_for_column(list(df.columns), col.lower())
                    # 检查selected_col_info的类型和内容
                    if isinstance(selected_col_info, tuple) and len(selected_col_info) == 2:
                        selected_col, remember = selected_col_info
                    else:
                        selected_col = selected_col_info  # 向后兼容
                        remember = False

                    if selected_col:
                        df = df.rename(columns={selected_col: col})
                        self.original_columns[col] = selected_col

                        # 如果用户选择记住这个选择，则添加到表头映射并保存
                        if remember:
                            # 将列名转换为小写以便于匹配
                            column_type = col.lower()
                            # 处理特殊的列名映射
                            if column_type == 'mfrpn':
                                column_type = 'mfr_pn'  # 映射到配置文件中的键名

                            # 定义列类型的中文名称
                            column_titles = {
                                'reference': "位号列",
                                'pn': "料号列",
                                'description': "描述列",
                                'manufacturer': "制造商列",
                                'mfr_pn': "厂商料号列",
                                'part': "元件类型列",
                                'quantity': "数量列"
                            }

                            # 确保表头映射字典中有对应的键
                            if column_type not in self.bom_header_mapping:
                                self.bom_header_mapping[column_type] = []

                            # 如果选择的列名不在映射列表中，则添加
                            if selected_col not in self.bom_header_mapping[column_type]:
                                self.bom_header_mapping[column_type].append(selected_col)
                                # 保存更新后的映射
                                self.save_header_mapping()
                                self._update_result_text(f"已将 '{selected_col}' 添加到{column_titles.get(column_type, column_type)}映射\n", False, 'info')
                    else:
                        self._show_error(f"无法继续，因为缺少必要的列: {', '.join(missing_columns)}")
                        return False

            # 输出表头映射结果，特别关注厂商料号列
            print(f"表头映射完成，映射后的列名: {list(df.columns)}")
            if 'MfrPN' in df.columns:
                print(f"成功识别厂商料号列: MfrPN")
                # 显示前几个厂商料号样本
                sample_mfrpns = df['MfrPN'].dropna().head(5).tolist()
                print(f"厂商料号样本: {sample_mfrpns}")
            else:
                print("警告: 未能识别厂商料号列，可能的原因:")
                print("  1. BOM文件中没有厂商料号列")
                print("  2. 厂商料号列的表头名称不在预设的别名列表中")
                print(f"  3. 当前BOM文件的列名: {list(df.columns)}")
                print("  建议检查表头映射设置或手动配置厂商料号列映射")

            # 验证BOM数据
            self.validation_results, self.ref_list = self._validate_bom_data(df)

            # 保存处理后的DataFrame和位号到料号的映射
            self.bom_df = df

            # 更新解析状态
            parsing_time = time.time() - t0
            self._update_status(f"BOM解析完成，用时 {parsing_time:.2f}秒")

            # 保存解析结果
            self._update_result_text(f"✓ BOM解析完成，处理了 {len(df)} 行数据，用时 {parsing_time:.2f}秒\n", False, 'success')
            self._update_result_text(f"  提取到 {len(self.ref_list)} 个唯一位号\n", False, 'info')

            # 显示验证结果
            self._display_validation_results(self.validation_results)

            return True

        except Exception as e:
            error_msg = str(e)
            self._show_error(f"解析BOM文件时出错: {error_msg}")
            # 记录详细错误到日志
            logging.error(f"解析BOM出错: {error_msg}", exc_info=True)
            return False

    def _validate_bom_data(self, bom_data):
        """验证BOM数据，检查重复料号、重复位号、替代料关系和位号数量"""
        # 结果存储
        validation_results = {
            'duplicate_pns': [],         # 重复料号
            'duplicate_refs': [],        # 重复位号（非替代料）
            'duplicate_mfrpns': [],      # 重复厂商料号
            'quantity_mismatch': [],     # 位号数量不匹配
            'alternative_parts': {}      # 替代料关系
        }

        # 预处理：清理数据并确保所有列都是字符串类型
        bom_data = bom_data.fillna('')

        # 确保关键列是字符串类型
        for col in ['PN', 'Reference', 'Item', 'MfrPN']:
            if col in bom_data.columns:
                bom_data[col] = bom_data[col].astype(str)
                # 将"nan"字符串替换为空字符串
                bom_data[col] = bom_data[col].replace('nan', '')

        # 用于存储处理后的所有唯一位号
        processed_refs = set()

        # 识别替代料关系 - 这是关键点
        alternative_groups = {}  # 存储替代料组

        # 创建一个料号到Item的映射
        pn_to_item_map = {}
        # 创建一个Item到PN的映射（一个Item可能对应多个PN）
        item_to_pns_map = {}

        # 先创建料号到Item的映射，方便后续快速查找
        if 'PN' in bom_data.columns and 'Item' in bom_data.columns:
            for _, row in bom_data.iterrows():
                pn = str(row.get('PN', '')).strip()
                item = str(row.get('Item', '')).strip()
                if pn and item:
                    pn_to_item_map[pn] = item
                    if item not in item_to_pns_map:
                        item_to_pns_map[item] = []
                    if pn not in item_to_pns_map[item]:
                        item_to_pns_map[item].append(pn)

        # 保存为类属性，以便在结果显示时使用
        self.pn_to_item_map = pn_to_item_map

        # 1. 识别替代料关系 - 基于Item格式（改进版本）
        if 'Item' in bom_data.columns:
            # 收集所有带有替代格式的Item (例如: 40.1, 40.2, ...)
            items_with_parts = {}
            for item in item_to_pns_map.keys():
                if '.' in item:
                    try:
                        # 尝试拆分基础编号和替代号，例如从"40.1"提取"40"
                        parts = item.split('.')
                        if len(parts) == 2:  # 确保格式正确（只有一个小数点）
                            base_item = parts[0]
                            sub_item = parts[1]
                            # 确保基础编号和子编号都是数字或合理的格式
                            if (base_item.isdigit() or base_item.replace('R', '').replace('C', '').replace('L', '').isdigit()) and sub_item.isdigit():
                                if base_item not in items_with_parts:
                                    items_with_parts[base_item] = []
                                items_with_parts[base_item].append(item)
                    except Exception:
                        pass  # 忽略格式不符合的项

            # 整理并保存替代料组信息
            for base_item, items in items_with_parts.items():
                if len(items) > 1:  # 只有当有多个替代项时才算作替代料组
                    alternative_groups[base_item] = items

            # 更新validation_results
            validation_results['alternative_parts'] = alternative_groups

        # 2. 检查重复料号
        if 'PN' in bom_data.columns:
            # 确保PN列被转换为字符串类型
            bom_data['PN'] = bom_data['PN'].astype(str)
            valid_pns = bom_data[bom_data['PN'].notna() & (bom_data['PN'].str.strip() != '')]['PN'].str.strip()
            pn_counts = valid_pns.value_counts()
            duplicate_pns = [pn for pn, count in pn_counts.items() if count > 1 and pn]

            if duplicate_pns:
                validation_results['duplicate_pns'] = duplicate_pns

        # 检查重复厂商料号 - 增强版本
        if 'MfrPN' in bom_data.columns:
            print(f"开始检查重复厂商料号，MfrPN列存在，共有 {len(bom_data)} 行数据")

            # 确保MfrPN列被转换为字符串类型
            bom_data['MfrPN'] = bom_data['MfrPN'].astype(str)

            # 更严格的数据清洗：去除空值、nan字符串、纯空格等
            valid_mfrpns_mask = (
                bom_data['MfrPN'].notna() &
                (bom_data['MfrPN'].str.strip() != '') &
                (bom_data['MfrPN'].str.lower() != 'nan') &
                (bom_data['MfrPN'].str.lower() != 'none') &
                (bom_data['MfrPN'].str.strip() != '-') &
                (bom_data['MfrPN'].str.strip() != 'n/a')
            )

            # 提取有效的厂商料号并去除前后空格
            valid_mfrpns = bom_data[valid_mfrpns_mask]['MfrPN'].str.strip()
            print(f"有效厂商料号数量: {len(valid_mfrpns)}")

            # 统计厂商料号出现次数
            mfrpn_counts = valid_mfrpns.value_counts()
            print(f"唯一厂商料号数量: {len(mfrpn_counts)}")

            # 找出重复的厂商料号（出现次数大于1）
            duplicate_mfrpns_basic = [mfrpn for mfrpn, count in mfrpn_counts.items() if count > 1 and mfrpn]
            print(f"发现重复厂商料号数量: {len(duplicate_mfrpns_basic)}")

            # 进一步分析重复厂商料号，考虑替代料关系
            duplicate_mfrpns_detailed = []

            if duplicate_mfrpns_basic:
                print("开始详细分析重复厂商料号...")
                for mfrpn in duplicate_mfrpns_basic:
                    print(f"  分析厂商料号: {mfrpn}")

                    # 获取使用该厂商料号的所有行
                    mfrpn_rows = bom_data[bom_data['MfrPN'].str.strip() == mfrpn]
                    print(f"    该厂商料号出现在 {len(mfrpn_rows)} 行中")

                    # 收集相关的料号和Item信息
                    related_pns = []
                    related_items = []
                    related_refs = []

                    for _, row in mfrpn_rows.iterrows():
                        pn = str(row.get('PN', '')).strip()
                        item = str(row.get('Item', '')).strip()
                        ref = str(row.get('Reference', '')).strip()

                        if pn and pn not in related_pns:
                            related_pns.append(pn)
                        if item and item not in related_items:
                            related_items.append(item)
                        if ref:
                            # 拆分位号
                            refs = [r.strip() for r in re.split(r'[,;，；\s]+', ref) if r.strip()]
                            related_refs.extend(refs)

                    print(f"    相关料号: {related_pns}")
                    print(f"    相关Item: {related_items}")

                    # 检查是否为替代料关系
                    is_alternative_group = False
                    if len(related_items) > 1:
                        # 检查Item是否符合替代料格式（如40.1, 40.2等）
                        items_with_format = [item for item in related_items if '.' in item]
                        print(f"    带格式的Item: {items_with_format}")

                        # 只有当所有Item都有格式（包含小数点）时才进行替代料判断
                        if len(items_with_format) == len(related_items) and len(items_with_format) > 1:
                            try:
                                base_items = set()
                                for item in items_with_format:
                                    base_item = item.split('.')[0]
                                    base_items.add(base_item)
                                print(f"    基础Item: {base_items}")
                                # 只有当所有Item都有相同的基础编号时，才认为是替代料组
                                if len(base_items) == 1:
                                    is_alternative_group = True
                                    print(f"    判定为替代料组")
                                else:
                                    print(f"    基础Item不同，不是替代料组")
                            except Exception:
                                print(f"    解析Item格式出错，不是替代料组")
                        else:
                            print(f"    不是所有Item都有格式或数量不足，不是替代料组")

                    # 记录重复厂商料号的详细信息
                    duplicate_info = {
                        'mfrpn': mfrpn,
                        'count': len(mfrpn_rows),
                        'related_pns': related_pns,
                        'related_items': related_items,
                        'related_refs': list(set(related_refs)),  # 去重
                        'is_alternative': is_alternative_group
                    }

                    duplicate_mfrpns_detailed.append(duplicate_info)

                # 保存详细的重复厂商料号信息
                validation_results['duplicate_mfrpns'] = duplicate_mfrpns_basic
                validation_results['duplicate_mfrpns_detailed'] = duplicate_mfrpns_detailed
                print(f"重复厂商料号检查完成，共发现 {len(duplicate_mfrpns_basic)} 个重复")
            else:
                print("未发现重复厂商料号")
        else:
            print("BOM数据中未找到MfrPN列，跳过重复厂商料号检查")

        # 创建一个位号到物料列表的映射
        ref_to_pns_map = {}

        # 处理位号和检查数量匹配
        if 'Reference' in bom_data.columns:
            for idx, row in bom_data.iterrows():
                refs_str = str(row['Reference']).strip()
                pn = str(row.get('PN', '')).strip()
                item = str(row.get('Item', '')).strip()
                quantity = row.get('Quantity', '')

                # 如果位号或料号为空，跳过
                if not refs_str or not pn:
                    continue

                # 将'nan'、'NaN'等值视为空字符串
                if isinstance(refs_str, str) and refs_str.lower() == 'nan':
                    continue
                if isinstance(pn, str) and pn.lower() == 'nan':
                    continue

                # 拆分位号，支持不同分隔符（包括中文分隔符），简化逻辑与V1.3版本保持一致
                refs = [ref.strip() for ref in re.split(r'[,;，；\s]+', refs_str) if ref.strip()]

                # 检查位号数量是否匹配Quantity列
                try:
                    qty = int(float(quantity)) if quantity and not isinstance(quantity, bool) else 0
                    if qty > 0 and len(refs) != qty:
                        validation_results['quantity_mismatch'].append({
                            'item': item,
                            'pn': pn,
                            'expected': qty,
                            'actual': len(refs)
                        })
                except (ValueError, TypeError):
                    # 数量无法转换为数字，跳过检查
                    pass

                # 将所有位号添加到位号到物料的映射
                for ref in refs:
                    processed_refs.add(ref)  # 添加到处理过的位号集合
                    if ref not in ref_to_pns_map:
                        ref_to_pns_map[ref] = []
                    if pn not in ref_to_pns_map[ref]:
                        ref_to_pns_map[ref].append(pn)

                # 检查同一物料内的位号重复
                ref_count = {}
                for ref in refs:
                    if ref in ref_count:
                        ref_count[ref] += 1
                    else:
                        ref_count[ref] = 1

                # 检查是否有重复位号
                for ref, count in ref_count.items():
                    if count > 1:
                        validation_results['duplicate_refs'].append({
                            'ref': ref,
                            'pn': pn,
                            'item': item,
                            'count': count
                        })

            # 检查不同物料之间的位号重复（非替代料关系）
            different_part_duplicates = []

            # 遍历每个位号及其对应的物料列表
            for ref, pns in ref_to_pns_map.items():
                # 如果位号对应多个物料，需要检查它们是否是替代关系
                if len(pns) > 1:
                    # 标记是否为替代料关系
                    is_alternative = False

                    # 方法1: 基于Item格式判断替代料关系
                    items = [pn_to_item_map.get(pn, '') for pn in pns if pn in pn_to_item_map]
                    items_with_format = [item for item in items if '.' in item]

                    if items_with_format:
                        # 尝试提取基础编号
                        try:
                            base_items = set()
                            for item in items_with_format:
                                base_item = item.split('.')[0]
                                base_items.add(base_item)

                                                       # 如果所有有格式的Item都有相同的基础编号，认为是替代料组
                            if len(base_items) == 1 and len(base_items) == len(items_with_format):
                                is_alternative = True
                        except Exception:
                            pass  # 忽略任何解析错误

                        # 方法2: 检查替代料组字典
                        if not is_alternative:
                            # 获取每个料号对应的基础组
                            item_to_base_groups = {}
                            for pn in pns:
                                if pn in pn_to_item_map:
                                    item = pn_to_item_map[pn]
                                    # 检查该Item是否属于某个替代料组
                                    for base_item, alt_items in alternative_groups.items():
                                        if item in alt_items:
                                            if pn not in item_to_base_groups:
                                                item_to_base_groups[pn] = set()
                                            item_to_base_groups[pn].add(base_item)

                            # 检查所有料号是否属于同一个替代料组
                            if item_to_base_groups:
                                # 收集所有基础组
                                all_base_groups = set()
                                for groups in item_to_base_groups.values():
                                    all_base_groups.update(groups)

                                # 检查是否有共同的基础组 且 所有料号都属于替代料组
                                if len(all_base_groups) == 1 and len(item_to_base_groups) == len(pns):
                                    is_alternative = True

                    # 如果不是替代料关系，则报告为重复位号
                    if not is_alternative:
                        different_part_duplicates.append({
                            'ref': ref,
                            'pns': pns,
                            'items': [pn_to_item_map.get(pn, '') for pn in pns],
                            'count': len(pns),
                            'same_part': False  # 标记为不同物料间的重复
                        })

            # 添加到验证结果中
            validation_results['duplicate_refs'].extend(different_part_duplicates)

        return validation_results, list(processed_refs)

    def _display_validation_results(self, validation_results):
        """显示BOM验证结果"""
        self.message_queue.put(("text", "=== BOM验证结果 ===\n", False))

        # 显示重复料号警告
        if validation_results['duplicate_pns']:
            self.message_queue.put(("text", f"警告: 发现 {len(validation_results['duplicate_pns'])} 个重复的料号\n", False))
            self.message_queue.put(("text", "重复的料号包括:\n", False))
            for i, pn in enumerate(validation_results['duplicate_pns']):
                if i < 5:  # 最多显示5个
                    self.message_queue.put(("text", f"   - {pn}\n", False))
            if len(validation_results['duplicate_pns']) > 5:
                self.message_queue.put(("text", f"   - ... 等 {len(validation_results['duplicate_pns'])-5} 个\n", False))

        # 显示重复厂商料号警告 - 增强版本
        if 'duplicate_mfrpns' in validation_results and validation_results['duplicate_mfrpns']:
            self.message_queue.put(("text", f"警告: 发现 {len(validation_results['duplicate_mfrpns'])} 个重复的厂商料号\n", False))

            # 如果有详细信息，显示详细分析
            if 'duplicate_mfrpns_detailed' in validation_results:
                detailed_info = validation_results['duplicate_mfrpns_detailed']

                # 分类显示：替代料组 vs 真正的重复
                alternative_duplicates = [info for info in detailed_info if info.get('is_alternative', False)]
                real_duplicates = [info for info in detailed_info if not info.get('is_alternative', False)]

                if alternative_duplicates:
                    self.message_queue.put(("text", f"  • 替代料组中的厂商料号重复 ({len(alternative_duplicates)}个):\n", False))
                    for info in alternative_duplicates[:3]:  # 最多显示3个
                        mfrpn = info['mfrpn']
                        related_items = info['related_items']
                        self.message_queue.put(("text", f"    - {mfrpn} (替代料组: {', '.join(related_items)})\n", False))
                    if len(alternative_duplicates) > 3:
                        self.message_queue.put(("text", f"    - ... 等 {len(alternative_duplicates)-3} 个\n", False))

                if real_duplicates:
                    self.message_queue.put(("text", f"  • 非替代料的厂商料号重复 ({len(real_duplicates)}个):\n", False))
                    for info in real_duplicates:  # 显示所有真正的重复
                        mfrpn = info['mfrpn']
                        related_pns = info['related_pns']
                        count = info['count']
                        self.message_queue.put(("text", f"    - {mfrpn} (出现{count}次，涉及料号: {', '.join(related_pns)})\n", False))
            else:
                # 如果没有详细信息，使用原来的简单显示方式
                self.message_queue.put(("text", "重复的厂商料号包括:\n", False))
                for i, mfrpn in enumerate(validation_results['duplicate_mfrpns']):
                    if i < 5:  # 最多显示5个
                        self.message_queue.put(("text", f"   - {mfrpn}\n", False))
                if len(validation_results['duplicate_mfrpns']) > 5:
                    self.message_queue.put(("text", f"   - ... 等 {len(validation_results['duplicate_mfrpns'])-5} 个\n", False))

        # 注释掉替代料关系的显示
        # # 显示替代料关系
        # if validation_results.get('alternative_parts') and isinstance(validation_results['alternative_parts'], dict):
        #     alt_parts = validation_results['alternative_parts']
        #     self.message_queue.put(("text", f"信息: 识别到 {len(alt_parts)} 组替代料关系\n", False))
        #
        #     # 只显示前3组替代料
        #     count = 0
        #     for base_item, items in alt_parts.items():
        #         if count < 3:
        #             # 收集对应的料号信息
        #             pns_list = []
        #             for item in items:
        #                 if item in self.item_to_pns_map:
        #                     pns_list.extend(self.item_to_pns_map[item])
        #
        #             self.message_queue.put(("text", f"  • 物料项 {base_item} 的替代料: {', '.join(items)}\n", False))
        #             if pns_list:
        #                 self.message_queue.put(("text", f"    对应料号: {', '.join(pns_list)}\n", False))
        #         count += 1
        #
        #     if len(alt_parts) > 3:
        #         self.message_queue.put(("text", f"    ... 等 {len(alt_parts)-3} 组\n", False))

        # 显示重复位号警告
        if validation_results['duplicate_refs']:
            # 分类重复位号
            same_part_dups = []  # 同一料号内重复
            diff_part_dups = []  # 不同料号间重复

            for ref_info in validation_results['duplicate_refs']:
                # 检查是否存在pns字段且包含多个料号
                if 'pns' in ref_info and isinstance(ref_info['pns'], list) and len(ref_info['pns']) > 1:
                    diff_part_dups.append(ref_info)
                # 兼容旧格式
                elif 'pn' in ref_info:
                    same_part_dups.append(ref_info)

            # 显示同一物料内重复的位号
            if same_part_dups:
                self.message_queue.put(("text", f"  • 同一物料内重复的位号 ({len(same_part_dups)}个):\n", False))
                for ref_info in same_part_dups[:3]:  # 最多显示3个
                    ref = ref_info['ref']
                    # 检查字段存在性，兼容不同格式
                    if 'pn' in ref_info:
                        pn = ref_info['pn']
                    elif 'pns' in ref_info and ref_info['pns']:
                        pn = ref_info['pns'][0]
                    else:
                        pn = "未知物料"
                    self.message_queue.put(("text", f"    - {ref} 在物料 {pn} 中重复出现\n", False))
                if len(same_part_dups) > 3:
                    self.message_queue.put(("text", f"    - ... 等 {len(same_part_dups)-3} 个\n", False))

            # 显示不同物料间重复的位号
            if diff_part_dups:
                self.message_queue.put(("text", f"  • 不同物料间重复的位号 ({len(diff_part_dups)}个):\n", False))
                for ref_info in diff_part_dups:  # 显示所有重复位号
                    ref = ref_info['ref']
                    pns = ref_info['pns']
                    # 更清晰地显示位号重复问题
                    if len(pns) > 2:
                        self.message_queue.put(("text", f"    - 位号 '{ref}' 在 {len(pns)} 个不同物料内重复出现:\n", False))
                        # 显示每个物料
                        for i, pn in enumerate(pns, 1):
                            self.message_queue.put(("text", f"      {i}. PN: {pn}\n", False))
                    else:
                        # 两个物料的情况
                        self.message_queue.put(("text", f"    - 位号 '{ref}' 在不同物料间重复: {pns[0]} 和 {pns[1]}\n", False))

        # 显示位号数量不匹配警告 - 改进此部分的显示格式
        if validation_results['quantity_mismatch']:
            self.message_queue.put(("text", f"警告: 发现 {len(validation_results['quantity_mismatch'])} 处位号数量不匹配\n", False))

            # 添加显示详细的数量不匹配信息，改进格式
            self.message_queue.put(("text", "位号数量不匹配详情:\n", False))
            for mismatch in validation_results['quantity_mismatch']:  # 显示所有不匹配项
                item_text = f"物料项: {mismatch['item']}" if 'item' in mismatch and mismatch['item'] else ""
                self.message_queue.put(("text", f"   - 料号: {mismatch['pn']} {item_text}\n", False))
                self.message_queue.put(("text", f"     BOM中数量: {mismatch['expected']}, 实际位号数: {mismatch['actual']}\n", False))

        # 如果没有问题，显示验证通过
        if not (validation_results['duplicate_pns'] or validation_results.get('duplicate_mfrpns') or
                validation_results['duplicate_refs'] or validation_results['quantity_mismatch']):
            self.message_queue.put(("text", "BOM验证通过，未发现问题\n", False))

        self.message_queue.put(("text", "===================\n", False))

    def _run_in_thread(self):
        """优化的后台处理线程，支持BOM数据验证和替代料处理"""
        total_start_time = time.time()
        stats = {}

        try:
            import traceback

            # 输出开始信息、当前时间和内存使用
            print(f"\n====== 开始处理 ======")
            print(f"时间: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())}")

            # 记录当前进程的内存使用情况
            try:
                if psutil:
                    process = psutil.Process(os.getpid())
                    memory_info = process.memory_info()
                    print(f"内存使用: {memory_info.rss / 1024 / 1024:.2f} MB")
            except Exception as e:
                print(f"获取内存信息出错: {str(e)}")

            self.processing = True

            # 将UI更新任务放入队列
            self.message_queue.put(("status", "正在处理中...", self.primary_color))

            # 计时：BOM解析开始
            bom_start_time = time.time()
            parse_success = self.parse_bom()  # 现在返回布尔值表示成功或失败
            # 计时：BOM解析结束
            bom_end_time = time.time()
            stats['bom_time'] = bom_end_time - bom_start_time

            # 如果解析失败或没有提取到有效位号，显示错误并停止处理
            if not parse_success or not hasattr(self, 'ref_list') or not self.ref_list:
                self.message_queue.put(("warning", "未能从BOM中提取有效位号，请检查BOM格式和表头设置。"))
                self.processing = False
                return

            # 使用self.ref_list代替之前的ref_designators
            ref_designators = self.ref_list

            # 检查是否存在PDF文件
            if not self.pdf_path:
                # 仅BOM检查模式
                self.message_queue.put(("text", "【仅BOM检查模式】\n", False))
                self.message_queue.put(("text", f"从BOM中提取了 {len(ref_designators)} 个唯一位号\n", False))

                # 计时：总处理结束
                total_end_time = time.time()
                stats['total_time'] = total_end_time - total_start_time

                # 创建一个仅包含BOM分析结果的字典
                result = {
                    'total_refs': len(ref_designators),
                    'found_refs': [],
                    'not_found_refs': [],
                    'bom_time': stats['bom_time'],
                    'total_time': stats['total_time'],
                    'mode': 'bom_only'
                }

                # 输出统计信息
                print(f"\n====== 仅BOM处理完成 ======")
                print(f"BOM解析时间: {stats['bom_time']:.2f}秒")
                print(f"总处理时间: {stats['total_time']:.2f}秒")

                # 显示结果
                self.message_queue.put(("show_results", result))
                return

            # 执行PDF处理
            pdf_start_time = time.time()

            # 使用单独进程处理PDF
            try:
                result = self.process_pdf(ref_designators)
                # 计时：PDF处理结束
                pdf_end_time = time.time()
                stats['pdf_time'] = pdf_end_time - pdf_start_time

                # 计时：总处理结束 - 移到这里确保在PDF处理后计算
                total_end_time = time.time()
                stats['total_time'] = total_end_time - total_start_time

                # 输出统计信息
                print(f"\n====== 处理完成 ======")
                print(f"BOM解析时间: {stats['bom_time']:.2f}秒")
                print(f"PDF处理时间: {stats['pdf_time']:.2f}秒")
                print(f"总处理时间: {stats['total_time']:.2f}秒")

                # 如果成功处理，保存输出PDF路径
                if 'output_pdf' in result and result['output_pdf']:
                    self.output_pdf_path = result['output_pdf']

                # 添加时间统计到结果字典
                for k, v in stats.items():
                    result[k] = v

                # 显示结果
                self.message_queue.put(("show_results", result))
            except Exception as e:
                # 处理PDF过程中的错误
                error_msg = f"PDF处理错误: {str(e)}"
                print(error_msg)
                traceback.print_exc()
                # 将错误信息发送到主线程进行处理
                # 在主线程中会调用_show_error方法，该方法会将错误信息转换为中文
                self.message_queue.put(("show_error", error_msg))

        except Exception as e:
            # 捕获所有异常
            error_msg = f"处理过程发生异常: {str(e)}"
            print(error_msg)
            traceback.print_exc()
            # 将错误信息发送到主线程进行处理
            # 在主线程中会调用_show_error方法，该方法会将错误信息转换为中文
            self.message_queue.put(("show_error", error_msg))

        finally:
            # 恢复界面状态
            self.message_queue.put(("update_button", {"state": 'normal'}))
            self.processing = False

    def _update_result_text(self, text, clear=False, tag=None):
        """更新结果文本框，使用after方法确保在主线程中执行UI更新"""
        # 使用after(0...)确保UI更新在主线程的空闲时间执行，避免阻塞
        # 使用lambda避免在调用时就执行函数
        self.master.after(0, lambda t=text, c=clear, tg=tag: self._do_update_result_text(t, c, tg))

    def _do_update_result_text(self, text, clear=False, tag=None):
        """实际执行文本更新的函数"""
        try:
            self.result_text.config(state='normal')
            if clear:
                self.result_text.delete(1.0, 'end')
            # 如果指定了标签，则应用该标签
            if tag:
                self.result_text.insert('end', text, tag)
            else:
                self.result_text.insert('end', text)
            self.result_text.see('1.0')  # 滚动到顶部
            self.result_text.config(state='disabled')
        except Exception as e:
            print(f"更新结果文本出错: {str(e)}")

    def _show_results(self, results):
        """显示处理结果"""
        if 'error' in results and results['error']:
            # 显示错误信息
            self._show_error(results['error'])
            return

        # 获取统计数据
        stats = {}
        for key in ['bom_time', 'pdf_time', 'total_time']:
            stats[key] = results.get(key, 0)

        # 计算时间
        pdf_minutes = int(stats.get('pdf_time', 0) / 60)
        pdf_seconds = stats.get('pdf_time', 0) % 60
        total_minutes = int(stats.get('total_time', 0) / 60)
        total_seconds = stats.get('total_time', 0) % 60
        bom_seconds = int(stats.get('bom_time', 0))
        bom_ms = int((stats.get('bom_time', 0) - bom_seconds) * 1000)

        # 保存输出PDF路径
        if 'output_pdf' in results and results['output_pdf']:
            self.output_pdf_path = results['output_pdf']

        # 清除旧的结果按钮（如果有）
        for widget in self.pdf_buttons_frame.winfo_children():
            widget.destroy()

        # 检查是否为仅BOM检查模式
        is_bom_only_mode = results.get('mode') == 'bom_only'

        # 如果不是仅BOM模式，添加PDF相关按钮
        if not is_bom_only_mode:
            # 添加简短标签和按钮
            ttk.Label(self.pdf_buttons_frame, text="PDF文件:").pack(side='left', padx=(0, 2))

            # 打开文件按钮 - 极小化
            ttk.Button(self.pdf_buttons_frame, text="打开",
                     command=self._open_output_pdf, width=4).pack(side='left', padx=1)

            # 打开文件夹按钮 - 极小化
            ttk.Button(self.pdf_buttons_frame, text="文件夹",
                     command=self._open_output_folder, width=5).pack(side='left', padx=1)

        # 清空之前的所有文本
        self._update_result_text("", True)

        # ===== 1. 标题和总结部分 =====
        header = (
            f"{'='*50}\n"
        )

        if is_bom_only_mode:
            header += f"📋  BOM检查结果报告 (仅BOM模式)\n"
        else:
            header += f"📋  BOM检查结果报告\n"

        header += f"{'='*50}\n\n"
        self._update_result_text(header, False, 'header')

        # ===== 2. 基本统计部分 =====
        stats_header = f"【基本统计】\n{'-'*30}\n"
        self._update_result_text(stats_header, False, 'subheader')

        # 根据模式显示不同的基本统计信息
        if is_bom_only_mode:
            basic_stats = (
                f"• BOM中位号总数: {results['total_refs']}\n\n"
            )
        else:
            found_percent = int((results['found'] / results['total']) * 100) if results['total'] > 0 else 0

            basic_stats = (
                f"• BOM中位号总数: {results['total']}\n"
                f"• 在PDF中找到: {results['found']} ({found_percent}%)\n"
                f"• 在PDF中未找到: {len(results['not_found'])}\n\n"
            )

        self._update_result_text(basic_stats, False, 'info')

        # ===== 3. BOM验证结果部分 =====
        if hasattr(self, 'validation_results'):
            validation_header = f"【BOM验证结果】\n{'-'*30}\n"
            self._update_result_text(validation_header, False, 'subheader')

            has_validation_issues = False

            # 替代料关系信息
            if self.validation_results.get('alternative_parts'):
                has_validation_issues = True
                # 不显示替代料关系信息，但仍保留计数逻辑
                # self._update_result_text(f"✓ 识别到 {len(self.validation_results['alternative_parts'])} 组替代料关系\n", False, 'success')
                # 显示部分替代料组详情
                # count = 0
                # for base_item, items in self.validation_results['alternative_parts'].items():
                #     if count < 3:  # 只显示前3组
                #         self._update_result_text(f"   - 组 {base_item}: {', '.join(items)}\n", False, 'info')
                #     count += 1
                # if count > 3:
                #     self._update_result_text(f"   - ... 更多替代料组 ...\n", False, 'info')

            # 检查是否有BOM验证问题
            validation_issues = []

            # 重复料号警告
            if self.validation_results.get('duplicate_pns'):
                has_validation_issues = True
                self._update_result_text(f"\n⚠️ 警告: 发现 {len(self.validation_results['duplicate_pns'])} 个重复料号\n", False, 'warning')
                # 显示重复料号
                if len(self.validation_results['duplicate_pns']) > 0:
                    pns_to_show = self.validation_results['duplicate_pns'][:5]  # 只显示前5个
                    self._update_result_text(f"   - {', '.join(pns_to_show)}" +
                                          (f" ... 等" if len(self.validation_results['duplicate_pns']) > 5 else "") +
                                          "\n", False, 'info')

            # 重复厂商料号警告 - 增强版本
            if self.validation_results.get('duplicate_mfrpns'):
                has_validation_issues = True
                self._update_result_text(f"\n⚠️ 警告: 发现 {len(self.validation_results['duplicate_mfrpns'])} 个重复厂商料号\n", False, 'warning')

                # 如果有详细信息，显示详细分析
                if self.validation_results.get('duplicate_mfrpns_detailed'):
                    detailed_info = self.validation_results['duplicate_mfrpns_detailed']

                    # 分类显示：替代料组 vs 真正的重复
                    alternative_duplicates = [info for info in detailed_info if info.get('is_alternative', False)]
                    real_duplicates = [info for info in detailed_info if not info.get('is_alternative', False)]

                    if alternative_duplicates:
                        self._update_result_text(f"  • 替代料组中的厂商料号重复 ({len(alternative_duplicates)}个):\n", False, 'emphasis')
                        for info in alternative_duplicates[:3]:  # 最多显示3个
                            mfrpn = info['mfrpn']
                            related_items = info['related_items']
                            self._update_result_text(f"    - {mfrpn} (替代料组: {', '.join(related_items)})\n", False, 'info')
                        if len(alternative_duplicates) > 3:
                            self._update_result_text(f"    - ... 等 {len(alternative_duplicates)-3} 个\n", False, 'info')

                    if real_duplicates:
                        self._update_result_text(f"  • 非替代料的厂商料号重复 ({len(real_duplicates)}个):\n", False, 'emphasis')
                        for info in real_duplicates:  # 显示所有真正的重复
                            mfrpn = info['mfrpn']
                            related_pns = info['related_pns']
                            count = info['count']
                            self._update_result_text(f"    - {mfrpn} (出现{count}次，涉及料号: {', '.join(related_pns)})\n", False, 'info')
                else:
                    # 如果没有详细信息，使用原来的简单显示方式
                    if len(self.validation_results['duplicate_mfrpns']) > 0:
                        mfrpns_to_show = self.validation_results['duplicate_mfrpns'][:5]  # 只显示前5个
                        self._update_result_text(f"   - {', '.join(mfrpns_to_show)}" +
                                               (f" ... 等" if len(self.validation_results['duplicate_mfrpns']) > 5 else "") +
                                               "\n", False, 'info')

            # 重复位号警告 - 显示所有重复位号
            if self.validation_results.get('duplicate_refs'):
                has_validation_issues = True
                self._update_result_text(f"\n⚠️ 警告: 发现 {len(self.validation_results['duplicate_refs'])} 个重复位号\n", False, 'warning')

                # 分类重复位号
                same_part_dups = []
                diff_part_dups = []

                for ref_info in self.validation_results['duplicate_refs']:
                    # 检查是否有same_part键，并使用它来判断类型
                    if 'same_part' in ref_info:
                        if ref_info['same_part']:
                            same_part_dups.append(ref_info)
                        else:
                            diff_part_dups.append(ref_info)
                    # 检查是否为包含多个物料的情况，即包含pns键且存在多个物料
                    elif 'pns' in ref_info and isinstance(ref_info['pns'], list) and len(ref_info['pns']) > 1:
                        diff_part_dups.append(ref_info)
                    # 兼容旧格式，只有pn键的情况归为同一物料内的重复
                    elif 'pn' in ref_info:
                        same_part_dups.append(ref_info)
                    # 如果无法判断类型，默认归为同一物料内的重复
                    else:
                        same_part_dups.append(ref_info)

                # 显示同一物料内重复的位号
                if same_part_dups:
                    self._update_result_text(f"  • 同一物料内重复的位号 ({len(same_part_dups)}个):\n", False, 'emphasis')
                    for ref_info in same_part_dups[:3]:  # 最多显示3个
                        ref = ref_info['ref']
                        # 检查字段存在性，兼容不同格式
                        if 'pn' in ref_info:
                            pn = ref_info['pn']
                        elif 'pns' in ref_info and ref_info['pns']:
                            pn = ref_info['pns'][0]
                        else:
                            pn = "未知物料"
                        self._update_result_text(f"    - {ref} 在物料 {pn} 中重复出现\n", False, 'info')
                    if len(same_part_dups) > 3:
                        self._update_result_text(f"    - ... 等 {len(same_part_dups)-3} 个\n", False, 'info')

                # 显示不同物料间重复的位号
                if diff_part_dups:
                    self._update_result_text(f"  • 不同物料间重复的位号 ({len(diff_part_dups)}个):\n", False, 'emphasis')
                    for ref_info in diff_part_dups:  # 显示所有重复位号
                        ref = ref_info['ref']
                        # 检查是否有'pns'键，如果没有则尝试使用其他键或创建一个默认值
                        if 'pns' in ref_info:
                            pns = ref_info['pns']
                        else:
                            # 如果没有pns键，可能是老格式，尝试构建一个pns列表
                            pns = []
                            if 'pn' in ref_info:
                                pns.append(ref_info['pn'])
                            # 确保pns至少有两个元素（因为是不同物料间的重复）
                            if len(pns) < 2:
                                pns.append("其他物料")

                        # 更清晰地显示位号重复问题，使用普通字体显示位号本身
                        if len(pns) > 2:
                            self._update_result_text(f"    - {ref}在 {len(pns)} 个物料中重复出现:\n", False, 'info')
                            # 显示每个物料
                            for i, pn in enumerate(pns, 1):
                                item = self.pn_to_item_map.get(pn, "") if hasattr(self, 'pn_to_item_map') else ""
                                if item:
                                    self._update_result_text(f"      {i}. {pn} (Item: {item})\n", False, 'info')
                                else:
                                    self._update_result_text(f"      {i}. {pn}\n", False, 'info')
                        else:
                            # 两个物料的情况
                            item1 = self.pn_to_item_map.get(pns[0], "") if hasattr(self, 'pn_to_item_map') else ""
                            item2 = self.pn_to_item_map.get(pns[1], "") if hasattr(self, 'pn_to_item_map') else ""
                            if item1 and item2:
                                self._update_result_text(f"    - {ref}在物料 {pns[0]} 和 {pns[1]} 中重复出现\n", False, 'info')
                            else:
                                self._update_result_text(f"    - {ref}在物料 {pns[0]} 和 {pns[1]} 中重复出现\n", False, 'info')

            # 位号数量不匹配警告
            if self.validation_results.get('quantity_mismatch'):
                has_validation_issues = True
                self._update_result_text(f"\n⚠️ 警告: 发现 {len(self.validation_results['quantity_mismatch'])} 处位号数量不匹配\n", False, 'warning')
                # 改进数量不匹配信息显示，使物料值和料号更加清晰
                for mismatch in self.validation_results['quantity_mismatch']:  # 显示所有不匹配项
                    self._update_result_text(f"   - Item: {mismatch['item']}，料号: {mismatch['pn']}，BOM中数量: {mismatch['expected']}，实际位号数: {mismatch['actual']}\n", False, 'info')

            # 如果没有任何问题
            if not has_validation_issues:
                self._update_result_text("✓ BOM验证通过，未发现问题\n", False, 'success')

            self._update_result_text("\n", False, 'info')

        # ===== 4. 未找到的位号部分 =====
        if not is_bom_only_mode and results['not_found']:
            not_found_header = f"【未找到的位号】\n{'-'*30}\n"
            self._update_result_text(not_found_header, False, 'subheader')

            # 每行最多显示8个位号，提高可读性
            not_found_refs = results['not_found']
            MAX_REFS_PER_LINE = 8
            MAX_DISPLAY_REFS = 100  # 最多显示前100个

            if len(not_found_refs) > MAX_DISPLAY_REFS:
                display_refs = not_found_refs[:MAX_DISPLAY_REFS]
                self._update_result_text(f"显示前 {MAX_DISPLAY_REFS} 个未找到的位号 (共 {len(not_found_refs)} 个):\n", False, 'info')
            else:
                display_refs = not_found_refs
                self._update_result_text(f"未找到的位号 (共 {len(not_found_refs)} 个):\n", False, 'info')

            # 分组显示位号
            for i in range(0, len(display_refs), MAX_REFS_PER_LINE):
                chunk = display_refs[i:i+MAX_REFS_PER_LINE]
                self._update_result_text(", ".join(chunk) + "\n", False, 'info')

            if len(not_found_refs) > MAX_DISPLAY_REFS:
                self._update_result_text(f"... 还有 {len(not_found_refs) - MAX_DISPLAY_REFS} 个位号未显示 ...\n", False, 'info')

            # 如果未匹配的位号很多，添加提示
            if len(not_found_refs) > 50:
                self._update_result_text("\n⚠️ 提示: 未匹配位号数量较多，请检查:\n", False, 'warning')
                self._update_result_text("1. PDF文件是否为文本可选择格式（非扫描版）\n", False, 'info')
                self._update_result_text("2. BOM中位号格式是否与原理图一致\n", False, 'info')
                self._update_result_text("3. 原理图中是否使用了特殊字体\n\n", False, 'info')

            self._update_result_text("\n", False, 'info')

        # ===== 5. 性能统计部分 =====
        performance_header = f"【性能统计】\n{'-'*30}\n"
        self._update_result_text(performance_header, False, 'subheader')

        # 根据模式显示不同的性能统计
        if is_bom_only_mode:
            time_stats = (
                f"• BOM解析耗时: {bom_seconds}.{bom_ms:03d}秒\n"
                f"• 总处理耗时: {int(total_minutes)}分{int(total_seconds)}秒\n"
            )
        else:
            time_stats = (
                f"• BOM解析耗时: {bom_seconds}.{bom_ms:03d}秒\n"
                f"• PDF处理耗时: {int(pdf_minutes)}分{int(pdf_seconds)}秒\n"
                f"• 总处理耗时: {int(total_minutes)}分{int(total_seconds)}秒\n"
            )

            # 添加输出文件信息
            if self.output_pdf_path:
                time_stats += f"\n标注后的原理图已保存为:\n{self.output_pdf_path}\n"

        self._update_result_text(time_stats, False, 'info')

        # 更新状态栏
        if not is_bom_only_mode and results.get('found') == results.get('total'):
            self._update_status("检查完成！所有位号已找到", self.success_color)
        elif is_bom_only_mode:
            self._update_status("BOM检查完成！", self.success_color)
        else:
            missing_count = len(results.get('not_found', []))
            self._update_status(f"检查完成！{missing_count} 个位号未找到", self.warning_color)

        # 启用保存结果按钮
        self.save_results_button.config(state='normal')

        # 保存当前结果以供导出使用
        self.current_results = results
        # 添加模式标记，便于导出时判断
        self.current_results['mode'] = 'bom_only' if is_bom_only_mode else 'full'

    def _open_output_pdf(self):
        """打开生成的PDF文件"""
        if hasattr(self, 'output_pdf_path') and os.path.exists(self.output_pdf_path):
            try:
                import subprocess
                import platform

                if platform.system() == 'Windows':
                    os.startfile(self.output_pdf_path)
                elif platform.system() == 'Darwin':  # macOS
                    subprocess.call(('open', self.output_pdf_path))
                else:  # Linux
                    subprocess.call(('xdg-open', self.output_pdf_path))

                self._update_status(f"已打开文件: {os.path.basename(self.output_pdf_path)}")
            except Exception as e:
                self._update_status(f"无法打开文件: {str(e)}", "red")
        else:
            self._update_status("找不到输出文件", "red")

    def _open_output_folder(self):
        """打开包含生成的PDF文件的文件夹"""
        if hasattr(self, 'output_pdf_path') and os.path.exists(self.output_pdf_path):
            folder_path = os.path.dirname(self.output_pdf_path)
            try:
                import subprocess
                import platform

                if platform.system() == 'Windows':
                    os.startfile(folder_path)
                elif platform.system() == 'Darwin':  # macOS
                    subprocess.call(('open', folder_path))
                else:  # Linux
                    subprocess.call(('xdg-open', folder_path))

                self._update_status(f"已打开文件夹: {folder_path}")
            except Exception as e:
                self._update_status(f"无法打开文件夹: {str(e)}", "red")
        else:
            self._update_status("找不到输出文件夹", "red")

    def save_results(self):
        """保存BOM检查结果为txt或Excel文件"""
        if not hasattr(self, 'current_results') or not self.current_results:
            messagebox.showinfo("无可保存的结果", "当前没有可保存的检查结果。")
            return

        # 使用复选框选择的格式
        save_as_excel = self.save_as_excel_var.get()

        # 生成默认文件名
        default_filename = self._generate_default_filename(save_as_excel)

        # 选择保存路径
        if not save_as_excel:  # 文本文件
            file_path = filedialog.asksaveasfilename(defaultextension=".txt",
                                                  filetypes=[("Text files", "*.txt")],
                                                  title="保存检查结果为文本文件",
                                                  initialfile=default_filename)
            if not file_path:
                return  # 用户取消

            try:
                # 获取文本内容
                text_content = self.result_text.get(1.0, tk.END)

                # 保存为文本文件
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(text_content)

                self._update_status(f"结果已保存为文本文件: {os.path.basename(file_path)}", self.success_color)

                # 询问是否打开文件
                if messagebox.askyesno("保存成功", f"结果已保存为文本文件: {os.path.basename(file_path)}\n\n是否打开文件？"):
                    self._open_file(file_path)
            except Exception as e:
                messagebox.showerror("保存错误", f"保存文本文件时出错: {str(e)}")

        else:  # Excel文件
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                  filetypes=[("Excel files", "*.xlsx")],
                                                  title="保存检查结果为Excel文件",
                                                  initialfile=default_filename)
            if not file_path:
                return  # 用户取消

            try:
                # 创建Excel文件
                workbook = pd.ExcelWriter(file_path, engine='openpyxl')

                # 准备数据
                results = self.current_results
                is_bom_only_mode = results.get('mode') == 'bom_only'

                # 创建单个工作表的数据列表
                all_data = []
                row_index = 0

                # 添加标题行
                all_data.append({
                    '类型': '标题',
                    '项目1': 'BOM检查结果报告',
                    '项目2': '',
                    '项目3': '',
                    '项目4': ''
                })
                row_index += 1

                # 添加空行
                all_data.append({
                    '类型': '',
                    '项目1': '',
                    '项目2': '',
                    '项目3': '',
                    '项目4': ''
                })
                row_index += 1

                # 添加基本信息标题
                all_data.append({
                    '类型': '分类标题',
                    '项目1': '基本信息',
                    '项目2': '',
                    '项目3': '',
                    '项目4': ''
                })
                row_index += 1

                # 添加基本信息列标题
                all_data.append({
                    '类型': '列标题',
                    '项目1': '检查项目',
                    '项目2': '数值',
                    '项目3': '',
                    '项目4': ''
                })
                row_index += 1

                # 添加BOM中位号总数
                all_data.append({
                    '类型': '数据',
                    '项目1': 'BOM中位号总数',
                    '项目2': results.get('total_refs', 0) if is_bom_only_mode else results.get('total', 0),
                    '项目3': '',
                    '项目4': ''
                })
                row_index += 1

                # 如果不是BOM只模式，添加PDF相关信息
                if not is_bom_only_mode:
                    all_data.append({
                        '类型': '数据',
                        '项目1': '在PDF中找到',
                        '项目2': results.get('found', 0),
                        '项目3': '',
                        '项目4': ''
                    })
                    row_index += 1

                    all_data.append({
                        '类型': '数据',
                        '项目1': '在PDF中未找到',
                        '项目2': len(results.get('not_found', [])),
                        '项目3': '',
                        '项目4': ''
                    })
                    row_index += 1

                # 添加空行
                all_data.append({
                    '类型': '',
                    '项目1': '',
                    '项目2': '',
                    '项目3': '',
                    '项目4': ''
                })
                row_index += 1

                # 如果有BOM验证结果，添加验证结果
                if hasattr(self, 'validation_results'):
                    # 重复料号
                    if self.validation_results.get('duplicate_pns'):
                        all_data.append({
                            '类型': '分类标题',
                            '项目1': '重复料号',
                            '项目2': '',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                        all_data.append({
                            '类型': '列标题',
                            '项目1': '重复料号',
                            '项目2': '',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                        for pn in self.validation_results['duplicate_pns']:
                            all_data.append({
                                '类型': '数据',
                                '项目1': pn,
                                '项目2': '',
                                '项目3': '',
                                '项目4': ''
                            })
                            row_index += 1

                        # 添加空行
                        all_data.append({
                            '类型': '',
                            '项目1': '',
                            '项目2': '',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                    # 重复厂商料号
                    if self.validation_results.get('duplicate_mfrpns'):
                        all_data.append({
                            '类型': '分类标题',
                            '项目1': '重复厂商料号',
                            '项目2': '',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                        all_data.append({
                            '类型': '列标题',
                            '项目1': '重复厂商料号',
                            '项目2': '',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                        for mfrpn in self.validation_results['duplicate_mfrpns']:
                            all_data.append({
                                '类型': '数据',
                                '项目1': mfrpn,
                                '项目2': '',
                                '项目3': '',
                                '项目4': ''
                            })
                            row_index += 1

                        # 添加空行
                        all_data.append({
                            '类型': '',
                            '项目1': '',
                            '项目2': '',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                    # 重复位号
                    if self.validation_results.get('duplicate_refs'):
                        all_data.append({
                            '类型': '分类标题',
                            '项目1': '重复位号',
                            '项目2': '',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                        all_data.append({
                            '类型': '列标题',
                            '项目1': '位号',
                            '项目2': '料号',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                        for ref_info in self.validation_results['duplicate_refs']:
                            ref = ref_info['ref']
                            if 'pns' in ref_info:
                                pns = ', '.join(ref_info['pns'])
                            elif 'pn' in ref_info:
                                pns = ref_info['pn']
                            else:
                                pns = '未知'

                            all_data.append({
                                '类型': '数据',
                                '项目1': ref,
                                '项目2': pns,
                                '项目3': '',
                                '项目4': ''
                            })
                            row_index += 1

                        # 添加空行
                        all_data.append({
                            '类型': '',
                            '项目1': '',
                            '项目2': '',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                    # 位号数量不匹配
                    if self.validation_results.get('quantity_mismatch'):
                        all_data.append({
                            '类型': '分类标题',
                            '项目1': '位号数量不匹配',
                            '项目2': '',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                        all_data.append({
                            '类型': '列标题',
                            '项目1': '物料项',
                            '项目2': '料号',
                            '项目3': 'BOM中数量',
                            '项目4': '实际位号数'
                        })
                        row_index += 1

                        for mismatch in self.validation_results['quantity_mismatch']:
                            all_data.append({
                                '类型': '数据',
                                '项目1': mismatch.get('item', ''),
                                '项目2': mismatch.get('pn', ''),
                                '项目3': mismatch.get('expected', 0),
                                '项目4': mismatch.get('actual', 0)
                            })
                            row_index += 1

                        # 添加空行
                        all_data.append({
                            '类型': '',
                            '项目1': '',
                            '项目2': '',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                # 如果有未找到的位号，添加未找到的位号
                if not is_bom_only_mode and results.get('not_found'):
                    all_data.append({
                        '类型': '分类标题',
                        '项目1': '未找到的位号',
                        '项目2': '',
                        '项目3': '',
                        '项目4': ''
                    })
                    row_index += 1

                    all_data.append({
                        '类型': '列标题',
                        '项目1': '未找到的位号',
                        '项目2': '',
                        '项目3': '',
                        '项目4': ''
                    })
                    row_index += 1

                    for ref in results['not_found']:
                        all_data.append({
                            '类型': '数据',
                            '项目1': ref,
                            '项目2': '',
                            '项目3': '',
                            '项目4': ''
                        })
                        row_index += 1

                # 创建数据帧并保存到Excel
                df = pd.DataFrame(all_data)
                df.to_excel(workbook, sheet_name='BOM检查结果', index=False)

                # 获取工作表并进行格式化
                worksheet = workbook.book['BOM检查结果']

                # 设置列宽
                worksheet.column_dimensions['A'].width = 12  # 类型列
                worksheet.column_dimensions['B'].width = 25  # 项目1列
                worksheet.column_dimensions['C'].width = 25  # 项目2列
                worksheet.column_dimensions['D'].width = 15  # 项目3列
                worksheet.column_dimensions['E'].width = 15  # 项目4列

                # 定义样式
                title_font = Font(name='Microsoft YaHei', size=12, bold=True)  # 标题字体
                category_font = Font(name='Microsoft YaHei', size=11, bold=True)  # 分类标题字体
                header_font = Font(name='Microsoft YaHei', size=10, bold=True)  # 列标题字体
                data_font = Font(name='Microsoft YaHei', size=10)  # 数据字体

                # 对齐方式
                center_alignment = Alignment(horizontal='center', vertical='center')
                left_alignment = Alignment(horizontal='left', vertical='center')
                right_alignment = Alignment(horizontal='right', vertical='center')

                # 填充颜色
                title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # 标题背景色
                category_fill = PatternFill(start_color='8EA9DB', end_color='8EA9DB', fill_type='solid')  # 分类标题背景色
                header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # 列标题背景色

                # 边框样式
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 应用样式
                row_index = 1  # Excel中的行索引从1开始
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                    row_type = worksheet.cell(row=row_index, column=1).value

                    # 合并标题和分类标题单元格
                    if row_type in ['标题', '分类标题']:
                        worksheet.merge_cells(f'B{row_index}:E{row_index}')

                    # 应用样式到每个单元格
                    for cell in row:
                        # 设置边框
                        cell.border = thin_border

                        # 根据行类型设置样式
                        if row_type == '标题':
                            cell.font = title_font
                            cell.fill = title_fill
                            cell.alignment = center_alignment
                        elif row_type == '分类标题':
                            cell.font = category_font
                            cell.fill = category_fill
                            cell.alignment = center_alignment
                        elif row_type == '列标题':
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_alignment
                        elif row_type == '数据':
                            cell.font = data_font
                            # 数据行的第一列居左，数值列居右
                            if cell.column == 2:  # 项目1列
                                cell.alignment = left_alignment
                            elif cell.column in [3, 4, 5]:  # 数值列
                                cell.alignment = right_alignment
                            else:
                                cell.alignment = center_alignment
                        else:  # 空行或其他
                            cell.font = data_font
                            cell.alignment = center_alignment

                    row_index += 1

                # 保存Excel文件
                workbook.close()

                self._update_status(f"结果已保存为Excel文件: {os.path.basename(file_path)}", self.success_color)

                # 询问是否打开文件
                if messagebox.askyesno("保存成功", f"结果已保存为Excel文件: {os.path.basename(file_path)}\n\n是否打开文件？"):
                    self._open_file(file_path)
            except Exception as e:
                messagebox.showerror("保存错误", f"保存Excel文件时出错: {str(e)}")

    def _generate_default_filename(self, is_excel=False):
        """生成默认文件名

        Args:
            is_excel (bool): 是否为Excel文件

        Returns:
            str: 默认文件名
        """
        # 获取当前日期时间
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 基本文件名
        base_filename = "BOM检查结果"

        # 如果有BOM文件，使用BOM文件名作为基础
        if hasattr(self, 'bom_path') and self.bom_path:
            bom_filename = os.path.basename(self.bom_path)
            # 去除扩展名
            bom_basename = os.path.splitext(bom_filename)[0]
            base_filename = f"{bom_basename}_检查结果"

        # 添加日期时间
        filename = f"{base_filename}_{current_time}"

        # 添加扩展名
        if is_excel:
            filename += ".xlsx"
        else:
            filename += ".txt"

        return filename

    def _open_file(self, file_path):
        """打开指定文件"""
        if os.path.exists(file_path):
            try:
                import subprocess
                import platform

                if platform.system() == 'Windows':
                    os.startfile(file_path)
                elif platform.system() == 'Darwin':  # macOS
                    subprocess.call(('open', file_path))
                else:  # Linux
                    subprocess.call(('xdg-open', file_path))

                self._update_status(f"已打开文件: {os.path.basename(file_path)}")
            except Exception as e:
                self._update_status(f"无法打开文件: {str(e)}", "red")
        else:
            self._update_status(f"找不到文件: {file_path}", "red")

    def _show_error(self, error, show_dialog=True):
        """显示错误信息，可选择是否弹出对话框"""
        # 将英文错误信息转换为中文
        chinese_error = self._translate_error_to_chinese(error)

        self._update_result_text("", True)  # 清空文本
        self._update_result_text(f"❌ 发生错误:\n\n{chinese_error}")
        self._update_status("处理出错", self.error_color)

        # 禁用保存结果按钮
        self.save_results_button.config(state='disabled')

        # 记录到日志
        logging.error(f"错误: {error}")

        # 可选择是否显示错误对话框
        if show_dialog:
            # 处理Traceback信息，只显示简洁的错误消息
            short_error = chinese_error.split('Traceback')[0] if 'Traceback' in chinese_error else chinese_error
            messagebox.showerror("处理错误", f"处理过程中发生错误:\n{short_error}")

        # 不需要返回值，让调用者决定是否继续

    def _translate_error_to_chinese(self, error):
        """将英文错误信息转换为中文错误信息"""
        # 常见错误信息的中英文映射
        error_mapping = {
            "name 'logging' is not defined": "程序内部错误：日志模块未定义，请联系开发者",
            "name 'chardet' is not defined": "缺少chardet模块，请安装：pip install chardet",
            "No module named": "缺少必要的模块，请安装相应的Python库",
            "Permission denied": "文件权限不足，请以管理员身份运行或检查文件权限",
            "File not found": "找不到指定的文件，请检查文件路径是否正确",
            "Invalid file": "无效的文件，请检查文件格式是否正确",
            "Failed to open": "打开文件失败，请确保文件未被其他程序占用",
            "Memory error": "内存不足，请关闭其他程序或增加系统内存",
            "Disk full": "磁盘空间不足，请清理磁盘空间",
            "Timeout": "操作超时，请检查网络连接或重试",
            "Connection error": "连接错误，请检查网络连接",
            "Invalid format": "无效的格式，请检查文件格式是否正确",
            "Cannot read": "无法读取文件，请检查文件是否损坏或格式不正确",
            "Cannot write": "无法写入文件，请检查文件权限或磁盘空间",
            "Excel file format": "Excel文件格式错误，请使用标准的Excel格式(.xlsx或.xls)",
            "CSV file format": "CSV文件格式错误，请检查CSV文件的编码和格式",
            "PDF file format": "PDF文件格式错误，请确保使用标准的PDF格式",
            "UnicodeDecodeError": "文件编码错误，请使用UTF-8或GBK编码保存文件",
            "KeyError": "程序内部错误：键值错误，请联系开发者",
            "IndexError": "程序内部错误：索引超出范围，请联系开发者",
            "TypeError": "程序内部错误：类型错误，请联系开发者",
            "ValueError": "程序内部错误：值错误，请联系开发者",
            "AttributeError": "程序内部错误：属性错误，请联系开发者",
            "ImportError": "程序内部错误：导入错误，请联系开发者",
            "ModuleNotFoundError": "缺少必要的模块，请安装相应的Python库",
            "OSError": "操作系统错误，请检查文件权限或磁盘空间",
            "FileNotFoundError": "找不到指定的文件，请检查文件路径是否正确",
            "PermissionError": "文件权限不足，请以管理员身份运行或检查文件权限",
            "FileExistsError": "文件已存在，请尝试使用其他文件名",
            "NotImplementedError": "功能尚未实现，请等待后续版本",
            "RuntimeError": "运行时错误，请重新启动程序或联系开发者",
            "Exception": "程序异常，请重新启动程序或联系开发者"
        }

        # 检查错误信息是否包含已知的错误模式
        for eng_error, cn_error in error_mapping.items():
            if eng_error in str(error):
                return cn_error

        # 如果没有匹配到已知错误，返回原始错误信息
        return str(error)

    def run_check(self):
        """开始检查流程"""
        if self.processing:
            messagebox.showinfo("处理中", "当前正在处理中，请等待完成...")
            return

        if not self.bom_path:
            messagebox.showwarning("文件缺失", "请先选择BOM文件！")
            return

        # 移除PDF文件检查的要求
        # if not self.pdf_path:
        #     messagebox.showwarning("文件缺失", "请先选择PDF原理图文件！")
        #     return

        self.run_button.config(state='disabled')
        self.save_results_button.config(state='disabled')  # 禁用保存结果按钮
        self.progress['value'] = 0
        self.progress_label.config(text="0%")

        self._update_result_text("", True)
        self._update_result_text("正在处理中，请稍候...\n")

        # 清空消息队列
        while not self.message_queue.empty():
            try:
                self.message_queue.get_nowait()
            except queue.Empty:
                break

        # 重置UI更新标志
        self.stop_ui_update = False

        # 启动后台线程处理
        threading.Thread(target=self._run_in_thread, daemon=True).start()

        # 启动UI更新线程
        threading.Thread(target=self._update_ui_from_queue, daemon=True).start()

    def _update_ui_from_queue(self):
        """从队列中获取消息并更新UI"""
        last_activity_time = time.time()

        while not self.stop_ui_update:
            try:
                # 非阻塞方式获取消息，超时后继续循环检查停止标志
                try:
                    message = self.message_queue.get(timeout=0.1)
                    # 更新活动时间
                    last_activity_time = time.time()
                except queue.Empty:
                    # 检查UI更新线程是否长时间没有活动，如果超过30秒则记录警告
                    if time.time() - last_activity_time > 30 and self.processing:
                        print(f"警告: UI更新线程长时间无活动 ({time.time() - last_activity_time:.1f}秒)")
                        # 发送一个心跳消息到队列
                        try:
                            self.message_queue.put(("status", "处理中...", self.primary_color))
                            last_activity_time = time.time()
                        except:
                            pass
                    continue

                # 处理不同类型的消息
                msg_type = message[0]

                try:
                    if msg_type == "status":
                        _, msg, color = message
                        self.master.after(0, lambda m=msg, c=color: self._update_status(m, c))

                    elif msg_type == "text":
                        _, text, clear = message
                        self.master.after(0, lambda t=text, c=clear: self._do_update_result_text(t, c))

                    elif msg_type == "progress":
                        _, value = message
                        self.master.after(0, lambda v=value: self._update_progress(v))

                    elif msg_type == "show_results":
                        _, results = message
                        self.master.after(0, lambda r=results: self._show_results(r))

                    elif msg_type == "show_error":
                        _, error = message
                        # 将错误信息转换为中文后再显示
                        self.master.after(0, lambda e=error: self._show_error(e))

                    elif msg_type == "update_button":
                        _, config = message
                        self.master.after(0, lambda c=config: self.run_button.config(**c))
                except Exception as e:
                    print(f"处理消息 {msg_type} 时出错: {str(e)}")

                # 标记任务完成
                self.message_queue.task_done()

            except Exception as e:
                print(f"UI更新错误: {str(e)}")
                # 短暂暂停，避免因持续错误导致CPU占用过高
                time.sleep(0.5)

        print("UI更新线程退出")

    def show_about(self):
        """显示关于窗口"""
        about_window = tk.Toplevel(self.master)
        about_window.title("关于本软件")
        about_window.withdraw()  # 先隐藏窗口
        about_window.minsize(550, 500)  # 增加窗口最小尺寸
        about_window.resizable(False, False)

        # 设置模态窗口
        about_window.transient(self.master)
        about_window.grab_set()

        # 主容器
        about_container = ttk.Frame(about_window, padding=20)
        about_container.pack(fill='both', expand=True)

        # 顶部标题区域
        header_frame = ttk.Frame(about_container)
        header_frame.pack(fill='x', pady=(0, 5))

        # 尝试添加软件图标
        try:
            if 'Image' in globals():  # 检查PIL是否可用
                # 使用PIL加载ico文件
                icon_path = "icon.ico"
                if os.path.exists(icon_path):
                    pil_img = Image.open(icon_path)
                    pil_img = pil_img.resize((64, 64), Image.LANCZOS)  # 调整大小
                    logo_img = ImageTk.PhotoImage(pil_img)
                    logo_label = ttk.Label(header_frame, image=logo_img)
                    logo_label.image = logo_img  # 保持引用防止被垃圾回收
                    logo_label.pack(side='left', padx=(0, 15))
        except Exception as e:
            print(f"加载图标出错: {e}")
            pass  # 忽略任何错误

        # 标题和版本信息
        title_info = ttk.Frame(header_frame)
        title_info.pack(side='left', fill='y', expand=True)

        title_label = ttk.Label(title_info, text="BOM检查工具",
                             font=('微软雅黑', 18, 'bold'),
                             foreground=self.primary_color)
        title_label.pack(anchor='w')

        version_label = ttk.Label(title_info, text="            V1.8",
                              font=('微软雅黑', 12, 'bold'))
        version_label.pack(anchor='w')

        # 分隔线
        separator = ttk.Separator(about_container, orient='horizontal')
        separator.pack(fill='x', pady=10)

        # 添加程序简介
        intro_frame = ttk.Frame(about_container)
        intro_frame.pack(fill='x', pady=(0, 10))

        intro_text = ttk.Label(intro_frame,
                            text="BOM检查工具是一款专为电子工程师设计的软件，能够快速对比BOM与原理图中的位号一致性，支持替代料识别、重复位号检测及物料信息注解，大幅提升元器件核对效率。",
                            font=('微软雅黑', 9),
                            wraplength=500,
                            justify='left')
        intro_text.pack(fill='x')

        # 核心功能展示（使用多列布局）
        features_frame = ttk.LabelFrame(about_container, text="✨ 核心功能", padding=(15, 10))
        features_frame.pack(fill='x', pady=(0, 10))

        # 创建两列功能列表
        features_grid = ttk.Frame(features_frame)
        features_grid.pack(fill='both', expand=True)

        # 使用grid布局而不是pack来更精确控制对齐
        features_grid.columnconfigure(0, weight=1)  # 第一列
        features_grid.columnconfigure(1, weight=1)  # 第二列

        # 第一列功能 - 使用grid
        ttk.Label(features_grid, text="📊 BOM检查", font=('微软雅黑', 10, 'bold')).grid(row=0, column=0, sticky='w', pady=(0, 5))
        ttk.Label(features_grid, text="• 自动对比位号一致性\n• 识别替代料组关系\n• 检测重复位号问题",
                 font=('微软雅黑', 9)).grid(row=1, column=0, sticky='w')

        ttk.Label(features_grid, text="🔍 PDF处理", font=('微软雅黑', 10, 'bold')).grid(row=2, column=0, sticky='w', pady=(10, 5))
        ttk.Label(features_grid, text="• 智能高亮标注位号\n• 支持多种位号格式\n• 可添加物料信息注解",
                 font=('微软雅黑', 9)).grid(row=3, column=0, sticky='w')

        # 第二列功能 - 使用grid确保标题对齐
        ttk.Label(features_grid, text="⚡ 性能优化", font=('微软雅黑', 10, 'bold')).grid(row=0, column=1, sticky='w', pady=(0, 5))
        ttk.Label(features_grid, text="• 优化大文件处理效率",
                 font=('微软雅黑', 9)).grid(row=1, column=1, sticky='w')

        ttk.Label(features_grid, text="🔧 实用工具", font=('微软雅黑', 10, 'bold')).grid(row=2, column=1, sticky='w', pady=(10, 5))
        ttk.Label(features_grid, text="• 自定义表头映射\n• 可定制高亮颜色\n• 详细的校验报告",
                 font=('微软雅黑', 9)).grid(row=3, column=1, sticky='w')

        # 技术支持联系方式
        highlight_frame = ttk.LabelFrame(about_container, text="📞 技术支持联系方式", padding=(15, 10))
        highlight_frame.pack(fill='x', pady=(0, 10))

        highlights = (
            "• 微信：XiaoHang_Sky\n"
        )

        # 增加文本框高度以显示所有内容
        highlight_text = tk.Text(highlight_frame, wrap='word', font=('微软雅黑', 9),
                               height=5, width=50, bg=self.secondary_color)
        highlight_text.insert('1.0', highlights)
        highlight_text.config(state='disabled')  # 设为只读
        highlight_text.pack(fill='both', expand=True, pady=5)

        # 底部信息部分
        info_frame = ttk.Frame(about_container)
        info_frame.pack(fill='x', pady=(5, 10))

        # 左侧开发者信息
        dev_frame = ttk.Frame(info_frame)
        dev_frame.pack(side='left', fill='y')

        ttk.Label(dev_frame, text="开发者：小航", font=('微软雅黑', 9, 'bold')).pack(anchor='w')
        ttk.Label(dev_frame, text=f"版权所有 © 2025", font=('微软雅黑', 9)).pack(anchor='w')
        ttk.Label(dev_frame, text=f"最近更新：2025年4月1日", font=('微软雅黑', 9)).pack(anchor='w')

        # 右侧技术支持信息
        support_frame = ttk.Frame(info_frame)
        support_frame.pack(side='right', fill='y')

        ttk.Label(support_frame, text="📞 技术支持", font=('微软雅黑', 9, 'bold')).pack(anchor='e')
        ttk.Label(support_frame, text="微信：XiaoHang_Sky", font=('微软雅黑', 9)).pack(anchor='e')
        ttk.Label(support_frame, text="工作时间：周一至周五 9:00-18:00", font=('微软雅黑', 9)).pack(anchor='e')

        # 底部按钮
        button_frame = ttk.Frame(about_container)
        button_frame.pack(pady=(10, 0))

        close_btn = ttk.Button(button_frame, text="关闭",
                           command=about_window.destroy,
                           style='Primary.TButton',
                           width=10)
        close_btn.pack()

        # 使用特殊方法确保窗口居中显示
        about_window.update_idletasks()
        self.center_window(about_window, 550, 500)  # 调整窗口大小
        about_window.deiconify()  # 准备好后再显示窗口

    def show_settings(self):
        """显示设置对话框"""
        settings_window = tk.Toplevel(self.master)
        settings_window.title("设置")
        settings_window.withdraw()  # 先隐藏窗口
        settings_window.minsize(450, 480)
        settings_window.transient(self.master)
        settings_window.grab_set()

        # 设置字体样式
        title_font = ('微软雅黑', 13, 'bold')
        label_font = ('微软雅黑', 10)
        note_font = ('微软雅黑', 9)

        # 设置容器
        container = ttk.Frame(settings_window, padding=15)
        container.pack(fill='both', expand=True)

        # 标题
        ttk.Label(container, text="应用设置",
                font=title_font).pack(pady=(0, 15))

        # 创建选项网格布局框架
        options_frame = ttk.LabelFrame(container, text="PDF处理选项", padding=10)
        options_frame.pack(fill='x', pady=(0, 10))

        # 创建两列网格布局
        options_grid = ttk.Frame(options_frame)
        options_grid.pack(fill='x', expand=True)
        options_grid.columnconfigure(0, weight=1)  # 左侧栏
        options_grid.columnconfigure(1, weight=1)  # 右侧栏

        # 左侧：高亮颜色选择
        color_frame = ttk.LabelFrame(options_grid, text="高亮颜色")
        color_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 5), pady=5)

        color_selection_frame = ttk.Frame(color_frame)
        color_selection_frame.pack(fill='both', expand=True, pady=5, padx=5)

        # 添加下拉菜单
        color_combobox = ttk.Combobox(color_selection_frame, textvariable=self.color_var,
                                     values=list(self.color_choices.keys()), state="readonly",
                                     width=15, font=label_font)
        color_combobox.pack(side='top', fill='x', pady=(0, 5))

        # 添加颜色预览画布
        preview_frame = ttk.LabelFrame(color_selection_frame, text="颜色预览")
        preview_frame.pack(side='top', fill='x', expand=True)

        self.color_preview = tk.Canvas(preview_frame, width=100, height=25, highlightthickness=1)
        self.color_preview.pack(padx=5, pady=5, fill='x', expand=True)

        # 初始更新颜色预览
        self._update_color_preview()

        # 绑定颜色变更事件
        color_combobox.bind('<<ComboboxSelected>>', lambda e: self._update_color_preview())

        # 右侧：添加注解选项
        annot_frame = ttk.LabelFrame(options_grid, text="PDF注解设置")
        annot_frame.grid(row=0, column=1, sticky='nsew', padx=(5, 0), pady=5)

        annot_content = ttk.Frame(annot_frame)
        annot_content.pack(fill='both', expand=True, pady=5, padx=5)

        annot_check = ttk.Checkbutton(annot_content, text="添加物料信息注解",
                                    variable=self.add_annotation_var,
                                    style='Large.TCheckbutton',
                                    command=self.save_settings)
        annot_check.pack(fill='x', pady=(5, 0))

        ttk.Label(annot_content, text="(注解默认隐藏)\n\n启用此选项将在高亮位号的\n同时添加该位号对应的物料\n信息作为注解",
                font=note_font, justify='left').pack(fill='x', pady=(5, 0))

        # 添加BOM表头映射设置按钮
        mapping_frame = ttk.LabelFrame(container, text="BOM表头映射", padding=10)
        mapping_frame.pack(fill='x', pady=(0, 10))

        ttk.Button(mapping_frame, text="编辑BOM表头映射",
                command=self.show_mapping_settings, style='Large.TButton').pack(fill='x', pady=5)

        ttk.Label(mapping_frame, text="自定义BOM文件中的表头与程序识别之间的映射关系",
                font=note_font).pack(fill='x', pady=(5, 0))

        # 按钮区域 - 增加空间
        button_frame = ttk.Frame(container)
        button_frame.pack(fill='x', pady=(20, 5))

        # 使用普通的tk.Button替代ttk.Button，更容易控制样式
        save_button = tk.Button(button_frame, text="确定",
                            command=lambda: [self.save_settings(), settings_window.destroy()],
                            width=12, height=2, font=('微软雅黑', 11, 'bold'),
                            bg='#4a6da7', fg='white', relief='raised')
        save_button.pack(side='right', padx=10, pady=10)

        # 添加悬停效果
        def on_enter(e):
            save_button['bg'] = '#3a5d97'  # 稍深色

        def on_leave(e):
            save_button['bg'] = '#4a6da7'  # 恢复原色

        save_button.bind("<Enter>", on_enter)
        save_button.bind("<Leave>", on_leave)

        # 设置自定义样式
        style = ttk.Style()
        style.configure('Large.TButton', font=label_font, padding=(5, 3))
        style.configure('Large.TCheckbutton', font=label_font)
        style.configure('TLabelframe.Label', font=label_font)

        # 防止窗口大小被改变，确保UI布局一致
        settings_window.resizable(False, False)

        # 使用特殊方法确保窗口居中显示并平滑显示
        self.center_window(settings_window, 450, 480)
        settings_window.deiconify()  # 准备好后再显示窗口

    def show_mapping_settings(self):
        """显示并编辑表头映射设置"""
        settings_window = tk.Toplevel(self.master)
        settings_window.title("BOM表头映射设置")
        settings_window.withdraw()  # 先隐藏窗口
        settings_window.minsize(800, 600)
        settings_window.transient(self.master)
        settings_window.grab_set()

        # 设置全局样式和变量
        style = ttk.Style()
        style.configure("Header.TLabel", font=('微软雅黑', 12, 'bold'))
        style.configure("Category.TLabel", font=('微软雅黑', 10, 'bold'))
        style.configure("Help.TLabel", font=('微软雅黑', 9), foreground="#666666")

        # 创建主容器
        main_frame = ttk.Frame(settings_window, padding=20)
        main_frame.pack(fill='both', expand=True)

        # 页面标题
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill='x', pady=(0, 20))

        ttk.Label(header_frame, text="BOM表头映射配置", style="Header.TLabel").pack(side='left')
        ttk.Label(header_frame,
                 text="配置不同BOM文件中各字段的可能表头名称，系统将自动识别匹配",
                 style="Help.TLabel").pack(side='left', padx=(15, 0))

        # 创建选项卡控件
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill='both', expand=True, pady=(0, 15))

        # 字段类型说明和标签
        field_labels = {
            'reference': {'title': '位号列', 'description': '元件在PCB上的位置标识，如R1, C2等'},
            'pn': {'title': '料号列', 'description': '元件的唯一识别编号'},
            'part': {'title': '元件类型列', 'description': '元件的类型，如电阻、电容等'},
            'quantity': {'title': '数量列', 'description': '元件在PCB上的使用数量'},
            'description': {'title': '描述列', 'description': '元件的详细描述信息'},
            'mfr_pn': {'title': '厂商料号列', 'description': '制造商的零件编号'},
            'manufacturer': {'title': '制造商列', 'description': '元件的生产厂商名称'}
        }

        # 存储所有选项卡中的列表控件
        self.mapping_listboxes = {}

        for field, info in field_labels.items():
            # 创建选项卡页面
            tab = ttk.Frame(notebook, padding=15)
            notebook.add(tab, text=info['title'])

            # 标题和说明区域
            ttk.Label(tab, text=info['title'], style="Category.TLabel").pack(anchor='w')
            ttk.Label(tab, text=info['description'], style="Help.TLabel").pack(anchor='w', pady=(0, 15))

            # 工具栏区域
            toolbar = ttk.Frame(tab)
            toolbar.pack(fill='x', pady=(0, 10))

            # 搜索框
            search_var = tk.StringVar()
            search_frame = ttk.Frame(toolbar)
            search_frame.pack(side='left')

            ttk.Label(search_frame, text="搜索:").pack(side='left', padx=(0, 5))
            search_entry = ttk.Entry(search_frame, width=20, textvariable=search_var)
            search_entry.pack(side='left')

            # 添加新项按钮
            btn_frame = ttk.Frame(toolbar)
            btn_frame.pack(side='right')

            add_var = tk.StringVar()
            ttk.Entry(btn_frame, width=20, textvariable=add_var).pack(side='left', padx=(0, 5))

            # 列表框和滚动条
            list_frame = ttk.Frame(tab)
            list_frame.pack(fill='both', expand=True)

            # 使用Listbox+Scrollbar替代Text
            listbox = tk.Listbox(list_frame, font=('微软雅黑', 10), selectmode=tk.EXTENDED, activestyle='none')
            listbox.pack(side='left', fill='both', expand=True)

            scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
            scrollbar.pack(side='right', fill='y')
            listbox.config(yscrollcommand=scrollbar.set)

            # 添加已有映射到列表框
            if field in self.bom_header_mapping:
                for item in self.bom_header_mapping[field]:
                    listbox.insert(tk.END, item)

            # 保存列表框引用
            self.mapping_listboxes[field] = listbox

            # 为添加按钮创建函数
            def add_item_factory(lb, entry_var):
                def add_item():
                    value = entry_var.get().strip()
                    if value and value not in lb.get(0, tk.END):
                        lb.insert(tk.END, value)
                        entry_var.set('')  # 清空输入框
                    elif value in lb.get(0, tk.END):
                        messagebox.showinfo("提示", f"'{value}'已存在，不能重复添加", parent=settings_window)
                    lb.see(tk.END)  # 滚动到最后
                return add_item

            add_btn = ttk.Button(btn_frame, text="添加",
                               command=add_item_factory(listbox, add_var))
            add_btn.pack(side='left')

            # 为删除按钮创建函数
            def delete_selected_factory(lb):
                def delete_selected():
                    selected = lb.curselection()
                    if selected:
                        # 从后往前删除，避免索引变化
                        for index in sorted(selected, reverse=True):
                            lb.delete(index)
                return delete_selected

            del_btn = ttk.Button(toolbar, text="删除所选",
                               command=delete_selected_factory(listbox))
            del_btn.pack(side='right', padx=5)

            # 为搜索框绑定事件
            def filter_list(event=None):
                search_text = search_var.get().lower()
                listbox.delete(0, tk.END)
                if field in self.bom_header_mapping:
                    for item in self.bom_header_mapping[field]:
                        if search_text in item.lower() or not search_text:
                            listbox.insert(tk.END, item)

            search_entry.bind('<KeyRelease>', filter_list)

            # 右键菜单
            context_menu = tk.Menu(listbox, tearoff=0)

            def show_context_menu(event):
                if listbox.size() > 0:
                    # 确保在点击的位置选中项目
                    clicked_index = listbox.nearest(event.y)
                    if clicked_index >= 0:
                        # 如果按住Ctrl，不清除已选择的项目
                        if not (event.state & 0x0004):  # 0x0004是Ctrl键的状态
                            listbox.selection_clear(0, tk.END)
                        listbox.selection_set(clicked_index)
                        context_menu.post(event.x_root, event.y_root)

            listbox.bind('<Button-3>', show_context_menu)

            context_menu.add_command(label="删除",
                                    command=delete_selected_factory(listbox))
            context_menu.add_command(label="复制",
                                    command=lambda: self.master.clipboard_append(
                                        listbox.get(listbox.curselection()[0]) if listbox.curselection() else ""))

            # 双击编辑功能
            def on_double_click(event):
                if listbox.curselection():
                    index = listbox.curselection()[0]
                    value = listbox.get(index)

                    # 创建编辑对话框
                    edit_dialog = tk.Toplevel(settings_window)
                    edit_dialog.title("编辑条目")
                    edit_dialog.transient(settings_window)
                    edit_dialog.grab_set()

                    dialog_frame = ttk.Frame(edit_dialog, padding=15)
                    dialog_frame.pack(fill='both', expand=True)

                    ttk.Label(dialog_frame, text="编辑值:").pack(anchor='w', pady=(0, 5))

                    edit_var = tk.StringVar(value=value)
                    entry = ttk.Entry(dialog_frame, width=40, textvariable=edit_var)
                    entry.pack(fill='x', pady=(0, 15))
                    entry.select_range(0, tk.END)
                    entry.focus_set()

                    btn_frame = ttk.Frame(dialog_frame)
                    btn_frame.pack(fill='x')

                    def save_edit():
                        new_value = edit_var.get().strip()
                        if new_value and new_value not in listbox.get(0, tk.END):
                            listbox.delete(index)
                            listbox.insert(index, new_value)
                            listbox.selection_set(index)
                            edit_dialog.destroy()
                        elif new_value in listbox.get(0, tk.END) and new_value != value:
                            messagebox.showinfo("提示", f"'{new_value}'已存在，不能重复添加", parent=edit_dialog)
                        else:
                            edit_dialog.destroy()

                    ttk.Button(btn_frame, text="保存", command=save_edit).pack(side='right', padx=5)
                    ttk.Button(btn_frame, text="取消", command=edit_dialog.destroy).pack(side='right')

                    # 响应回车键
                    entry.bind('<Return>', lambda e: save_edit())

                    # 居中显示对话框
                    edit_dialog.update_idletasks()
                    x = settings_window.winfo_rootx() + (settings_window.winfo_width() // 2) - (edit_dialog.winfo_width() // 2)
                    y = settings_window.winfo_rooty() + (settings_window.winfo_height() // 2) - (edit_dialog.winfo_height() // 2)
                    edit_dialog.geometry(f"+{x}+{y}")

            listbox.bind('<Double-1>', on_double_click)

        # 底部按钮区域
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill='x', pady=(15, 0))

        def save_mappings():
            # 保存所有列表框中的内容到映射字典
            for field, listbox in self.mapping_listboxes.items():
                values = list(listbox.get(0, tk.END))
                if values:
                    self.bom_header_mapping[field] = values

            # 保存到配置文件
            self.save_header_mapping()
            settings_window.destroy()
            messagebox.showinfo("保存成功", "BOM表头映射配置已保存！", parent=self.master)

        def reset_defaults():
            if messagebox.askyesno("重置确认", "确定要重置为默认映射吗？这将覆盖您的自定义设置。",
                                parent=settings_window):
                # 重新初始化默认映射
                self.__init_default_mapping()

                # 更新列表框内容
                for field, listbox in self.mapping_listboxes.items():
                    listbox.delete(0, tk.END)
                    if field in self.bom_header_mapping:
                        for item in self.bom_header_mapping[field]:
                            listbox.insert(tk.END, item)

        # 帮助文本
        help_text = "提示：双击列表项可以编辑，右键点击可显示更多操作菜单"
        ttk.Label(button_frame, text=help_text, style="Help.TLabel").pack(side='left')

        # 操作按钮
        ttk.Button(button_frame, text="保存设置", command=save_mappings,
                  style="Accent.TButton" if hasattr(style, 'map') and 'Accent.TButton' in style.theme_names() else '').pack(side='right', padx=5)
        ttk.Button(button_frame, text="重置默认", command=reset_defaults).pack(side='right', padx=5)
        ttk.Button(button_frame, text="取消", command=settings_window.destroy).pack(side='right', padx=5)

        # 使用特殊方法确保窗口居中显示
        settings_window.update_idletasks()
        self.center_window(settings_window, 800, 600)
        settings_window.deiconify()  # 准备好后再显示窗口

    def _update_color_preview(self):
        """更新颜色预览"""
        selected_color = self.color_var.get()
        rgb_color = self.color_choices[selected_color]

        # 将 0-1 的RGB值转换为十六进制颜色
        hex_color = "#{:02x}{:02x}{:02x}".format(
            int(rgb_color[0] * 255),
            int(rgb_color[1] * 255),
            int(rgb_color[2] * 255)
        )

        self.color_preview.delete("all")
        self.color_preview.create_rectangle(0, 0, 100, 30, fill=hex_color, outline="")

    def on_closing(self):
        """处理窗口关闭事件"""
        # 先保存当前设置
        try:
            self.save_settings()
            print("程序关闭前已保存设置")
        except Exception as e:
            print(f"关闭时保存设置失败: {str(e)}")

        if self.processing:
            if messagebox.askyesno("确认退出", "处理正在进行中，确定要退出吗？"):
                # 设置停止标志
                self.stop_ui_update = True

                # 等待消息队列处理完成
                try:
                    self.message_queue.join(timeout=1.0)
                except:
                    pass

                # 强制执行垃圾回收，确保资源释放
                gc.collect()

                # 销毁窗口
                self.master.destroy()
        else:
            # 设置停止标志并销毁窗口
            self.stop_ui_update = True

            # 执行最终清理
            try:
                # 清空消息队列
                while not self.message_queue.empty():
                    try:
                        self.message_queue.get_nowait()
                        self.message_queue.task_done()
                    except:
                        break
                # 强制垃圾回收
                gc.collect()
            except:
                pass

            self.master.destroy()

    def _update_status(self, message, color=None):
        """更新状态栏信息"""
        if color is None:
            color = self.text_color
        self.status_label.config(text=message, foreground=color)

    def _detect_header_row(self, excel_path):
        """
        检测Excel文件中真正的表头行，跳过项目信息行
        返回表头行的索引（从0开始计数）
        """
        try:
            # 读取Excel文件的前15行进行分析
            max_rows_to_check = 15
            file_ext = os.path.splitext(excel_path)[1].lower()

            try:
                if file_ext == '.csv':
                    # 对于CSV文件，尝试不同编码，强制使用字符串类型
                    for encoding in ['utf-8', 'gbk', 'gb18030', 'latin1']:
                        try:
                            df_preview = pd.read_csv(excel_path, encoding=encoding, header=None, nrows=max_rows_to_check, dtype=str)
                            break
                        except:
                            continue
                else:
                    # 对于Excel文件，试用不同的引擎，强制使用字符串类型
                    try:
                        df_preview = pd.read_excel(excel_path, engine='openpyxl', header=None, nrows=max_rows_to_check, dtype=str)
                    except:
                        try:
                            df_preview = pd.read_excel(excel_path, engine='xlrd', header=None, nrows=max_rows_to_check, dtype=str)
                        except:
                            # 最后尝试默认引擎，强制使用字符串类型
                            df_preview = pd.read_excel(excel_path, header=None, nrows=max_rows_to_check, dtype=str)
            except Exception as e:
                print(f"读取预览数据失败: {str(e)}")
                return 0  # 出错时使用第一行作为表头

            # 获取更全面的表头关键词
            header_keywords = []

            # 添加位号相关关键词
            ref_keywords = ['reference', 'ref', 'refdes', 'designator', 'location', 'position', '位号', '元件位号']
            header_keywords.extend(ref_keywords)

            # 添加料号相关关键词
            pn_keywords = ['pn', 'part number', 'partnumber', 'p/n', 'material', '料号', '型号']
            header_keywords.extend(pn_keywords)

            # 添加数量相关关键词
            qty_keywords = ['quantity', 'qty', 'count', 'amount', '数量', '用量']
            header_keywords.extend(qty_keywords)

            # 添加更多通用表头关键词
            general_keywords = ['item', 'description', 'value', 'footprint', 'package', 'manufacturer', 'vendor',
                               '项目', '描述', '封装', '生产商', '供应商', '备注']
            header_keywords.extend(general_keywords)

            # 移除重复关键词并转换为小写
            header_keywords = list(set([kw.lower() for kw in header_keywords]))

            # 统计每行包含关键词的数量
            row_scores = []
            for row_idx in range(min(max_rows_to_check, len(df_preview))):
                row = df_preview.iloc[row_idx]

                # 计算该行非空单元格数量
                non_empty_cells = row.notna().sum()

                # 计算该行包含关键词的单元格数量
                keyword_matches = 0
                for cell in row:
                    if pd.notna(cell):
                        cell_str = str(cell).strip().lower()
                        if any(keyword.lower() in cell_str for keyword in header_keywords):
                            keyword_matches += 1
                        # 精确匹配加分
                        if cell_str in header_keywords:
                            keyword_matches += 1

                # 计算行得分：关键词匹配数 * 2 + 非空单元格比例
                if non_empty_cells > 0:
                    # 行的得分 = 关键词匹配数 * 2 + 非空单元格数量/总列数
                    score = (keyword_matches * 2) + (non_empty_cells / len(row))
                    row_scores.append((row_idx, score, keyword_matches, non_empty_cells))
                else:
                    row_scores.append((row_idx, 0, 0, 0))

            # 按得分排序
            row_scores.sort(key=lambda x: x[1], reverse=True)

            # 打印调试信息
            print(f"表头行得分排名（前3名）:")
            for i, (row_idx, score, matches, cells) in enumerate(row_scores[:3]):
                row_preview = ' | '.join([str(x) for x in df_preview.iloc[row_idx].values[:5]])
                print(f"行 {row_idx+1}: 得分={score:.2f}, 关键词匹配={matches}, 非空单元格={cells}, 内容: {row_preview}...")

            # 如果得分最高的行有明显高于其他行的得分，则认为是表头行
            if row_scores and row_scores[0][1] > 0:
                best_row, best_score, best_matches, best_non_empty = row_scores[0]

                # 如果该行有较多非空单元格且匹配关键词
                if best_non_empty >= 3 and best_matches > 0:
                    print(f"选择第 {best_row+1} 行作为表头行: 得分={best_score:.2f}")
                    return best_row

                # 如果关键词匹配数超过3个，也认为是表头行
                if best_matches >= 3:
                    print(f"选择第 {best_row+1} 行作为表头行（基于关键词匹配）: 匹配数={best_matches}")
                    return best_row

                # 如果非空单元格数量大于5且有关键词匹配，可能是表头行
                if best_non_empty >= 5 and best_matches > 0:
                    print(f"选择第 {best_row+1} 行作为表头行（基于非空单元格）: 非空单元格={best_non_empty}")
                    return best_row

            # 备选策略：查找第一个包含"Reference"或"PN"关键词的行
            key_header_terms = ["reference", "ref", "pn", "partnumber", "part number", "位号", "料号"]
            for row_idx in range(min(max_rows_to_check, len(df_preview))):
                row = df_preview.iloc[row_idx]
                for cell in row:
                    if pd.notna(cell):
                        cell_str = str(cell).strip().lower()
                        if any(term == cell_str or term in cell_str for term in key_header_terms):
                            print(f"备选策略: 选择第 {row_idx+1} 行作为表头行（发现关键表头词）: {cell_str}")
                            return row_idx

            # 如果无法确定，默认返回0（第一行）
            print("无法确定表头行，使用第1行作为默认表头行")
            return 0
        except Exception as e:
            print(f"检测表头行时出错: {str(e)}")
            return 0  # 出错时默认使用第一行

    def ask_user_for_column(self, column_names, column_type='reference'):
        """
        弹出对话框，让用户选择表头列并选择是否记住选择

        Args:
            column_names: 列名列表
            column_type: 列类型，可以是 'reference'(位号列), 'pn'(料号列),
                        'description'(描述列), 'manufacturer'(制造商列), 'mfr_pn'(厂商料号列)等
        """
        # 根据列类型设置窗口标题和提示文本
        titles = {
            'reference': "选择位号列",
            'pn': "选择料号列",
            'description': "选择描述列",
            'manufacturer': "选择制造商列",
            'mfr_pn': "选择厂商料号列",
            'part': "选择元件类型列",
            'quantity': "选择数量列"
        }

        prompts = {
            'reference': "未能自动识别位号列，请手动选择:",
            'pn': "未能自动识别料号列，请手动选择:",
            'description': "未能自动识别描述列，请手动选择:",
            'manufacturer': "未能自动识别制造商列，请手动选择:",
            'mfr_pn': "未能自动识别厂商料号列，请手动选择:",
            'part': "未能自动识别元件类型列，请手动选择:",
            'quantity': "未能自动识别数量列，请手动选择:"
        }

        hints = {
            'reference': "(位号列通常包含如 R1, C2, U3 等器件编号)",
            'pn': "(料号列通常包含物料的唯一识别编号)",
            'description': "(描述列通常包含物料的详细描述信息)",
            'manufacturer': "(制造商列通常包含物料的生产厂商名称)",
            'mfr_pn': "(厂商料号列通常包含制造商的零件编号)",
            'part': "(元件类型列通常包含如电阻、电容等类型信息)",
            'quantity': "(数量列通常包含物料的使用数量)"
        }

        # 获取对应的标题和提示文本，如果没有则使用默认值
        title = titles.get(column_type, f"选择{column_type}列")
        prompt = prompts.get(column_type, f"未能自动识别{column_type}列，请手动选择:")
        hint = hints.get(column_type, "")

        # 创建对话框窗口但不立即显示
        root = tk.Toplevel(self.master)
        root.title(title)
        root.withdraw()  # 先隐藏窗口
        root.transient(self.master)
        root.grab_set()

        frame = ttk.Frame(root, padding=10)
        frame.pack(fill='both', expand=True)

        ttk.Label(frame, text=prompt,
                font=('微软雅黑', 10, 'bold')).pack(pady=(0, 5))

        if hint:
            ttk.Label(frame, text=hint,
                    font=('微软雅黑', 9)).pack(pady=(0, 10))

        # 创建列表框和滚动条
        list_frame = ttk.Frame(frame)
        list_frame.pack(fill='both', expand=True, pady=5)

        listbox = tk.Listbox(list_frame, font=('微软雅黑', 9), selectmode=tk.SINGLE)
        listbox.pack(side='left', fill='both', expand=True)

        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
        scrollbar.pack(side='right', fill='y')
        listbox.config(yscrollcommand=scrollbar.set)

        # 在列表中添加列名
        for col in column_names:
            listbox.insert(tk.END, col)

        # 添加双击选择功能
        def on_double_click(event):
            on_ok()
        listbox.bind('<Double-1>', on_double_click)

        # 添加"记住我的选择"复选框
        remember_var = tk.BooleanVar(value=True)
        remember_check = ttk.Checkbutton(frame, text="记住我的选择", variable=remember_var)
        remember_check.pack(anchor='w', pady=(10, 0))

        # 结果变量
        result = [None, False]  # [选择的列名, 是否记住]

        # 添加按钮
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x', pady=(10, 0))

        def on_ok():
            selected_idx = listbox.curselection()
            if selected_idx:
                result[0] = column_names[selected_idx[0]]
                result[1] = remember_var.get()
                root.destroy()
            else:
                messagebox.showinfo("提示", "请选择一个列名", parent=root)

        def on_cancel():
            root.destroy()

        ttk.Button(btn_frame, text="确定", command=on_ok).pack(side='right', padx=5)
        ttk.Button(btn_frame, text="取消", command=on_cancel).pack(side='right', padx=5)

        # 使用特殊方法确保窗口居中显示
        self.center_window(root, 450, 350)
        root.deiconify()  # 显示窗口

        # 使窗口模态
        root.focus_set()
        root.wait_window(root)
        return result[0], result[1]

    def load_settings(self):
        """从统一配置文件加载用户设置，增强可靠性"""
        try:
            # 使用统一配置文件加载
            config = self._load_config()

            # 加载高亮颜色设置
            if 'highlight_color' in config:
                color_name = config['highlight_color']
                if color_name in self.color_choices:
                    self.color_var.set(color_name)
                    self.highlight_color = self.color_choices[color_name]
                    print(f"已加载颜色设置: {color_name}")

            # 加载注解设置
            if 'add_annotation' in config:
                self.add_annotation_var.set(config['add_annotation'])
                print(f"已加载注解设置: {self.add_annotation_var.get()}")

            # 加载保存格式设置
            if 'last_save_format' in config and hasattr(self, 'save_as_excel_var'):
                save_format = config['last_save_format']
                # 设置保存格式选项
                self.save_as_excel_var.set(save_format == 'xlsx')
                print(f"已加载保存格式设置: {save_format}")

            print("已从统一配置文件加载用户设置")

            # 尝试加载备用位置的配置文件
            config_file = self._get_config_path()
            if not os.path.exists(config_file):
                # 尝试从当前目录加载
                current_config = os.path.join(os.getcwd(), "bom_checker_config.json")
                if os.path.exists(current_config):
                    try:
                        with open(current_config, 'r', encoding='utf-8') as f:
                            backup_config = json.load(f)
                            # 应用配置
                            if 'highlight_color' in backup_config:
                                color_name = backup_config['highlight_color']
                                if color_name in self.color_choices:
                                    self.color_var.set(color_name)
                                    self.highlight_color = self.color_choices[color_name]
                            if 'add_annotation' in backup_config:
                                self.add_annotation_var.set(backup_config['add_annotation'])
                            if 'last_save_format' in backup_config and hasattr(self, 'save_as_excel_var'):
                                save_format = backup_config['last_save_format']
                                self.save_as_excel_var.set(save_format == 'xlsx')
                            print(f"已从备用位置加载配置: {current_config}")
                            # 保存到标准位置
                            self.save_settings()
                    except Exception as e:
                        print(f"加载备用配置失败: {str(e)}")

                # 尝试从用户目录加载
                user_config = os.path.join(os.path.expanduser("~"), "bom_checker_config.json")
                if os.path.exists(user_config):
                    try:
                        with open(user_config, 'r', encoding='utf-8') as f:
                            user_backup_config = json.load(f)
                            # 应用配置
                            if 'highlight_color' in user_backup_config:
                                color_name = user_backup_config['highlight_color']
                                if color_name in self.color_choices:
                                    self.color_var.set(color_name)
                                    self.highlight_color = self.color_choices[color_name]
                            if 'add_annotation' in user_backup_config:
                                self.add_annotation_var.set(user_backup_config['add_annotation'])
                            if 'last_save_format' in user_backup_config and hasattr(self, 'save_as_excel_var'):
                                save_format = user_backup_config['last_save_format']
                                self.save_as_excel_var.set(save_format == 'xlsx')
                            print(f"已从用户目录加载配置: {user_config}")
                            # 保存到标准位置
                            self.save_settings()
                    except Exception as e:
                        print(f"加载用户目录配置失败: {str(e)}")

            # 检查旧的配置文件并提示
            old_settings_file = "user_settings.json"
            old_mapping_file = "bom_header_mapping.json"

            if os.path.exists(old_settings_file) or os.path.exists(old_mapping_file):
                print("发现旧的配置文件，程序将使用新的统一配置文件")

        except Exception as e:
            print(f"加载设置出错: {str(e)}")
            # 出错时仍然尝试保存当前默认设置
            try:
                self.save_settings()
                print("已保存默认设置")
            except:
                pass

    def save_settings(self):
        """保存用户设置到统一配置文件"""
        try:
            # 先读取现有配置
            config = self._load_config()

            # 更新配置
            config.update({
                'highlight_color': self.color_var.get(),
                'add_annotation': self.add_annotation_var.get(),
                'last_save_format': 'xlsx' if self.save_as_excel_var.get() else 'txt'  # 保存当前选择的保存格式
            })

            # 保存配置
            self._save_config(config)

            print("已保存用户设置到统一配置文件")

            # 尝试删除旧的配置文件
            try:
                old_files = ["user_settings.json", "bom_header_mapping.json"]
                for old_file in old_files:
                    if os.path.exists(old_file):
                        os.remove(old_file)
                        print(f"已删除旧配置文件: {old_file}")
            except Exception as e:
                print(f"删除旧配置文件出错: {str(e)}")

        except Exception as e:
            print(f"保存设置出错: {str(e)}")

    def process_pdf(self, ref_designators):
        """
        使用多进程并行处理PDF页面，提高搜索和标注效率，并防止UI卡顿

        优化特性：
        - 自动根据CPU核心数计算最佳线程数(核心数×1.5)，在性能与资源消耗间取得平衡
        - 使用多处理进程隔离PDF处理，避免UI卡顿
        - 实时更新处理进度
        """
        # 确保ref_designators不为None且是列表
        if not ref_designators or not isinstance(ref_designators, list):
            raise ValueError("参数 ref_designators 必须是非空列表")

        # 创建进程间通信队列
        result_queue = Queue()
        progress_queue = Queue()

        # 自动计算最佳线程数
        cpu_cores = os.cpu_count() or 4
        num_threads = min(32, max(4, int(cpu_cores * 1.5)))

        # 确保有bom_header_mapping
        header_mapping = getattr(self, 'bom_header_mapping', None)

        # 创建新进程处理PDF，避免UI卡顿
        pdf_process = Process(
            target=pdf_processing_worker,
            args=(
                self.pdf_path,                # PDF文件路径
                ref_designators,              # 位号列表
                num_threads,                  # 线程数
                self.highlight_color,         # 高亮颜色
                result_queue,                 # 结果队列
                progress_queue,               # 进度队列
                self.add_annotation_var.get(),# 是否添加注解
                self.bom_path,                # BOM文件路径
                header_mapping                # 表头映射
            ),
            daemon=True,  # 设置为后台进程，主进程退出时自动终止
            name="PDF_Processor"  # 设置进程名称，便于调试
        )

        # 开始处理
        self.message_queue.put(("status", "在后台进程中处理PDF...", self.primary_color))
        pdf_process.start()

        # 监控处理进度和结果
        last_progress = 0
        result = None

        # 每100ms检查一次进度更新，避免阻塞UI线程
        check_interval = 0.1  # 秒
        start_time = time.time()

        while pdf_process.is_alive() or not progress_queue.empty() or not result_queue.empty():
            # 处理所有待处理的进度消息
            while not progress_queue.empty():
                try:
                    msg = progress_queue.get_nowait()
                    msg_type = msg[0]

                    if msg_type == "progress":
                        progress_value = msg[1]
                        if progress_value > last_progress:
                            last_progress = progress_value
                            self.message_queue.put(("progress", progress_value))
                    elif msg_type == "warning":
                        self.message_queue.put(("text", f"警告: {msg[1]}\n", False))
                except:
                    break

            # 检查是否有结果数据
            if result is None and not result_queue.empty():
                try:
                    result = result_queue.get_nowait()
                except:
                    pass

            # 如果进程仍在运行，等待一小段时间再检查
            if pdf_process.is_alive():
                time.sleep(check_interval)
            else:
                # 进程已结束但仍未获取结果，再等待一小段时间
                if result is None:
                    time.sleep(0.05)
                else:
                    break

        # 确保进程已终止
        if pdf_process.is_alive():
            print("PDF处理进程仍在运行，尝试终止...")
            pdf_process.terminate()
            pdf_process.join(timeout=1.0)

        # 如果没有获取到结果，创建一个错误结果
        if result is None:
            result = {
                'error': "未能从PDF处理进程获取结果",
                'total': len(ref_designators),
                'found': 0,
                'not_found': ref_designators,
                'output_pdf': None
            }

        return result

    def _update_progress(self, value):
        """更新进度条，确保在主线程中执行。"""
        if hasattr(self, 'progress_var'):
            self.progress_var.set(value)
            self.progress_label.config(text=f"{int(value)}%")
            self.master.update_idletasks()

    def center_window(self, window, width=None, height=None):
        """将窗口居中显示在屏幕上，避免闪烁和抖动"""
        # 若窗口已可见，先隐藏
        was_visible = window.winfo_viewable()
        if was_visible:
            window.withdraw()

        # 如果没有指定宽高，使用窗口当前大小
        if width is None:
            width = max(window.winfo_reqwidth(), 100)
        if height is None:
            height = max(window.winfo_reqheight(), 100)

        # 确保窗口大小合理
        if width < 100: width = 450
        if height < 100: height = 480

        # 计算居中位置
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)

        # 一次性设置窗口位置和大小
        window.geometry(f"{width}x{height}+{x}+{y}")

        # 确保几何设置已应用
        window.update_idletasks()

        # 如果之前窗口是可见的，重新显示
        if was_visible:
            window.deiconify()

if __name__ == "__main__":
    # 在Windows系统上启用多进程支持
    freeze_support()

    # 在Windows系统上，确保设置正确的多进程启动方法
    if sys.platform == 'win32':
        # 设置多进程启动方法为spawn，这在Windows上是默认值，但明确设置可以避免打包问题
        try:
            mp.set_start_method('spawn')
        except RuntimeError:
            # 可能已经设置了启动方法
            print("多进程启动方法已设置")

    root = tk.Tk()
    app = BOMCheckerApp(root)
    root.mainloop()











